import concurrent.futures
import datetime
import requests
import urllib3
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from collections import defaultdict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
matplotlib.rcParams["font.family"] = "DejaVu Sans"
matplotlib.rcParams["axes.unicode_minus"] = False

API_URL = "https://ows.goszakup.gov.kz/v3/graphql"
HEADERS = {
    "Authorization": "Bearer 8530636543c42453dc78b894f3c32ab6",
    "Content-Type": "application/json",
}

GRAPHQL_QUERY = """
query EmfReportBpr($after: Int, $date: String!) {
  EmfReportBpr(limit: 200, after: $after, filter: { dateReport: $date }) {
    grbId, strId, katFgr, katFgrName,
    klsFpgr, klsFpgrName, pklAbp, pklAbpName,
    adtype, utv, utch, plg, sumrg
  }
}
"""

DATE_2023 = "2024-01-01"
DATE_2024 = "2025-01-01"
DATE_2025 = "2026-01-01"


def _get_latest_month_candidates(n=4):
    """Generate n month-start dates going back from last month, skipping Jan-01 annual dates."""
    annual = {f"{y}-01-01" for y in range(2016, datetime.date.today().year + 3)}
    today = datetime.date.today()
    y, m = today.year, today.month - 1
    if m == 0:
        m, y = 12, y - 1
    candidates = []
    for _ in range(n):
        s = f"{y}-{m:02d}-01"
        if s not in annual:
            candidates.append(s)
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return candidates


# ─────────────────────────── API fetch ───────────────────────────

def _fetch_all(date_str: str) -> list:
    """Fetch all paginated rows for a given datereport."""
    import time
    rows = []
    after = None
    while True:
        payload = {
            "query": GRAPHQL_QUERY,
            "variables": {"after": after, "date": date_str},
        }
        for attempt in range(4):
            try:
                r = requests.post(API_URL, json=payload, headers=HEADERS,
                                  verify=False, timeout=120)
                r.raise_for_status()
                break
            except requests.exceptions.Timeout:
                if attempt == 3:
                    raise
                time.sleep(5 * (attempt + 1))
        data = r.json()
        batch = data.get("data", {}).get("EmfReportBpr", []) or []
        if not batch:
            break
        rows.extend(batch)
        page_info = (data.get("extensions") or {}).get("pageInfo") or {}
        last_id = page_info.get("lastId")
        if not page_info.get("hasNextPage") or not last_id:
            break
        after = last_id
    return rows


def _fetch_with_fallback(primary: str, fallback: str) -> list:
    rows = _fetch_all(primary)
    if not any(_is_tax(r) for r in rows):
        rows = _fetch_all(fallback)
    return rows


def _find_latest_month_tax() -> tuple[str, list]:
    """Return (label, rows) for the newest month that has tax data."""
    for candidate in _get_latest_month_candidates():
        rows = _fetch_all(candidate)
        if any(_is_tax(r) for r in rows):
            label = f"{candidate[5:7]}.{candidate[:4]} Исп."
            return label, rows
    return None, []


# ─────────────────────────── filtering ───────────────────────────

def _safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _is_tax(r: dict) -> bool:
    """Tax revenues: grbId=1, katFgr=1, strId=1."""
    return (
        _safe_int(r.get("grbId")) == 1 and
        _safe_int(r.get("katFgr")) == 1 and
        _safe_int(r.get("strId")) == 1
    )


# ─────────────────────────── aggregation ───────────────────────────

def _to_mln(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ".")) / 1_000_000
    except Exception:
        return 0.0


def _group_by_kls(rows: list) -> dict:
    """Aggregate by klsFpgrName → {kls: {sumrg, plg, utv, utch}}"""
    totals = defaultdict(lambda: defaultdict(float))
    for r in rows:
        if not _is_tax(r):
            continue
        kls = r.get("klsFpgrName") or ""
        totals[kls]["sumrg"] += _to_mln(r.get("sumrg"))
        totals[kls]["plg"]   += _to_mln(r.get("plg"))
        totals[kls]["utv"]   += _to_mln(r.get("utv"))
        totals[kls]["utch"]  += _to_mln(r.get("utch"))
    return {k: dict(v) for k, v in totals.items()}


def _pivot_pkl(rows: list) -> dict:
    """Pivot by (klsFpgrName, pklAbpName) → {sumrg, plg, utv, utch}"""
    data = defaultdict(lambda: defaultdict(float))
    for r in rows:
        if not _is_tax(r):
            continue
        key = (r.get("klsFpgrName") or "", r.get("pklAbpName") or "")
        data[key]["sumrg"] += _to_mln(r.get("sumrg"))
        data[key]["plg"]   += _to_mln(r.get("plg"))
        data[key]["utv"]   += _to_mln(r.get("utv"))
        data[key]["utch"]  += _to_mln(r.get("utch"))
    return {k: dict(v) for k, v in data.items()}


def _all_kls_ordered(rows_list: list) -> list:
    """Ordered unique klsFpgrName values from all datasets."""
    seen = []
    for rows in rows_list:
        for r in rows:
            if not _is_tax(r):
                continue
            kls = r.get("klsFpgrName") or ""
            if kls and kls not in seen:
                seen.append(kls)
    return seen


def _all_pkl_ordered(rows_list: list) -> list:
    """Ordered unique (kls, pkl) pairs from all datasets."""
    seen = []
    for rows in rows_list:
        for r in rows:
            if not _is_tax(r):
                continue
            key = (r.get("klsFpgrName") or "", r.get("pklAbpName") or "")
            if key not in seen:
                seen.append(key)
    return seen


# ─────────────────────────── Excel builder ───────────────────────────

def build_excel_taxes(filename: str) -> str:
    """
    Sheet 1 — tax revenues table (kls group + pkl detail × periods)
    Sheet 2 — bar chart comparing 2023 / 2024 / 2025 execution by tax type
    """
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        f2023 = executor.submit(_fetch_all, DATE_2023)
        f2024 = executor.submit(_fetch_all, DATE_2024)
        f2025 = executor.submit(_fetch_with_fallback, DATE_2025, "2025-12-01")
        f_month = executor.submit(_find_latest_month_tax)

    r2023 = f2023.result()
    r2024 = f2024.result()
    r2025 = f2025.result()
    month_label, r_month = f_month.result()

    all_rows = [r2023, r2024, r2025] + ([r_month] if r_month else [])
    g23, g24, g25 = _group_by_kls(r2023), _group_by_kls(r2024), _group_by_kls(r2025)
    d23, d24, d25 = _pivot_pkl(r2023), _pivot_pkl(r2024), _pivot_pkl(r2025)
    g_month = _group_by_kls(r_month) if r_month else {}
    d_month = _pivot_pkl(r_month) if r_month else {}

    all_kls  = _all_kls_ordered(all_rows)
    all_pkls = _all_pkl_ordered(all_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Налоги"

    col_labels = ["2023 Исп.", "2024 Исп.", "2025 Исп.", "2025 Утв.", "2025 Уточн.", "2025 Скорр."]
    if r_month:
        col_labels.append(month_label)

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_cols = 1 + len(col_labels)

    # ── Title ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = "НАЛОГОВЫЕ ПОСТУПЛЕНИЯ КАЗАХСТАНА"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"] = "(млн тенге)"
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Header ──
    ws.append(["Наименование"] + col_labels)
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 30

    def _val_row_kls(kls):
        def get(d): return d.get(kls, {})
        vals = [
            round(get(g23).get("sumrg", 0), 1),
            round(get(g24).get("sumrg", 0), 1),
            round(get(g25).get("sumrg", 0), 1),
            round(get(g25).get("utv",   0), 1),
            round(get(g25).get("utch",  0), 1),
            round(get(g25).get("plg",   0), 1),
        ]
        if r_month:
            vals.append(round(get(g_month).get("sumrg", 0), 1))
        return vals

    def _val_row_pkl(key):
        def get(d): return d.get(key, {})
        vals = [
            round(get(d23).get("sumrg", 0), 1),
            round(get(d24).get("sumrg", 0), 1),
            round(get(d25).get("sumrg", 0), 1),
            round(get(d25).get("utv",   0), 1),
            round(get(d25).get("utch",  0), 1),
            round(get(d25).get("plg",   0), 1),
        ]
        if r_month:
            vals.append(round(get(d_month).get("sumrg", 0), 1))
        return vals

    def _append_row(label, vals, font):
        ws.append([label] + vals)
        r = ws.max_row
        ws.cell(r, 1).font = font
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="center")
        for c in range(1, total_cols + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if c > 1:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.0"

    # ── Republic total ──
    total = [
        round(sum(v.get("sumrg", 0) for v in g23.values()), 1),
        round(sum(v.get("sumrg", 0) for v in g24.values()), 1),
        round(sum(v.get("sumrg", 0) for v in g25.values()), 1),
        round(sum(v.get("utv",   0) for v in g25.values()), 1),
        round(sum(v.get("utch",  0) for v in g25.values()), 1),
        round(sum(v.get("plg",   0) for v in g25.values()), 1),
    ]
    if r_month:
        total.append(round(sum(v.get("sumrg", 0) for v in g_month.values()), 1))
    _append_row("I. НАЛОГОВЫЕ ПОСТУПЛЕНИЯ", total, Font(bold=True, size=10))

    # ── By klsFpgrName (group) + pklAbpName (detail) ──
    for kls in all_kls:
        _append_row(kls, _val_row_kls(kls), Font(bold=True, size=9))
        for (kl, pk) in all_pkls:
            if kl != kls or not pk:
                continue
            _append_row(f"   {pk}", _val_row_pkl((kl, pk)), Font(size=8))

    # Column widths
    ws.column_dimensions["A"].width = 55
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ── Chart sheet ──
    ws_chart = wb.create_sheet("График")

    chart_kls = all_kls[:10]
    cv23 = [g23.get(k, {}).get("sumrg", 0) / 1_000 for k in chart_kls]
    cv24 = [g24.get(k, {}).get("sumrg", 0) / 1_000 for k in chart_kls]
    cv25 = [g25.get(k, {}).get("sumrg", 0) / 1_000 for k in chart_kls]

    if chart_kls:
        x = range(len(chart_kls))
        w = 0.25
        fig, ax = plt.subplots(figsize=(12, 6))
        ax.bar([i - w for i in x], cv23, w, label="2023 Исп.", color="#003366")
        ax.bar(list(x),            cv24, w, label="2024 Исп.", color="#0066CC")
        ax.bar([i + w for i in x], cv25, w, label="2025 Исп.", color="#00AAFF")

        short = [lb[:30] + "…" if len(lb) > 30 else lb for lb in chart_kls]
        ax.set_xticks(list(x))
        ax.set_xticklabels(short, rotation=30, ha="right", fontsize=8)
        ax.set_ylabel("млрд тенге")
        ax.set_title(
            "Налоговые поступления Казахстана по видам налогов",
            fontweight="bold", pad=15
        )
        ax.legend()
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        plt.figtext(0.01, -0.02, "По данным Министерства финансов РК", fontsize=8, color="gray")
        plt.figtext(0.85, -0.02, "TENGENOMIKA", fontsize=9, fontweight="bold", color="#003366")
        plt.tight_layout()

        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=200, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        ws_chart.add_image(Image(buf), "A1")

    wb.save(filename)
    return filename
