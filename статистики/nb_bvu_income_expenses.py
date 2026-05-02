import requests
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter, MonthLocator
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import matplotlib

# --- Настройки кириллицы ---
matplotlib.rcParams["font.family"] = "DejaVu Sans"
matplotlib.rcParams["axes.unicode_minus"] = False

API_URL = (
    "https://data.nationalbank.kz/api/report/stat-data-locale"
    "?form_id=3&date_from=2019-01-01&date_to=2050-01-01&lang=ru"
)

MONTHS_RU = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
}


# ---------------------------------------------------------------------
# Загрузка данных
# ---------------------------------------------------------------------
def load_bvu_income_expenses() -> pd.DataFrame:
    """
    Загружает сводный отчет о доходах и расходах БВУ.
    Возвращает DataFrame:
    Дата | Код | Наименование | Сумма_тыс_тенге
    """
    r = requests.get(API_URL, timeout=(10, 60))
    r.raise_for_status()
    data = r.json()

    rows = []
    for item in data:
        report_date = item.get("reportDate")
        amount = item.get("amount")

        if not report_date or amount is None:
            continue

        attrs = item.get("attributes", []) or []

        code = None
        name = None
        for a in attrs:
            if a.get("attrCode") == "account_group":
                code = a.get("valueCode")
                name = a.get("valueNameRu")
                break

        if not code:
            continue

        rows.append(
            {
                "Дата": pd.to_datetime(report_date, errors="coerce"),
                "Код": str(code),
                "Наименование": name or "",
                "Сумма_тыс_тенге": float(amount),
            }
        )

    df = pd.DataFrame(rows)
    df = df.dropna(subset=["Дата"])
    return df


# ---------------------------------------------------------------------
# Excel + график
# ---------------------------------------------------------------------
def build_excel_bvu_income_expenses(filename: str, year: int) -> str:
    """
    Excel:
    Лист 1 — таблица (группы счетов × месяцы) за выбранный год
    Лист 2 — график (итого доходы/расходы по месяцам) за выбранный год
    """

    df = load_bvu_income_expenses()
    if df.empty:
        raise ValueError("Нет данных из API")

    # ✅ Фильтр строго по выбранному году
    df = df[df["Дата"].dt.year == year].copy()
    if df.empty:
        raise ValueError(f"Нет данных за {year} год")

    # Месяц (reportDate обычно 01.MM)
    df["Месяц"] = df["Дата"].dt.to_period("M").dt.to_timestamp()

    # На случай дублей
    grouped = (
        df.groupby(["Код", "Наименование", "Месяц"], as_index=False)["Сумма_тыс_тенге"]
          .sum()
    )

    pivot = grouped.pivot_table(
        index=["Код", "Наименование"],
        columns="Месяц",
        values="Сумма_тыс_тенге",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()

    # Сортировка по коду (если код не числовой — сортируем как строку)
    def _safe_int(x: str):
        try:
            return int(x)
        except Exception:
            return 10**18

    pivot = pivot.sort_values(by="Код", key=lambda s: s.map(_safe_int))

    month_cols = sorted([c for c in pivot.columns if isinstance(c, pd.Timestamp)])

    # ================= Excel =================
    wb = Workbook()
    ws = wb.active
    ws.title = "Доходы и расходы БВУ"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(month_cols))
    ws["A1"] = f"СВОДНЫЙ ОТЧЕТ О ДОХОДАХ И РАСХОДАХ БВУ РК ({year} г.)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    header = ["Код", "Наименование доходов / расходов"]
    header += [f"на {m.strftime('%d.%m.%Y')} г." for m in month_cols]
    ws.append(header)

    for _, row in pivot.iterrows():
        ws.append([row["Код"], row["Наименование"]] + [round(float(row[m]), 1) for m in month_cols])

    # ИТОГО
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)

    for j in range(3, 3 + len(month_cols)):
        col = get_column_letter(j)
        ws.cell(row=total_row, column=j, value=f"=SUM({col}3:{col}{total_row-1})").font = Font(bold=True)

    # Оформление
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    for j in range(3, 3 + len(month_cols)):
        ws.column_dimensions[get_column_letter(j)].width = 18

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2 + len(month_cols)):
        for c in row:
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in ws[2]:
        c.font = Font(bold=True)

    for r in range(3, ws.max_row + 1):
        for j in range(3, 3 + len(month_cols)):
            ws.cell(row=r, column=j).number_format = "#,##0"

    # ================= График =================
    ws_chart = wb.create_sheet("График")

    total_series = (
        grouped.groupby("Месяц", as_index=False)["Сумма_тыс_тенге"]
               .sum()
               .sort_values("Месяц")
    )

    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(
        total_series["Месяц"],
        total_series["Сумма_тыс_тенге"],
        linewidth=2.5,
        marker="o",
        label="Итого доходы и расходы",
    )

    ax.set_title(
        f"Доходы и расходы БВУ РК — ИТОГО ({year} г.)",
        fontsize=12,
        fontweight="bold",
        pad=18,
    )
    ax.set_ylabel("тыс. тенге")
    ax.grid(True, linestyle="--", alpha=0.5)

    ax.xaxis.set_major_locator(MonthLocator(interval=1))
    ax.xaxis.set_major_formatter(DateFormatter("%b.%y"))
    plt.xticks(rotation=45, ha="right")

    ax.legend()
    plt.figtext(0.01, -0.02, "Источник: НБ РК", fontsize=9, color="gray")
    plt.figtext(0.87, -0.02, "TENGENOMIKA", fontsize=9, fontweight="bold")

    plt.tight_layout()

    img_data = BytesIO()
    plt.savefig(img_data, format="png", dpi=200, bbox_inches="tight")
    img_data.seek(0)
    plt.close()

    img = Image(img_data)
    ws_chart.add_image(img, "A1")

    wb.save(filename)
    return filename
