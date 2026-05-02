from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Абайская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "1000%", "sheet_name": "Абайская область"},
    {"kato_pattern": "1010%", "sheet_name": "Семей Г.А."},
    {"kato_pattern": "1018%", "sheet_name": "Курчатов Г.А."},
    {"kato_pattern": "1032%", "sheet_name": "Абайский район"},
    {"kato_pattern": "1034%", "sheet_name": "район Ақсуат"},
    {"kato_pattern": "1036%", "sheet_name": "Аягозский район"},
    {"kato_pattern": "1038%", "sheet_name": "Бескарагайский район"},
    {"kato_pattern": "1040%", "sheet_name": "Бородулихинский район"},
    {"kato_pattern": "1042%", "sheet_name": "Бородулихинский район"},
    {"kato_pattern": "1044%", "sheet_name": "Кокпектинский район"},
    {"kato_pattern": "1046%", "sheet_name": "Урджарский район"},
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

