from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Кызылординская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "4300%", "sheet_name": "Кызылординская область"},
    {"kato_pattern": "4310%", "sheet_name": "Кызылорда"},
    {"kato_pattern": "4319%", "sheet_name": "Байконыр"},
    {"kato_pattern": "4332%", "sheet_name": "Аральский район"},
    {"kato_pattern": "4336%", "sheet_name": "Жалагашский район"},
    {"kato_pattern": "4340%", "sheet_name": "Жанакорганский район"},
    {"kato_pattern": "4344%", "sheet_name": "Казалинский район"},
    {"kato_pattern": "4346%", "sheet_name": "Кармакшинский район"},
    {"kato_pattern": "4348%", "sheet_name": "Сырдарьинский район"},
    {"kato_pattern": "4352%", "sheet_name": "Чиилийский район"}
]



first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

