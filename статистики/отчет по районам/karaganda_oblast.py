from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Карагандинская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "3500%", "sheet_name": "Карагандинская область"},
    {"kato_pattern": "3510%", "sheet_name": "Караганда"},
    {"kato_pattern": "3516%", "sheet_name": "Балхаш"},
    {"kato_pattern": "3518%", "sheet_name": "Жезказган"},
    {"kato_pattern": "3520%", "sheet_name": "Каражал"},
    {"kato_pattern": "3521%", "sheet_name": "Приозерск"},
    {"kato_pattern": "3522%", "sheet_name": "Сарань"},
    {"kato_pattern": "3523%", "sheet_name": "Сатпаев"},
    {"kato_pattern": "3524%", "sheet_name": "Темиртау"},
    {"kato_pattern": "3528%", "sheet_name": "Шахтинск"},
    {"kato_pattern": "3532%", "sheet_name": "Абайский район"},
    {"kato_pattern": "3536%", "sheet_name": "Актогайский район"},
    {"kato_pattern": "3540%", "sheet_name": "Бухар-Жырауский район"},
    {"kato_pattern": "3544%", "sheet_name": "Жанааркинский район"},
    {"kato_pattern": "3548%", "sheet_name": "Каркаралинский район"},
    {"kato_pattern": "3552%", "sheet_name": "Нуринский район"},
    {"kato_pattern": "3556%", "sheet_name": "Осакаровский район"},
    {"kato_pattern": "3560%", "sheet_name": "Улытауский район"},
    {"kato_pattern": "3564%", "sheet_name": "Шетский район"}
]



first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

