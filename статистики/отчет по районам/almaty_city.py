from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Алматы.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "750000000%", "sheet_name": "Г. Алматы"},
    {"kato_pattern": "75111%", "sheet_name": "район Алмалинский"},
    {"kato_pattern": "75121%", "sheet_name": "район Алатауский"},
    {"kato_pattern": "75131%", "sheet_name": "район Ауэзовский"},
    {"kato_pattern": "75141%", "sheet_name": "район Бостандыкский"},
    {"kato_pattern": "75151%", "sheet_name": "район Жетысуский"},
    {"kato_pattern": "75171%", "sheet_name": "район Медеуский"},
    {"kato_pattern": "75181%", "sheet_name": "район Наурызбайский"},
    {"kato_pattern": "75191%", "sheet_name": "район Турксибский"},
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

