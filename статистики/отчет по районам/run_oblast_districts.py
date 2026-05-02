from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'ERI',
    'user': 'postgres',
    'password': '20041015R',
    'host': 'localhost',
    'port': '5432'
}

output_file = "2025июль/Абайская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "1000%", "sheet_name": "Абайская область", "file_name": "всеобласти/Алматинская область.xlsx"},
    {"kato_pattern": "1100%", "sheet_name": "Акмолинская область", "file_name": "всеобласти/Акмолинская область.xlsx"},
    {"kato_pattern": "1500%", "sheet_name": "Актюбинская область", "file_name": "всеобласти/Актюбинская область.xlsx"},
    {"kato_pattern": "1900%", "sheet_name": "Алматинская область", "file_name": "всеобласти/Алматинская область.xlsx"},
    {"kato_pattern": "2300%", "sheet_name": "Атырауская область", "file_name": "всеобласти/Атырауская область.xlsx"},
    {"kato_pattern": "2700%", "sheet_name": "Западно-Казахстанская область", "file_name": "всеобласти/Западно-Казахстанская область.xlsx"},
    {"kato_pattern": "3100%", "sheet_name": "Жамбылская область", "file_name": "всеобласти/Жамбылская область.xlsx"},
    {"kato_pattern": "3300%", "sheet_name": "область Жетісу", "file_name": "всеобласти/область Жетісу.xlsx"},
    {"kato_pattern": "3500%", "sheet_name": "Карагандинская область", "file_name": "всеобласти/Карагандинская область.xlsx"},
    {"kato_pattern": "3900%", "sheet_name": "Костанайская область", "file_name": "всеобласти/Костанайская область.xlsx"},
    {"kato_pattern": "4300%", "sheet_name": "Кызылординская область", "file_name": "всеобласти/Кызылординская область.xlsx"},
    {"kato_pattern": "4700%", "sheet_name": "Мангистауская область", "file_name": "всеобласти/Мангистауская область.xlsx"},
    {"kato_pattern": "5100%", "sheet_name": "Южно-Казахстанская область", "file_name": "всеобласти/Южно-Казахстанская область.xlsx"},
    {"kato_pattern": "5500%", "sheet_name": "Павлодарская область", "file_name": "всеобласти/Павлодарская область.xlsx"},
    {"kato_pattern": "5900%", "sheet_name": "Северо-Казахстанская область", "file_name": "всеобласти/Северо-Казахстанская область.xlsx"},
    {"kato_pattern": "6100%", "sheet_name": "Туркестанская область", "file_name": "всеобласти/Туркестанская область.xlsx"},
    {"kato_pattern": "6200%", "sheet_name": "область Ұлытау", "file_name": "всеобласти/область Ұлытау.xlsx"},
    {"kato_pattern": "6300%", "sheet_name": "Восточно-Казахстанская область", "file_name": "всеобласти/Восточно-Казахстанская область.xlsx"},
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")
