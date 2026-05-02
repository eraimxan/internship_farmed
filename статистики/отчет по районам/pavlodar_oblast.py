from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Павлодарская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "5500%", "sheet_name": "Павлодарская область"},
    {"kato_pattern": "5510%", "sheet_name": "Павлодар"},
    {"kato_pattern": "5516%", "sheet_name": "Аксу"},
    {"kato_pattern": "5522%", "sheet_name": "Экибастуз"},
    {"kato_pattern": "5532%", "sheet_name": "Актогайский район"},
    {"kato_pattern": "5536%", "sheet_name": "Баянаульский район"},
    {"kato_pattern": "5542%", "sheet_name": "Железинский район"},
    {"kato_pattern": "5546%", "sheet_name": "Иртышский район"},
    {"kato_pattern": "5548%", "sheet_name": "Тереңкөл район"},
    {"kato_pattern": "5552%", "sheet_name": "Аққулы район"},
    {"kato_pattern": "5556%", "sheet_name": "Майский район"},
    {"kato_pattern": "5560%", "sheet_name": "Павлодарский район"},
    {"kato_pattern": "5564%", "sheet_name": "Успенский район"},
    {"kato_pattern": "5568%", "sheet_name": "Щербактинский район"}
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

