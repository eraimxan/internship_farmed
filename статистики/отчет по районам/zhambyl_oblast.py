from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Жамбылская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "3100%", "sheet_name": "Жамбылская область"},
    {"kato_pattern": "3110%", "sheet_name": "Тараз"},
    {"kato_pattern": "3136%", "sheet_name": "Байзакский район"},
    {"kato_pattern": "3140%", "sheet_name": "Жамбылский район"},
    {"kato_pattern": "3142%", "sheet_name": "Жуалынский район"},
    {"kato_pattern": "3148%", "sheet_name": "Турара Рыскулова район"},
    {"kato_pattern": "3150%", "sheet_name": "Кордайский район"},
    {"kato_pattern": "3154%", "sheet_name": "Меркенский район"},
    {"kato_pattern": "3156%", "sheet_name": "Мойынкумский район"},
    {"kato_pattern": "3160%", "sheet_name": "Сарысуский район"},
    {"kato_pattern": "3162%", "sheet_name": "Таласский район"},
    {"kato_pattern": "3166%", "sheet_name": "Шуский район"}
]


first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

