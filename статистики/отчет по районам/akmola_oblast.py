from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Акмолинская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "1100%", "sheet_name": "Акмолинская область"},
    {"kato_pattern": "1110%", "sheet_name": "Кокшетау Г.А"},
    {"kato_pattern": "1116%", "sheet_name": "Косшы Г.А."},
    {"kato_pattern": "1118%", "sheet_name": "Степногорск Г.А."},
    {"kato_pattern": "1132%", "sheet_name": "Аккольский район"},
    {"kato_pattern": "1134%", "sheet_name": "Аршалынский район"},
    {"kato_pattern": "1136%", "sheet_name": "Астраханский район"},
    {"kato_pattern": "1138%", "sheet_name": "Атбасарский район"},
    {"kato_pattern": "1140%", "sheet_name": "Буландынский район"},
    {"kato_pattern": "1144%", "sheet_name": "Егиндыкольский район"},
    {"kato_pattern": "1145%", "sheet_name": "район Биржан сал"},   
    {"kato_pattern": "1146%", "sheet_name": "Ерейментауский район"},
    {"kato_pattern": "1148%", "sheet_name": "Есильский район"},
    {"kato_pattern": "1152%", "sheet_name": "Жаксынский район"},
    {"kato_pattern": "1154%", "sheet_name": "Жаркаинский район"},
    {"kato_pattern": "1156%", "sheet_name": "Зерендинский район"},
    {"kato_pattern": "1160%", "sheet_name": "Коргалжынский район"},
    {"kato_pattern": "1164%", "sheet_name": "Сандыктауский район"},
    {"kato_pattern": "1166%", "sheet_name": "Целиноградский район"},
    {"kato_pattern": "1168%", "sheet_name": "Шортандинский район"},
    {"kato_pattern": "1170%", "sheet_name": "Бурабайский район"},    
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

