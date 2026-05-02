from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Западно-Казахстанская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "2700%", "sheet_name": "Западно-Казахстанская область"},
    {"kato_pattern": "2710%", "sheet_name": "Уральск Г.А."},
    {"kato_pattern": "2732%", "sheet_name": "Акжаикский район"},
    {"kato_pattern": "2736%", "sheet_name": "Бурлинский район"},
    {"kato_pattern": "2740%", "sheet_name": "Жангалинский район"},
    {"kato_pattern": "2742%", "sheet_name": "Жанибекский район"},
    {"kato_pattern": "2744%", "sheet_name": "район Бәйтерек"},
    {"kato_pattern": "2748%", "sheet_name": "Казталовский район"},
    {"kato_pattern": "2750%", "sheet_name": "Каратобинский район"},
    {"kato_pattern": "2754%", "sheet_name": "Бокейординский район"},
    {"kato_pattern": "2758%", "sheet_name": "Сырымский район"},
    {"kato_pattern": "2760%", "sheet_name": "Таскалинский район"},
    {"kato_pattern": "2762%", "sheet_name": "Теректинский район"},
    {"kato_pattern": "2766%", "sheet_name": "Чингирлауский район"},
]


first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

