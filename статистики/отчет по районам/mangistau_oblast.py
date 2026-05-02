from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Мангистауская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "4700%", "sheet_name": "Мангистауская область"},
    {"kato_pattern": "4710%", "sheet_name": "Актау"},
    {"kato_pattern": "4718%", "sheet_name": "Жанаозен"},
    {"kato_pattern": "4736%", "sheet_name": "Бейнеуский район"},
    {"kato_pattern": "4742%", "sheet_name": "Каракиянский район"},
    {"kato_pattern": "4746%", "sheet_name": "Мангистауский район"},
    {"kato_pattern": "4750%", "sheet_name": "Мунайлинский район"},
    {"kato_pattern": "4752%", "sheet_name": "Тупкараганский район"}
]


first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

