from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Туркестанская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "6100%", "sheet_name": "Туркестанская область"},
    {"kato_pattern": "6110%", "sheet_name": "Туркестан"},
    {"kato_pattern": "6116%", "sheet_name": "Арысь"},
    {"kato_pattern": "6120%", "sheet_name": "Кентау"},
    {"kato_pattern": "6136%", "sheet_name": "Байдибека район"},
    {"kato_pattern": "6138%", "sheet_name": "Жетисайский район"},
    {"kato_pattern": "6139%", "sheet_name": "Келесский район"},
    {"kato_pattern": "6140%", "sheet_name": "Казыгуртский район"},
    {"kato_pattern": "6144%", "sheet_name": "Мактааральский район"},
    {"kato_pattern": "6146%", "sheet_name": "Ордабасынский район"},
    {"kato_pattern": "6148%", "sheet_name": "Отрарский район"},
    {"kato_pattern": "6152%", "sheet_name": "Сайрамский район"},
    {"kato_pattern": "6154%", "sheet_name": "Сарыагашский район"},
    {"kato_pattern": "6155%", "sheet_name": "Сауран район"},
    {"kato_pattern": "6156%", "sheet_name": "Сузакский район"},
    {"kato_pattern": "6158%", "sheet_name": "Толебийский район"},
    {"kato_pattern": "6160%", "sheet_name": "Тюлькубасский район"},
    {"kato_pattern": "6164%", "sheet_name": "Шардаринский район"},
]



first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

