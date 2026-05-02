from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Костанайская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "3900%", "sheet_name": "Костанайская область"},
    {"kato_pattern": "3910%", "sheet_name": "Костанай"},
    {"kato_pattern": "3916%", "sheet_name": "Аркалык"},
    {"kato_pattern": "3920%", "sheet_name": "Лисаковск"},
    {"kato_pattern": "3924%", "sheet_name": "Рудный"},
    {"kato_pattern": "3932%", "sheet_name": "Алтынсаринский район"},
    {"kato_pattern": "3934%", "sheet_name": "Амангельдинский район"},
    {"kato_pattern": "3936%", "sheet_name": "Аулиекольский район"},
    {"kato_pattern": "3940%", "sheet_name": "Денисовский район"},
    {"kato_pattern": "3942%", "sheet_name": "Жангельдинский район"},
    {"kato_pattern": "3944%", "sheet_name": "Житикаринский район"},
    {"kato_pattern": "3948%", "sheet_name": "Камыстинский район"},
    {"kato_pattern": "3950%", "sheet_name": "Карабалыкский район"},
    {"kato_pattern": "3952%", "sheet_name": "Карасуский район"},
    {"kato_pattern": "3954%", "sheet_name": "Костанайский район"},
    {"kato_pattern": "3956%", "sheet_name": "Мендыкаринский район"},
    {"kato_pattern": "3958%", "sheet_name": "Наурзумский район"},
    {"kato_pattern": "3962%", "sheet_name": "Сарыкольский район"},
    {"kato_pattern": "3964%", "sheet_name": "Беимбета Майлина район"},
    {"kato_pattern": "3966%", "sheet_name": "Узункольский район"},
    {"kato_pattern": "3968%", "sheet_name": "Федоровский район"}
]




first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

