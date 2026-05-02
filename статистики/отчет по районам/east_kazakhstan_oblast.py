from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Восточно-Казахстанская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "6300%", "sheet_name": "Восточно-Казахстанская область.xlsx"},
    {"kato_pattern": "6310%", "sheet_name": "Усть-Каменогорск"},
    {"kato_pattern": "6318%", "sheet_name": "Аягоз"},
    {"kato_pattern": "6320%", "sheet_name": "Зыряновск"},
    {"kato_pattern": "6322%", "sheet_name": "Курчатов"},
    {"kato_pattern": "6324%", "sheet_name": "Риддер"},
    {"kato_pattern": "6328%", "sheet_name": "Семей"},
    {"kato_pattern": "6332%", "sheet_name": "Абайский район"},
    {"kato_pattern": "6334%", "sheet_name": "Аягозский район"},
    {"kato_pattern": "6336%", "sheet_name": "Бескарагайский район"},
    {"kato_pattern": "6338%", "sheet_name": "Бородулихинский район"},
    {"kato_pattern": "6340%", "sheet_name": "Глубоковский район"},
    {"kato_pattern": "6344%", "sheet_name": "Жарминский район"},
    {"kato_pattern": "6346%", "sheet_name": "Зайсанский район"},
    {"kato_pattern": "6348%", "sheet_name": "Алтай район"},
    {"kato_pattern": "6350%", "sheet_name": "Кокпектинский район"},
    {"kato_pattern": "6352%", "sheet_name": "Курчумский район"},
    {"kato_pattern": "6354%", "sheet_name": "Катон-Карагайский район"},
    {"kato_pattern": "6355%", "sheet_name": "район Марқакөл"},
    {"kato_pattern": "6356%", "sheet_name": "Самар район"},
    {"kato_pattern": "6358%", "sheet_name": "Тарбагатайский район"},
    {"kato_pattern": "6362%", "sheet_name": "Уланский район"},
    {"kato_pattern": "6363%", "sheet_name": "район Үлкен Нарын"},
    {"kato_pattern": "6364%", "sheet_name": "Урджарский район"},
    {"kato_pattern": "6368%", "sheet_name": "Шемонаихинский район"},
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

