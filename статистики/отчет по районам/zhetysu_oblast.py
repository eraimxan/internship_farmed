from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Жетісу область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "3300%", "sheet_name": "Жетісу область"},
    {"kato_pattern": "3310%", "sheet_name": "Талдыкорган"},
    {"kato_pattern": "3318%", "sheet_name": "Текели"},
    {"kato_pattern": "3332%", "sheet_name": "Аксуский район"},
    {"kato_pattern": "3334%", "sheet_name": "Алакольский район"},
    {"kato_pattern": "3336%", "sheet_name": "Ескельдинский район"},
    {"kato_pattern": "3340%", "sheet_name": "Кербулакский район"},
    {"kato_pattern": "3342%", "sheet_name": "Коксуский район"},
    {"kato_pattern": "3344%", "sheet_name": "Каратальский район"},
    {"kato_pattern": "3346%", "sheet_name": "Панфиловский район"},
    {"kato_pattern": "3348%", "sheet_name": "Саркандский район"}
]


first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

