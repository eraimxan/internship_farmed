from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Алматинская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "1900%", "sheet_name": "Алматинская область"},
    {"kato_pattern": "1916%", "sheet_name": "Конаев Г.А."},
    {"kato_pattern": "1918%", "sheet_name": "Алатау Г.А."},
    {"kato_pattern": "1936%", "sheet_name": "Балхашский район"},
    {"kato_pattern": "1940%", "sheet_name": "Енбекшиказахский район"},
    {"kato_pattern": "1942%", "sheet_name": "Жамбылский район"},
    {"kato_pattern": "1944%", "sheet_name": "Кегенский район"},
    {"kato_pattern": "1952%", "sheet_name": "Карасайский район"},
    {"kato_pattern": "1958%", "sheet_name": "Райымбекский район"},
    {"kato_pattern": "1962%", "sheet_name": "Талгарский район"},
    {"kato_pattern": "1966%", "sheet_name": "Уйгурский район"},
    {"kato_pattern": "1968%", "sheet_name": "Илийский район"},
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

