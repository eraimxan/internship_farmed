from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Атырауская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "2300%", "sheet_name": "Атырауская область"},
    {"kato_pattern": "2310%", "sheet_name": "Атырау Г.А."},
    {"kato_pattern": "2336%", "sheet_name": "Жылыойский район"},
    {"kato_pattern": "2340%", "sheet_name": "Индерский район"},
    {"kato_pattern": "2342%", "sheet_name": "Исатайский район"},
    {"kato_pattern": "2346%", "sheet_name": "Курмангазинский район"},
    {"kato_pattern": "2348%", "sheet_name": "Кзылкогинский район"},
    {"kato_pattern": "2352%", "sheet_name": "Макатский район"},
    {"kato_pattern": "2356%", "sheet_name": "Махамбетский район"},
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

