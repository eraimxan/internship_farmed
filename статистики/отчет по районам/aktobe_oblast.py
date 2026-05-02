from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Актюбинская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "1500%", "sheet_name": "Актюбинская область"},
    {"kato_pattern": "1510%", "sheet_name": "Актобе Г.А."},
    {"kato_pattern": "1532%", "sheet_name": "Алгинский район"},
    {"kato_pattern": "1534%", "sheet_name": "Айтекебийский район"},
    {"kato_pattern": "1536%", "sheet_name": "Байганинский район"},
    {"kato_pattern": "1540%", "sheet_name": "Каргалинский район"},
    {"kato_pattern": "1542%", "sheet_name": "Хобдинский район"},
    {"kato_pattern": "1546%", "sheet_name": "Мартукский район"},
    {"kato_pattern": "1548%", "sheet_name": "Мугалжарский район"},
    {"kato_pattern": "1552%", "sheet_name": "Уилский район"},
    {"kato_pattern": "1556%", "sheet_name": "Темирский район"},
    {"kato_pattern": "1560%", "sheet_name": "Хромтауский район"},
    {"kato_pattern": "1564%", "sheet_name": "Шалкарский район"},
    {"kato_pattern": "1568%", "sheet_name": "Иргизский район"},
]

first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

