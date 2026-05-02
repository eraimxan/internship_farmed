from district_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import os

db_params = {
    'dbname': 'emfreportbpr',
    'user': 'postgres',
    'password': '159753NBER',
    'host': '56.228.80.80',
    'port': '5432'
}

output_file = "2025август/Северо-Казахстанская область.xlsx"

# Создаём папку, если её нет
os.makedirs(os.path.dirname(output_file), exist_ok=True)


# Если файла нет — создаём с временным листом
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Временный лист"
    wb.save(output_file)

reports = [
    {"kato_pattern": "5900%", "sheet_name": "Северо-Казахстанская область"},
    {"kato_pattern": "5910%", "sheet_name": "Петропавловск"},
    {"kato_pattern": "5932%", "sheet_name": "Айыртауский район"},
    {"kato_pattern": "5934%", "sheet_name": "Акжарский район"},
    {"kato_pattern": "5936%", "sheet_name": "Магжана Жумабаева район"},
    {"kato_pattern": "5942%", "sheet_name": "Есильский район"},
    {"kato_pattern": "5946%", "sheet_name": "Жамбылский район"},
    {"kato_pattern": "5950%", "sheet_name": "Кызылжарский район"},
    {"kato_pattern": "5952%", "sheet_name": "Мамлютский район"},
    {"kato_pattern": "5956%", "sheet_name": "Шал акына район"},
    {"kato_pattern": "5958%", "sheet_name": "Аккайынский район"},
    {"kato_pattern": "5960%", "sheet_name": "Тайыншинский район"},
    {"kato_pattern": "5962%", "sheet_name": "Тимирязевский район"},
    {"kato_pattern": "5964%", "sheet_name": "Уалихановский район"},
    {"kato_pattern": "5966%", "sheet_name": "Габита Мусрепова район"},
]



first = True
for report in reports:
    create_table_in_excel(
        db_params=db_params,
        file_path=output_file,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"]
    )

    # Удаляем временный лист только после первой вставки
    if first:
        wb = load_workbook(output_file)
        if "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Временный лист"]
            wb.save(output_file)
        first = False

print(f"Файл успешно создан: {output_file}")

