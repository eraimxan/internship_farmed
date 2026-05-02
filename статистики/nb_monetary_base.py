import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def build_excel_monetary_base(data, file_path="monetary_base.xlsx", chart_path="chart.png"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Денежная база"

    # Заголовки
    ws["A1"] = "Денежная база и агрегаты широкой денежной массы"
    ws["A2"] = "млн. тенге / на конец периода"
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Формируем подписи месяцев
    months = [pd.to_datetime(d["reporting_date"]).strftime("%m.%Y") for d in data]
    ws.append(["Категория"] + months)

    # стиль для заголовков
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Основные строки (как у тебя)
    rows = [
        ("1. Денежная база (резервные деньги)", "m_b_r_m", True),
        ("Изменения за месяц, %", "m_b_r_m_monthly_change", False),
        ("Изменение за период с начала года, %", "m_b_r_m_change_begin_year", False),
        ("1.1 наличные деньги вне НБК", "m_b_r_m_cash_outside_nbk", False),
        ("1.2 депозиты БВУ и других организаций в НБК", "m_b_r_m_deposit_stb_and_other_nbk", False),
        ("Денежная база (в узком выражении)", "m_b_n_t", True),
        ("Изменения за месяц, %", "m_b_n_t_monthly_change", False),
        ("Изменение за период с начала года, %", "m_b_n_t_change_begin_year", False),
        ("резервные депозиты БВУ в НБРК", "m_b_n_t_stb_reserve_nbk", False),
        ("2. M0 (наличные деньги в обращении)", "m_0", True),
        ("Изменения за месяц, %", "m_0_monthly_change", False),
        ("Изменение за период с начала года, %", "m_0_change_begin_year", False),
        ("3. M1", "m_1", True),
        ("Изменения за месяц, %", "m_1_monthly_change", False),
        ("Изменение за период с начала года, %", "m_1_change_begin_year", False),
        ("3.1 переводимые депозиты населения в тенге", "m_1_trans_deposit_population_tenge", False),
        ("3.2 переводимые депозиты небанковских юр. лиц в тенге", "m_1_trans_deposit_not_banking_tenge", False),
        ("4. M2", "m_2", True),
        ("Изменения за месяц, %", "m_2_monthly_change", False),
        ("Изменение за период с начала года, %", "m_2_change_begin_year", False),
        ("4.1 другие депозиты населения (тенге + FX)", "m_2_deposit_tenge_and_foreign_currency_population", False),
        ("4.2 другие депозиты небанковских юр. лиц (тенге + FX)", "m_2_deposit_tenge_and_foreign_currency_not_banking", False),
        ("5. M3 (денежная масса)", "m_3", True),
        ("Изменения за месяц, %", "m_3_monthly_change", False),
        ("Изменение за период с начала года, %", "m_3_change_begin_year", False),
        ("5.1 другие депозиты населения в иностранной валюте", "m_3_other_foreign_currency_population", False),
        ("5.2 другие депозиты небанковских юр. лиц в иностранной валюте", "m_3_other_foreign_currency_not_banking", False),
    ]

    # Заполняем таблицу
    for row_title, key, bold in rows:
        values = [d.get(key, "") if key else "" for d in data]
        row = [row_title] + values
        ws.append(row)

        r = ws.max_row
        ws[f"A{r}"].font = Font(bold=bold)
        for c in range(2, len(values) + 2):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(horizontal="right")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '# ##0.0'

    # Автоширина колонок
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            try:
                if cell.value is not None:
                    # Если число — считаем длину форматированного отображения
                    if isinstance(cell.value, (int, float)):
                        text = f"{cell.value:,.1f}".replace(",", " ")
                    else:
                        text = str(cell.value)
                    max_length = max(max_length, len(text))
            except:
                pass
        # ставим ширину с запасом
        ws.column_dimensions[col_letter].width = max_length + 2

    # ---------------- Второй лист с графиком ----------------
    ws2 = wb.create_sheet("График агрегатов")

    df = pd.DataFrame([{
        "Дата": pd.to_datetime(d["reporting_date"]).strftime("%m.%y"),
        "M3 YTD": d.get("m_3_change_begin_year"),
        "База YTD": d.get("m_b_r_m_change_begin_year"),
        "Наличные YTD": d.get("m_0_change_begin_year"),
        "Инфляция": d.get("cpi_ytd"),
    } for d in data])

    df = df.fillna(0)

    # сохраняем таблицу
    ws2.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws2.append(row)

    # строим красивый график
    fig, ax1 = plt.subplots(figsize=(10, 6))
    width = 0.25
    x = range(len(df))

    ax1.bar([i - width for i in x], df["M3 YTD"], width, label="M3 YTD", color="#003366")
    ax1.bar([i for i in x], df["База YTD"], width, label="Денежная база YTD", color="#008000")
    ax1.bar([i + width for i in x], df["Наличные YTD"], width, label="Наличные YTD", color="#00BFBF")

    ax1.set_ylabel("% YTD")
    ax1.set_xticks(x)
    ax1.set_xticklabels(df["Дата"])

    # диапазон Y учитывает отрицательные значения
    ymin = min(df[["M3 YTD", "База YTD", "Наличные YTD"]].min().min(), 0)
    ymax = max(df[["M3 YTD", "База YTD", "Наличные YTD"]].max().max(), 0)
    ax1.set_ylim(ymin * 1.3, ymax * 1.3)

    # подписи процентов (над/под в зависимости от знака)
    for i, v in enumerate(df["M3 YTD"]):
        ax1.text(i - width, v + (0.5 if v >= 0 else -0.5), f"{v:.1f}%",
                ha="center", va="bottom" if v >= 0 else "top", fontsize=8)

    for i, v in enumerate(df["База YTD"]):
        ax1.text(i, v + (0.5 if v >= 0 else -0.5), f"{v:.1f}%",
                ha="center", va="bottom" if v >= 0 else "top", fontsize=8)

    for i, v in enumerate(df["Наличные YTD"]):
        ax1.text(i + width, v + (0.5 if v >= 0 else -0.5), f"{v:.1f}%",
                ha="center", va="bottom" if v >= 0 else "top", fontsize=8)

    # инфляция линия
    ax2 = ax1.twinx()
    ax2.plot(x, df["Инфляция"], color="red", marker="o", label="Инфляция")
    ax2.set_ylabel("Инфляция %")
    ax2.set_ylim(0, max(df["Инфляция"]) * 1.3 if df["Инфляция"].max() > 0 else 10)

    # легенда снизу
    lines, labels = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines + lines2, labels + labels2,
            loc="upper center", bbox_to_anchor=(0.5, -0.1), ncol=4)
    ax1.tick_params(axis="x", pad=10)
    
    for spine in ["top", "right", "left", "bottom"]:
        ax1.spines[spine].set_visible(False)
        ax2.spines[spine].set_visible(False)


    plt.title("Динамика денежных агрегатов Казахстана с начала года в %")
    plt.tight_layout()
    plt.savefig(chart_path, dpi=200)
    plt.close()

    # вставляем картинку в Excel
    img = Image(chart_path)
    ws2.add_image(img, "H2")


        # ---------------- Третий график ----------------
    ws3 = wb.create_sheet("Агрегаты (линии)")

    df2 = pd.DataFrame([{
        "Дата": pd.to_datetime(d["reporting_date"]).strftime("%m.%y"),
        "Денежная база": d.get("m_b_r_m"),
        "M0": d.get("m_0"),
        "M2": d.get("m_2"),
        "M3": d.get("m_3"),
    } for d in data])

    df2 = df2.fillna(0)

    # сохраняем таблицу
    ws3.append(list(df2.columns))
    for row in df2.itertuples(index=False):
        ws3.append(row)

    # строим линейный график
    fig, ax = plt.subplots(figsize=(10,6))

    ax.plot(df2["Дата"], df2["Денежная база"], marker="o", label="1. Денежная база (резервные деньги)")
    ax.plot(df2["Дата"], df2["M0"], marker="o", label="2. M0 (наличные деньги в обращении)")
    ax.plot(df2["Дата"], df2["M2"], marker="o", label="M2, млн тенге")
    ax.plot(df2["Дата"], df2["M3"], marker="o", label="M3, млн тенге")

    ax.set_title("Денежные агрегаты Казахстана")
    ax.set_ylabel("млн тенге")
    ax.set_xlabel("Дата")
    ax.grid(True, linestyle="--", alpha=0.6)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.1), ncol=2)

    plt.tight_layout()
    chart_path2 = "chart2.png"
    plt.savefig(chart_path2, dpi=200)
    plt.close()

    # вставляем картинку в Excel
    img2 = Image(chart_path2)
    ws3.add_image(img2, "H2")


    wb.save(file_path)
    return file_path
