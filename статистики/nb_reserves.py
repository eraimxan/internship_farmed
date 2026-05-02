import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image


def build_excel_reserves(data, file_path="international_reserves.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Международные резервы"

    # Заголовок
    ws["A1"] = "Международные резервы и валютные активы Национального фонда РК"
    ws.merge_cells("A1:P1")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Подзаголовки
    headers = [
        "Дата",

        "Валовые международные резервы (объем, млн.$)",
        "изменение к предыдущему месяцу, %",
        "изменение к началу года, %",

        "Активы в СКВ (объем, млн.$)",
        "изменение к предыдущему месяцу, %",
        "изменение к началу года, %",

        "Монетарное золото (объем, млн.$)",
        "изменение к предыдущему месяцу, %",
        "изменение к началу года, %",

        "Чистые международные резервы (объем, млн.$)",
        "изменение к предыдущему месяцу, %",
        "изменение к началу года, %",

        "Валютные активы Нацфонда РК (объем, млн.$)",
        "изменение к предыдущему месяцу, %",
        "изменение к началу года, %",
    ]
    ws.append(headers)

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Данные
    df = pd.DataFrame([{
        "Дата": pd.to_datetime(d["reporting_date"]),
        "Валовые": d.get("gross_international_reserve_volume_million_dollar"),
        "СКВ": d.get("assets_hard_currency_volume_million_dollar"),
        "Золото": d.get("gold_volume_million_dollar"),
        "Чистые": d.get("net_international_reserve_volume_million_dollar"),
        "Нацфонд": d.get("national_fund_republic_of_kazakhstan_volume_million_dollar"),
    } for d in data])

    df = df.sort_values("Дата")
    df["Месяц"] = df["Дата"].dt.strftime("%m.%y")

    for _, row in df.iterrows():
        ws.append([
            row["Месяц"],
            row["Валовые"], "", "",
            row["СКВ"], "", "",
            row["Золото"], "", "",
            row["Чистые"], "", "",
            row["Нацфонд"], "", ""
        ])

    # Форматирование
    for row in ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # Автоширина
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2


    # ---------------- График ----------------
    ws2 = wb.create_sheet("Графики")

    # 1 Чистые золотовалютные резервы
    fig1, ax1 = plt.subplots(figsize=(9, 4))
    ax1.fill_between(df["Дата"], df["Чистые"], color="#3169b3")
    ax1.set_title("Динамика чистых золотовалютных резервов НБ РК", fontsize=12, fontweight="bold")
    ax1.set_ylabel("млн USD")
    ax1.grid(axis='y', color='gray', alpha=0.3)
    ax1.set_facecolor("white")
    plt.xticks(rotation=45, ha='right')
    img1 = BytesIO()
    plt.tight_layout()
    plt.savefig(img1, format="png", dpi=200)
    plt.close(fig1)
    img1.seek(0)
    ws2.add_image(Image(img1), "A1")

    # 2️ Столбчатый график (СКВ, Золото, Нацфонд)
    fig2, ax2 = plt.subplots(figsize=(9, 4))
    width = 0.6
    x = range(len(df))
    ax2.bar(x, df["СКВ"], color="#00A85A", label="СКВ")
    ax2.bar(x, df["Золото"], bottom=df["СКВ"], color="#FFEE00", label="Золото")
    ax2.bar(x, df["Нацфонд"], bottom=df["СКВ"] + df["Золото"], color="white", edgecolor="black", label="Нацфонд")
    ax2.set_facecolor("#001848")
    ax2.set_title("Международные резервы Казахстана, в млн USD", fontsize=12, fontweight="bold", color="white")
    ax2.tick_params(colors="white")
    ax2.legend(facecolor="#001848", labelcolor="white")
    ax2.set_xticks(x)
    ax2.set_xticklabels(df["Месяц"], rotation=45, ha="right", color="white")
    ax2.set_yticks(range(0, int(df[["СКВ", "Золото", "Нацфонд"]].sum(axis=1).max()) + 20000, 20000))
    ax2.set_yticklabels([f"{y:,}".replace(",", " ") for y in ax2.get_yticks()], color="white")
    plt.tight_layout()
    img2 = BytesIO()
    plt.savefig(img2, format="png", dpi=200, facecolor="#001848")
    plt.close(fig2)
    img2.seek(0)
    ws2.add_image(Image(img2), "A25")

    # 3️ Общая динамика резервов и Нацфонда
    fig3, ax3 = plt.subplots(figsize=(9, 4))
    ax3.fill_between(df["Дата"], df["Нацфонд"], color="#00A85A", label="Нацфонд")
    ax3.fill_between(df["Дата"], df["Нацфонд"] + df["Чистые"], df["Нацфонд"], color="#FFEE00", label="Чистые ЗВР")
    ax3.set_title("Динамика международных резервов Казахстана за выбранный период", fontsize=12, fontweight="bold")
    ax3.set_ylabel("млн USD")
    ax3.legend()
    ax3.grid(axis='y', alpha=0.3)
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    img3 = BytesIO()
    plt.savefig(img3, format="png", dpi=200)
    plt.close(fig3)
    img3.seek(0)
    ws2.add_image(Image(img3), "A50")

    wb.save(file_path)
    return file_path
