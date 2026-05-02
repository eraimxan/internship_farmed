import concurrent.futures
import datetime
import requests
import urllib3
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from collections import OrderedDict, defaultdict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
matplotlib.rcParams["font.family"] = "DejaVu Sans"
matplotlib.rcParams["axes.unicode_minus"] = False

API_URL = "https://ows.goszakup.gov.kz/v3/graphql"
HEADERS = {
    "Authorization": "Bearer 8530636543c42453dc78b894f3c32ab6",
    "Content-Type": "application/json",
}

GRAPHQL_QUERY = """
query EmfReportBpr($after: Int, $date: String!) {
  EmfReportBpr(limit: 200, after: $after, filter: { dateReport: $date }) {
    grbId, strId, katFgr, katFgrName,
    klsFpgr, klsFpgrName, pklAbp, pklAbpName,
    adtype, utv, utch, plg, sumrg
  }
}
"""

# datereport values: Jan-01 of next year = annual execution of that year
DATE_2023 = "2024-01-01"
DATE_2024 = "2025-01-01"
DATE_2025 = "2026-01-01"


def _get_latest_month_candidates(n=4):
    """Generate n month-start dates going back from last month, skipping Jan-01 annual dates."""
    annual = {f"{y}-01-01" for y in range(2016, datetime.date.today().year + 3)}
    today = datetime.date.today()
    y, m = today.year, today.month - 1
    if m == 0:
        m, y = 12, y - 1
    candidates = []
    for _ in range(n):
        s = f"{y}-{m:02d}-01"
        if s not in annual:
            candidates.append(s)
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return candidates


# ─────────────────────────── API fetch ───────────────────────────

def _fetch_all(date_str: str) -> list:
    """Fetch all paginated rows for a given datereport."""
    import time
    rows = []
    after = None
    while True:
        payload = {
            "query": GRAPHQL_QUERY,
            "variables": {"after": after, "date": date_str},
        }
        for attempt in range(4):
            try:
                r = requests.post(API_URL, json=payload, headers=HEADERS,
                                  verify=False, timeout=120)
                r.raise_for_status()
                break
            except requests.exceptions.Timeout:
                if attempt == 3:
                    raise
                time.sleep(5 * (attempt + 1))
        data = r.json()
        batch = data.get("data", {}).get("EmfReportBpr", []) or []
        if not batch:
            break
        rows.extend(batch)
        page_info = (data.get("extensions") or {}).get("pageInfo") or {}
        last_id = page_info.get("lastId")
        if not page_info.get("hasNextPage") or not last_id:
            break
        after = last_id
    return rows


def _fetch_with_fallback(primary: str, fallback: str) -> list:
    """Try primary date; if no income rows found, try fallback."""
    rows = _fetch_all(primary)
    if not any(_is_income(r) for r in rows):
        rows = _fetch_all(fallback)
    return rows


def _find_latest_month_income() -> tuple[str, list]:
    """Return (label, rows) for the newest month that has income data."""
    candidates = _get_latest_month_candidates()
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(candidates)) as ex:
        futures = {c: ex.submit(_fetch_all, c) for c in candidates}
    for c in candidates:
        try:
            rows = futures[c].result()
        except Exception:
            continue
        if any(_is_income(r) for r in rows):
            return f"{c[5:7]}.{c[:4]} Исп.", rows
    return None, []


# ─────────────────────────── filtering ───────────────────────────

def _safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _is_income(r: dict) -> bool:
    """Replicate the SQL grbid=1 revenue filter in Python."""
    grb    = _safe_int(r.get("grbId"))
    katfgr = _safe_int(r.get("katFgr"))
    strid  = _safe_int(r.get("strId"))
    adtype = _safe_int(r.get("adtype"))
    if grb != 1:
        return False
    if katfgr == 1 and strid == 1:
        return True
    if katfgr == 2 and strid == 2 and adtype is None:
        return True
    if katfgr == 3 and strid == 3:
        return True
    if katfgr == 4 and strid == 4 and adtype not in (3, 4):
        return True
    return False


# ─────────────────────────── aggregation ───────────────────────────

def _to_mln(v) -> float:
    """Convert raw string/number value to millions of tenge."""
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ".")) / 1_000_000
    except Exception:
        return 0.0


def _compute_maps(rows: list) -> dict:
    """
    Build lookup dicts keyed by:
      ("SEC",)                         — republic total
      ("FGR", katFgr, katFgrName)      — category subtotal
      ("PKL", katFgr, katFgrName, pkl) — line item
    Each value: {sumrg, plg, utv, utch}
    """
    fields = ("sumrg", "plg", "utv", "utch")
    sec  = defaultdict(float)
    fgr  = defaultdict(lambda: defaultdict(float))
    pkl_ = defaultdict(lambda: defaultdict(float))

    for r in rows:
        if not _is_income(r):
            continue
        katFgr    = _safe_int(r.get("katFgr"))
        katFgrName = r.get("katFgrName") or ""
        pklAbpName = r.get("pklAbpName") or ""

        fgr_key = ("FGR", katFgr, katFgrName)
        pkl_key = ("PKL", katFgr, katFgrName, pklAbpName)

        for f in fields:
            val = _to_mln(r.get(f))
            sec[f]         += val
            fgr[fgr_key][f] += val
            pkl_[pkl_key][f] += val

    result = {("SEC",): dict(sec)}
    result.update({k: dict(v) for k, v in fgr.items()})
    result.update({k: dict(v) for k, v in pkl_.items()})
    return result


def _build_scheme(rows_list: list) -> list:
    """Build ordered row scheme from the union of all data sets."""
    fgr_groups = OrderedDict()
    for rows in rows_list:
        for r in rows:
            if not _is_income(r):
                continue
            katFgr     = _safe_int(r.get("katFgr"))
            katFgrName  = r.get("katFgrName") or ""
            pklAbpName  = r.get("pklAbpName") or ""
            pkl_list = fgr_groups.setdefault((katFgr, katFgrName), [])
            if pklAbpName and pklAbpName not in pkl_list:
                pkl_list.append(pklAbpName)

    scheme = [(("SEC",), "I. ДОХОДЫ", "sec")]
    for (katFgr, katFgrName), pkl_list in fgr_groups.items():
        scheme.append((("FGR", katFgr, katFgrName), katFgrName, "fgr"))
        for pkl in pkl_list:
            scheme.append((("PKL", katFgr, katFgrName, pkl), f"   {pkl}", "pkl"))
    return scheme


# ─────────────────────────── Excel builder ───────────────────────────

def build_excel_income_report(filename: str) -> str:
    """
    Sheet 1 — hierarchy table (revenue categories × periods)
    Sheet 2 — bar chart comparing 2023 / 2024 / 2025 execution by category
    """
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        f2023 = executor.submit(_fetch_all, DATE_2023)
        f2024 = executor.submit(_fetch_all, DATE_2024)
        f2025 = executor.submit(_fetch_with_fallback, DATE_2025, "2025-12-01")
        f_month = executor.submit(_find_latest_month_income)

    r2023 = f2023.result()
    r2024 = f2024.result()
    r2025 = f2025.result()
    month_label, r_month = f_month.result()

    all_rows = [r2023, r2024, r2025] + ([r_month] if r_month else [])
    scheme = _build_scheme(all_rows)

    map2025 = _compute_maps(r2025)
    vm = {
        "2023_sumrg": _compute_maps(r2023),
        "2024_sumrg": _compute_maps(r2024),
        "2025_sumrg": map2025,
        "2025_utv":   map2025,
        "2025_utch":  map2025,
        "2025_plg":   map2025,
    }
    if r_month:
        vm["month_sumrg"] = _compute_maps(r_month)

    col_labels = ["2023 Исп.", "2024 Исп.", "2025 Исп.", "2025 Утв.", "2025 Уточн.", "2025 Скорр."]
    DATA_COLS = [
        ("2023_sumrg", "sumrg"),
        ("2024_sumrg", "sumrg"),
        ("2025_sumrg", "sumrg"),
        ("2025_utv",   "utv"),
        ("2025_utch",  "utch"),
        ("2025_plg",   "plg"),
    ]
    if r_month:
        col_labels.append(month_label)
        DATA_COLS.append(("month_sumrg", "sumrg"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Доходы"

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_cols = 1 + len(col_labels)

    # ── Title ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = "ИСПОЛНЕНИЕ РЕСПУБЛИКАНСКОГО БЮДЖЕТА — ДОХОДЫ"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"] = "(млн тенге)"
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Header ──
    ws.append(["Наименование"] + col_labels)
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 30

    # ── Data rows ──
    for key, label, style in scheme:
        vals = [round(vm[dk].get(key, {}).get(field, 0.0), 1) for dk, field in DATA_COLS]
        ws.append([label] + vals)
        r = ws.max_row
        if style == "sec":
            ws.cell(r, 1).font = Font(bold=True, size=10)
        elif style == "fgr":
            ws.cell(r, 1).font = Font(bold=True, size=9)
        else:
            ws.cell(r, 1).font = Font(size=8)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="center")
        for c in range(1, total_cols + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if c > 1:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.0"

    ws.column_dimensions["A"].width = 55
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ── Chart ──
    ws_chart = wb.create_sheet("График")

    fgr_labels, v23, v24, v25 = [], [], [], []
    for key, label, style in scheme:
        if style == "fgr":
            _, katFgr, katFgrName = key
            lookup = ("FGR", katFgr, katFgrName)
            fgr_labels.append(katFgrName)
            v23.append(vm["2023_sumrg"].get(lookup, {}).get("sumrg", 0.0) / 1_000)
            v24.append(vm["2024_sumrg"].get(lookup, {}).get("sumrg", 0.0) / 1_000)
            v25.append(vm["2025_sumrg"].get(lookup, {}).get("sumrg", 0.0) / 1_000)

    if fgr_labels:
        x = range(len(fgr_labels))
        w = 0.25
        fig, ax = plt.subplots(figsize=(12, 6))
        ax.bar([i - w for i in x], v23, w, label="2023 Исп.", color="#003366")
        ax.bar(list(x),            v24, w, label="2024 Исп.", color="#0066CC")
        ax.bar([i + w for i in x], v25, w, label="2025 Исп.", color="#00AAFF")

        short = [lb[:30] + "…" if len(lb) > 30 else lb for lb in fgr_labels]
        ax.set_xticks(list(x))
        ax.set_xticklabels(short, rotation=25, ha="right", fontsize=8)
        ax.set_ylabel("млрд тенге")
        ax.set_title("Исполнение республиканского бюджета по категориям доходов",
                     fontweight="bold", pad=15)
        ax.legend()
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        plt.figtext(0.01, -0.02, "По данным Министерства финансов РК", fontsize=8, color="gray")
        plt.figtext(0.85, -0.02, "TENGENOMIKA", fontsize=9, fontweight="bold", color="#003366")
        plt.tight_layout()

        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=200, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        ws_chart.add_image(Image(buf), "A1")

    wb.save(filename)
    return filename
