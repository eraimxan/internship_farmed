import requests
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter, MonthLocator
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import matplotlib

# ----------------------------- Настройка отображения кириллицы -----------------------------
matplotlib.rcParams['font.family'] = 'DejaVu Sans'
matplotlib.rcParams['axes.unicode_minus'] = False

API_URL = "https://data.nationalbank.kz/api/report/stat-data-locale?form_id=299&date_from=2014-01-01&date_to=2050-01-01&lang=ru"
TARGET_CATEGORY = "Реальный эффективный обменный курс"


def load_data():
    """Загружает данные РЭОК из API НБ РК"""
    r = requests.get(API_URL, timeout=(10, 30))
    r.raise_for_status()
    data = r.json()

    rows = []
    for item in data:
        date = item.get("reportDate")
        amount = item.get("amount")
        attrs = item.get("attributes", [])

        category = None
        period_type = None
        for a in attrs:
            if a.get("attrCode") == "category":
                category = a.get("valueNameRu")
            if a.get("attrCode") == "period":
                period_type = a.get("valueCode")

        # Только месячные данные для категории "РЭОК"
        if category == TARGET_CATEGORY and period_type != "quarter" and date and amount is not None:
            d = pd.to_datetime(date)
            if d >= pd.Timestamp("2015-01-01"):
                rows.append({"Дата": d, "Индекс": float(amount)})

    df = pd.DataFrame(rows).sort_values("Дата").reset_index(drop=True)
    return df


def build_excel_with_chart(filename):
    """
    Строит Excel-файл с таблицей и графиком РЭОК.
    Если в названии файла есть год — фильтрует данные до этого года включительно.
    """
    df = load_data()

    # --- Определяем год из имени файла (если есть, например "РЭОК_месячные_2023.xlsx") ---
    import re
    match = re.search(r"(\d{4})", filename)
    if match:
        selected_year = int(match.group(1))
        df = df[df["Дата"].dt.year <= selected_year]

    # --- Создание Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "РЭОК данные"

    ws.merge_cells("A1:B1")
    ws["A1"] = "Реальный эффективный обменный курс (месячные данные)"
    ws["A1"].font = Font(size=13, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(["Дата", "Индекс"])
    for _, row in df.iterrows():
        ws.append([row["Дата"].strftime("%Y-%m"), round(row["Индекс"], 2)])

    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for c in row:
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # --- Второй лист: график ---
    ws_chart = wb.create_sheet("График РЭОК")

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(df["Дата"], df["Индекс"], color="#1236c2", linewidth=2.5, label="РЭОК")

    ax.set_title("Индексы реального эффективного обменного курса тенге",
                 fontsize=12, fontweight="bold", pad=20)
    ax.set_ylabel("Индекс", fontsize=10)
    ax.grid(True, linestyle="--", alpha=0.5)

    for spine in ax.spines.values():
        spine.set_linewidth(1.2)
        spine.set_color("#000000")

    MONTHS_RU = {
        'Jan': 'янв', 'Feb': 'фев', 'Mar': 'мар', 'Apr': 'апр', 'May': 'май',
        'Jun': 'июн', 'Jul': 'июл', 'Aug': 'авг', 'Sep': 'сен', 'Oct': 'окт',
        'Nov': 'ноя', 'Dec': 'дек'
    }

    ax.xaxis.set_major_locator(MonthLocator(interval=4))
    ax.xaxis.set_major_formatter(DateFormatter('%b.%y'))

    plt.xticks(rotation=45, ha="right")
    plt.tick_params(axis="x", labelsize=8)
    plt.tick_params(axis="y", labelsize=8)

    plt.draw()
    tick_positions = ax.get_xticks()
    labels = [item.get_text() for item in ax.get_xticklabels()]
    new_labels = []
    for lbl in labels:
        if '.' in lbl:
            month_eng, year = lbl.split('.')
            month_ru = MONTHS_RU.get(month_eng.capitalize(), month_eng)
            new_labels.append(f"{month_ru}.{year}")
        else:
            new_labels.append(lbl)

    ax.set_xticks(tick_positions)
    ax.set_xticklabels(new_labels, rotation=45, ha="right")

    ax.legend(["РЭОК"], loc="upper right", fontsize=9)
    plt.figtext(0.01, -0.02, "Источник: по данным НБ РК", fontsize=9, color="gray")
    plt.figtext(0.87, -0.02, "TENGENOMIKA", fontsize=9, color="#1236c2", fontweight="bold")

    plt.tight_layout()

    img_data = BytesIO()
    plt.savefig(img_data, format="png", dpi=200, bbox_inches="tight")
    img_data.seek(0)
    plt.close()

    img = Image(img_data)
    ws_chart.add_image(img, "A1")

    wb.save(filename)
    return filename
