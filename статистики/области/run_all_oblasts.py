from oblast_report_builder import create_table_in_excel
from openpyxl import Workbook, load_workbook
import psycopg2
import os

db_params = {
    'dbname': 'ERI',
    'user': 'postgres',
    'password': '20041015R',
    'host': 'localhost',
    'port': '5432'
}

output_file = "2023-2025/2025год.xlsx"
datereport = "2026-02-01"

os.makedirs(os.path.dirname(output_file), exist_ok=True)

# Создаём файл, если его нет
if not os.path.exists(output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "202602"
    wb.save(output_file)

# Открываем книгу один раз
wb = load_workbook(output_file)

# Открываем соединение один раз
conn = psycopg2.connect(**db_params)

reports = [
    {"kato_pattern": "10%", "sheet_name": "Абайская область"},
    {"kato_pattern": "11%", "sheet_name": "Акмолинская область"},
    {"kato_pattern": "15%", "sheet_name": "Актюбинская область"},
    {"kato_pattern": "19%", "sheet_name": "Алматинская область"},
    {"kato_pattern": "23%", "sheet_name": "Атырауская область"},
    {"kato_pattern": "27%", "sheet_name": "Западно-Казахстанская область"},
    {"kato_pattern": "31%", "sheet_name": "Жамбылская область"},
    {"kato_pattern": "33%", "sheet_name": "область Жетісу"},
    {"kato_pattern": "35%", "sheet_name": "Карагандинская область"},
    {"kato_pattern": "39%", "sheet_name": "Костанайская область"},
    {"kato_pattern": "43%", "sheet_name": "Кызылординская область"},
    {"kato_pattern": "47%", "sheet_name": "Мангистауская область"},
    {"kato_pattern": "51%", "sheet_name": "Южно-Казахстанская область"},
    {"kato_pattern": "55%", "sheet_name": "Павлодарская область"},
    {"kato_pattern": "59%", "sheet_name": "Северо-Казахстанская область"},
    {"kato_pattern": "61%", "sheet_name": "Туркестанская область"},
    {"kato_pattern": "62%", "sheet_name": "область Ұлытау"},
    {"kato_pattern": "63%", "sheet_name": "Восточно-Казахстанская область"},
    {"kato_pattern": "71%", "sheet_name": "г.Астана"},
    {"kato_pattern": "75%", "sheet_name": "г.Алматы"},
    {"kato_pattern": "79%", "sheet_name": "г.Шымкент"},
]

first = True
for report in reports:
    create_table_in_excel(
        conn=conn,
        wb=wb,
        kato_pattern=report["kato_pattern"],
        sheet_name=report["sheet_name"],
        datereport=datereport
    )

    if first and "Временный лист" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Временный лист"]
        first = False

# Сохраняем один раз в конце
wb.save(output_file)

# Закрываем соединение
conn.close()

print(f"✅ Файл успешно создан: {output_file}")
