import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import openpyxl


# --------------------- Получение данных из БД ---------------------
def fetch_data(conn, kato_pattern, datereport):
    """Загружает данные из таблицы emfreportbpr_v9 по kato и дате."""
    cur = conn.cursor()
    cur.execute(f"""
        SELECT
            grbId,
            katFgr,
            katFgrName,
            klsFpgrName,
            pklAbpName,
            spkPrgName,
            sumrg,
            plg,
            utv,
            utch,
            plgp,
            plgo,
            kato
        FROM emfreportbpr_202602
        WHERE datereport = '{datereport}'
          AND CAST(kato AS TEXT) LIKE '{kato_pattern}'
          AND (
                (
                    (
                        (grbid = 1 AND katfgr = 1 AND strid = 1) OR
                        (grbid = 1 AND katfgr = 2 AND strid = 2 AND adtype IS NULL) OR
                        (grbid = 1 AND katfgr = 3 AND strid = 3) OR
                        (grbid = 1 AND katfgr = 4) OR
                        (grbid = 2 AND katfgr = 5)
                    )
                )
                OR
                (
                    grbid = 2 AND (
                        (katfgr = 1 AND adtype IS NULL) OR
                        (katfgr = 2) OR
                        (katfgr = 3) OR
                        (katfgr = 4) OR
                        (katfgr = 5 AND (adtype IS NULL OR adtype NOT IN (3, 4))) OR
                        (katfgr = 6 AND adtype IS NULL) OR
                        (katfgr = 7 AND adtype IS NULL) OR
                        (katfgr = 8 AND adtype IS NULL) OR
                        (katfgr = 9) OR
                        (katfgr = 10 AND adtype IS NULL) OR
                        (katfgr = 11 AND adtype IS NULL) OR
                        (katfgr = 12 AND adtype IS NULL) OR
                        (katfgr = 13) OR
                        (katfgr = 14 AND (adtype IS NULL OR adtype NOT IN (3, 4))) OR
                        (katfgr = 15 AND adtype NOT IN (3, 4))
                    )
                )
                OR grbid IN (3, 4, 5, 6)
          )
        GROUP BY
            grbId, katFgr, katFgrName,
            klsFpgrName, pklAbpName, spkPrgName,
            sumrg, plg, utv, utch, plgp, plgo, kato;
    """)
    data = cur.fetchall()
    cur.close()
    return data


# --------------------- Агрегация данных ---------------------
def aggregate_data(data, grb_id, field_index):
    agg, total = {}, 0.0

    for r in data:
        if len(r) != 13 or r[0] != grb_id:
            continue

        val = str(r[field_index]).replace(',', '.')

        try:
            v = float(val) / 1000 if val else 0.0
        except ValueError:
            v = 0.0

        total += v

        _, katFgr, katFgrName, klsFpgrName, pklAbpName, spkPrgName, *_ = r

        fgr = agg.setdefault(katFgr, {
            "name": katFgrName,
            "total": 0,
            "lvl1": {}
        })

        l1 = fgr["lvl1"].setdefault(klsFpgrName, {
            "total": 0,
            "lvl2": {}
        })

        l2 = l1["lvl2"].setdefault(pklAbpName, {
            "total": 0,
            "prog": {}
        })

        l2["prog"][spkPrgName] = l2["prog"].get(spkPrgName, 0) + v

        l2["total"] += v
        l1["total"] += v
        fgr["total"] += v

    return agg, total


# --------------------- Запись секции в Excel ---------------------
def write_section(ws, start_row, agg, total, col, font_main, font_small):
    ws.cell(row=start_row - 1, column=col, value=total).font = Font(
        name='Times New Roman', bold=True, size=9
    )

    row = start_row

    for fgr, d in agg.items():
        ws.cell(row=row, column=7, value=d["name"].upper()).font = font_main
        ws.cell(row=row, column=col, value=d["total"]).font = font_main
        row += 1

        for l1n, l1d in d["lvl1"].items():
            ws.cell(row=row, column=7, value=l1n).font = Font(
                name='Times New Roman', size=8, italic=True
            )
            ws.cell(row=row, column=col, value=l1d["total"]).font = font_small
            row += 1

            for l2n, l2d in l1d["lvl2"].items():
                ws.cell(row=row, column=7, value=f" {l2n}").font = font_small
                ws.cell(row=row, column=col, value=l2d["total"]).font = font_small
                row += 1

                for pn, pv in l2d["prog"].items():
                    ws.cell(row=row, column=7, value=f" – {pn}").font = font_small
                    ws.cell(row=row, column=col, value=pv).font = font_small
                    row += 1

    return row


# --------------------- Основная функция ---------------------
def create_table_in_excel(conn, wb, kato_pattern, sheet_name, datereport):
    """Создает один лист Excel в открытой книге."""

    data = fetch_data(conn, kato_pattern, datereport)

    if not data:
        print(f"⚠️ Нет данных для {sheet_name}")
        return

    # Удаляем старый лист
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # ---------- Настройки оформления ----------
    font_main = Font(name='Times New Roman', bold=True, size=9)
    font_small = Font(name='Times New Roman', size=8)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ---------- Заголовки ----------
    ws.merge_cells('A1:M1')
    ws['A1'] = f"Отчет об исполнении бюджета на {datereport}"
    ws['A1'].font = Font(name='Times New Roman', bold=True, size=12)
    ws['A1'].alignment = center

    ws.merge_cells('A3:F3')
    ws['A3'] = "Период: месячная"

    ws.merge_cells('A4:F4')
    ws['A4'] = "Единица измерения: тыс. тенге"
    ws['A4'].alignment = center

    # ---------- Колонки ----------
    widths = {
        'A': 4, 'B': 4, 'C': 4, 'D': 4, 'E': 4, 'F': 4,
        'G': 30, 'H': 15, 'I': 15, 'J': 15, 'K': 15,
        'L': 15, 'M': 15, 'N': 15, 'O': 15, 'P': 15, 'Q': 15
    }

    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    ws.row_dimensions[5].height = 40
    ws.row_dimensions[6].height = 40

    # ---------- Заголовок таблицы ----------
    headers = {
        'A5:F6': "Коды бюджетной классификации",
        'G5:G6': "Наименование",
        'H5:H6': "Утвержденный бюджет на отчетный финансовый год",
        'I5:I6': "Уточненный бюджет на отчетный финансовый год",
        'J5:J6': "Скорректированный бюджет на отчетный финансовый год",
        'K5:L5': "Сводный план поступлений и финансирования",
        'M5:M6': "Принятые обязательства",
        'N5:N6': "Неоплаченные обязательства",
        'O5:O6': "Исполнение поступлений бюджета",
        'P5:P6': "Исп-е поступлений к свод. плану, %",
        'Q5:Q6': "Исп-е поступлений к бюджету, %"
    }

    for cells, text in headers.items():
        ws.merge_cells(cells)
        c1 = cells.split(':')[0]
        ws[c1] = text
        ws[c1].font = Font(name='Times New Roman', bold=True, size=9)
        ws[c1].alignment = center

    # ---------- Нумерация колонок ----------
    for i, col in enumerate(
        ['A', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q'],
        start=1
    ):
        ws[f"{col}7"] = str(i)

    # ---------- Конфигурация ----------
    FIELD_INDEX = {
        "utv": 8,
        "utch": 9,
        "plg": 7,
        "plgp": 11,
        "sumrg": 6
    }

    FIELD_COL = {
        "utv": 'H',
        "utch": 'I',
        "plg": 'J',
        "plgp": 'K',
        "sumrg": 'O'
    }

    SECTIONS = {
        1: "I. Доходы",
        2: "II. Затраты",
        3: "III. ЧИСТОЕ БЮДЖЕТНОЕ КРЕДИТОВАНИЕ",
        4: "IV. САЛЬДО ПО ОПЕРАЦИЯМ С ФИНАНСОВЫМИ АКТИВАМИ",
        5: "V. ДЕФИЦИТ (ПРОФИЦИТ) БЮДЖЕТА",
        6: "VI. ФИНАНСИРОВАНИЕ ДЕФИЦИТА (ИСПОЛЬЗОВАНИЕ ПРОФИЦИТА)"
    }

    # ---------- Генерация таблицы ----------
    for field, idx in FIELD_INDEX.items():
        col_num = openpyxl.utils.column_index_from_string(FIELD_COL[field])
        row = 8

        for grb_id, title in SECTIONS.items():
            ws.cell(row=row, column=7, value=title).font = font_main
            row += 1

            agg, total = aggregate_data(data, grb_id, idx)

            row = write_section(
                ws, row, agg, total, col_num, font_main, font_small
            )

    print(f"✅ {sheet_name} успешно добавлен ({len(data)} строк)")