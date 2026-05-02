import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
import openpyxl


# ─────────────────────── Dynamic table / date helpers ───────────────────────

def _get_latest_table(conn):
    """Return the most recent emfreportbpr_YYYYMM table, or fall back to emfreportbpr_v9."""
    cur = conn.cursor()
    # Prefer monthly tables (emfreportbpr_202603 etc.)
    cur.execute("""
        SELECT table_name FROM information_schema.tables
        WHERE table_schema = 'public'
          AND table_name ~ '^emfreportbpr_[0-9]{6}$'
        ORDER BY table_name DESC LIMIT 1
    """)
    row = cur.fetchone()
    if row:
        cur.close()
        return row[0]
    # Fall back to the versioned view (remote server)
    cur.execute("""
        SELECT table_name FROM information_schema.tables
        WHERE table_schema = 'public' AND table_name = 'emfreportbpr_v9'
        UNION ALL
        SELECT table_name FROM information_schema.views
        WHERE table_schema = 'public' AND table_name = 'emfreportbpr_v9'
        LIMIT 1
    """)
    row = cur.fetchone()
    cur.close()
    return row[0] if row else None


def _get_latest_datereport(conn, table):
    """Return the latest datereport value in the given table."""
    cur = conn.cursor()
    cur.execute(f"SELECT MAX(datereport) FROM {table}")
    row = cur.fetchone()
    cur.close()
    return row[0] if row else None


# ─────────────────────── Data fetch ───────────────────────────────────────

def fetch_data(conn, kato_pattern, datereport, table):
    cur = conn.cursor()
    query = f"""
        SELECT
            grbId,
            katFgr,
            katFgrName,
            klsFpgrName,
            pklAbpName,
            spkPrgName,
            SUM(sumrg)  AS sumrg,
            SUM(plg)    AS plg,
            SUM(utv)    AS utv,
            SUM(utch)   AS utch,
            SUM(plgp)   AS plgp,
            SUM(plgo)   AS plgo
        FROM {table}
        WHERE datereport = %s
          AND CAST(kato AS TEXT) LIKE %s
          AND (
              (grbid = 1 AND (
                  (katfgr = 1 AND strid = 1) OR
                  (katfgr = 2 AND strid = 2 AND adtype IS NULL) OR
                  (katfgr = 3 AND strid = 3) OR
                  (katfgr = 4 AND strid = 4 AND adtype NOT IN (3,4))
              ))
              OR
              (grbid = 2 AND (
                  (katfgr = 1  AND adtype IS NULL) OR
                  (katfgr = 2) OR
                  (katfgr = 3) OR
                  (katfgr = 4) OR
                  (katfgr = 5  AND (adtype IS NULL OR adtype NOT IN (3, 4))) OR
                  (katfgr = 6  AND adtype IS NULL) OR
                  (katfgr = 7  AND adtype IS NULL) OR
                  (katfgr = 8  AND adtype IS NULL) OR
                  (katfgr = 9) OR
                  (katfgr = 10 AND adtype IS NULL) OR
                  (katfgr = 11 AND adtype IS NULL) OR
                  (katfgr = 12 AND adtype IS NULL) OR
                  (katfgr = 13) OR
                  (katfgr = 14 AND (adtype IS NULL OR adtype NOT IN (3, 4))) OR
                  (katfgr = 15 AND adtype NOT IN (3, 4))
              ))
              OR grbid IN (3, 4, 5, 6)
          )
        GROUP BY grbId, katFgr, katFgrName, klsFpgrName, pklAbpName, spkPrgName
        ORDER BY grbId, katFgr, klsFpgrName, pklAbpName, spkPrgName
    """
    cur.execute(query, (datereport, kato_pattern))
    data = cur.fetchall()
    cur.close()
    return data


# ─────────────────────── Aggregation ──────────────────────────────────────

# Field positions in the SELECT result (0-based):
# 0=grbId, 1=katFgr, 2=katFgrName, 3=klsFpgrName, 4=pklAbpName,
# 5=spkPrgName, 6=sumrg, 7=plg, 8=utv, 9=utch, 10=plgp, 11=plgo

FIELD_INDEX = {
    "utv":   8,
    "utch":  9,
    "plg":   7,
    "plgp": 10,
    "plgo": 11,
    "sumrg": 6,
}

FIELD_COL = {
    "utv":   'H',
    "utch":  'I',
    "plg":   'J',
    "plgp":  'K',
    "plgo":  'L',
    "sumrg": 'O',
}

SECTIONS = {
    1: "I. Доходы",
    2: "II. Затраты",
    3: "III. Чистое бюджетное кредитование",
    4: "IV. Сальдо по операциям с финансовыми активами",
    5: "V. Дефицит (профицит) бюджета",
    6: "VI. Финансирование дефицита (использование профицита) бюджета",
}


def aggregate_data(data, grb_id, field_index):
    agg, total = {}, 0.0

    for r in data:
        if r[0] != grb_id:
            continue

        val = r[field_index]
        try:
            v = float(val) / 1000 if val is not None else 0.0
        except (TypeError, ValueError):
            v = 0.0

        total += v

        _, katFgr, katFgrName, klsFpgrName, pklAbpName, spkPrgName, *_ = r

        klsFpgrName = klsFpgrName or "Без группы"
        pklAbpName  = pklAbpName  or "Без подгруппы"
        spkPrgName  = spkPrgName  or "Без программы"

        fgr = agg.setdefault(katFgr, {"name": katFgrName, "total": 0.0, "lvl1": {}})
        l1  = fgr["lvl1"].setdefault(klsFpgrName, {"total": 0.0, "lvl2": {}})
        l2  = l1["lvl2"].setdefault(pklAbpName,   {"total": 0.0, "prog": {}})

        l2["prog"][spkPrgName] = l2["prog"].get(spkPrgName, 0.0) + v
        l2["total"] += v
        l1["total"] += v
        fgr["total"] += v

    return agg, total


# ─────────────────────── Excel writer ─────────────────────────────────────

def write_section(ws, start_row, agg, total, col, font_main, font_small):
    ws.cell(row=start_row - 1, column=col, value=round(total, 2)).font = font_main
    row = start_row

    for fgr in sorted(agg.keys(), key=lambda x: (x is None, x)):
        d = agg[fgr]
        ws.cell(row=row, column=7, value=d["name"].upper()).font = font_main
        ws.cell(row=row, column=col, value=round(d["total"], 2)).font = font_main
        row += 1

        for l1n in sorted(d["lvl1"].keys()):
            l1d = d["lvl1"][l1n]
            ws.cell(row=row, column=7, value=l1n).font = font_small
            ws.cell(row=row, column=col, value=round(l1d["total"], 2)).font = font_small
            row += 1

            for l2n in sorted(l1d["lvl2"].keys()):
                l2d = l1d["lvl2"][l2n]
                ws.cell(row=row, column=7, value=f"    {l2n}").font = font_small
                ws.cell(row=row, column=col, value=round(l2d["total"], 2)).font = font_small
                row += 1

                for pn in sorted(l2d["prog"].keys()):
                    ws.cell(row=row, column=7, value=f"      – {pn}").font = font_small
                    ws.cell(row=row, column=col, value=round(l2d["prog"][pn], 2)).font = font_small
                    row += 1

    return row


# ─────────────────────── Main ──────────────────────────────────────────────

def create_table_in_excel(conn, wb, kato_pattern, sheet_name, datereport=None):
    """
    Build one Excel sheet for the given kato_pattern.
    If datereport is None, auto-detects the latest available date.
    """
    table = _get_latest_table(conn)
    if not table:
        print(f"⚠️ Не найдена таблица emfreportbpr_* в БД")
        return

    if datereport is None:
        datereport = _get_latest_datereport(conn, table)
    if not datereport:
        print(f"⚠️ Нет данных в таблице {table}")
        return

    data = fetch_data(conn, kato_pattern, datereport, table)

    if not data:
        print(f"⚠️ Нет данных для {sheet_name} (kato={kato_pattern}, date={datereport})")
        return

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    font_main  = Font(name='Times New Roman', bold=True, size=9)
    font_small = Font(name='Times New Roman', size=8)
    center     = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Title ──
    ws.merge_cells('A1:Q1')
    ws['A1'] = f"Отчет об исполнении бюджета на {datereport}"
    ws['A1'].font = Font(name='Times New Roman', bold=True, size=12)
    ws['A1'].alignment = center

    ws.merge_cells('A3:F3')
    ws['A3'] = "Период: месячная"
    ws['A3'].font = font_main

    ws.merge_cells('A4:F4')
    ws['A4'] = "Единица измерения: тыс. тенге"
    ws['A4'].font = font_main
    ws['A4'].alignment = center

    # ── Column widths ──
    widths = {'A':4,'B':4,'C':4,'D':4,'E':4,'F':4,'G':40,
              'H':15,'I':15,'J':15,'K':15,'L':15,'M':15,'N':15,'O':15,'P':15,'Q':15}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.row_dimensions[5].height = 50
    ws.row_dimensions[6].height = 30

    # ── Header row ──
    col_headers = {
        'A5:F6': "Коды бюджетной классификации",
        'G5:G6': "Наименование",
        'H5:H6': "Утвержденный бюджет на отчетный финансовый год",
        'I5:I6': "Уточненный бюджет на отчетный финансовый год",
        'J5:J6': "Скорректированный бюджет на отчетный финансовый год",
        'K5:L5': "Сводный план поступлений и финансирования по платежам/обязательствам",
        'M5:M6': "Принятые обязательства",
        'N5:N6': "Неоплаченные обязательства",
        'O5:O6': "Исполнение поступлений бюджета и/или оплаченных обязательств",
        'P5:P6': "Исп-е к своду. плану, %",
        'Q5:Q6': "Исп-е к бюджету, %",
    }
    for rng, text in col_headers.items():
        ws.merge_cells(rng)
        c0 = rng.split(':')[0]
        ws[c0] = text
        ws[c0].font = font_main
        ws[c0].alignment = center

    ws['K6'] = "по платежам"
    ws['K6'].font = font_main
    ws['K6'].alignment = center
    ws['L6'] = "по обязательствам"
    ws['L6'].font = font_main
    ws['L6'].alignment = center

    # Column numbering row
    for i, col in enumerate(['A','G','H','I','J','K','L','M','N','O','P','Q'], 1):
        ws[f"{col}7"] = str(i)

    # ── Data ──
    for field, idx in FIELD_INDEX.items():
        col_num = openpyxl.utils.column_index_from_string(FIELD_COL[field])
        row = 8

        for grb_id, title in SECTIONS.items():
            ws.cell(row=row, column=7, value=title).font = font_main
            row += 1
            agg, total = aggregate_data(data, grb_id, idx)
            row = write_section(ws, row, agg, total, col_num, font_main, font_small)

    print(f"✅ {sheet_name} готов ({len(data)} строк, таблица={table}, дата={datereport})")
