"""
BNS АСПР indicator config + Excel report builder.

Each menu button maps to one or more Taldau index IDs. The builder fetches
data for all sub-indicators in parallel and writes a single Excel workbook.
"""
import concurrent.futures
import datetime
import re
from io import BytesIO
from collections import OrderedDict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

import bns_taldau_loader as taldau

matplotlib.rcParams["font.family"] = "DejaVu Sans"
matplotlib.rcParams["axes.unicode_minus"] = False


# ──────────────── Indicator catalog (callback_data → config) ────────────────

INDICATORS = {
    "bns_profit": {
        "label": "Уровень рентабельности крупных & средних предприятий",
        "ids": [(703165, "Рентабельность (убыточность)")],
    },
    "bns_gdp_ifo": {
        "label": "Динамика ИФО ВВП",
        "ids": [
            (2979005, "ИФО ВВП методом производства"),
            (700974,  "ИФО ВВП методом конечного использования"),
        ],
    },
    "bns_labor": {
        "label": "Основные показатели рынка труда",
        "ids": [
            (702944, "Уровень безработицы"),
            (702943, "Безработное население (15+)"),
            (702841, "Уровень занятости в эк. активном населении"),
        ],
    },
    "bns_cars": {
        "label": "Продажа и постановка на учет авто",
        "ids": [
            (702228, "Парк ТС, состоящих на учете"),
            (702233, "Число прибывших ТС"),
        ],
    },
    "bns_trade": {
        "label": "Внешняя торговля Казахстана",
        "ids": [
            (701054, "Импорт товаров (млн USD)"),
            (701074, "Экспорт товаров (млн USD)"),
        ],
    },
    "bns_inflation_base": {
        "label": "Показатели базовой инфляции",
        "ids": [
            (703082,  "Базовый ИПЦ"),
            (55056856, "Базовый ИПЦ без 3 составляющих"),
            (55056857, "Базовый ИПЦ без 7 составляющих"),
        ],
    },
    "bns_inflation": {
        "label": "Месячная инфляция Казахстана",
        "ids": [(703076, "Индекс потребительских цен")],
    },
}


# ──────────────── Period discovery ────────────────

def available_periods(callback: str) -> list[dict]:
    """Union of available periods across all sub-indicators of a button."""
    cfg = INDICATORS[callback]
    seen = OrderedDict()
    for index_id, _ in cfg["ids"]:
        for p in taldau.get_period_list(index_id):
            seen.setdefault(p["id"], p["name"])
    return [{"id": pid, "name": name} for pid, name in seen.items()]


def available_years(callback: str, period_id: int) -> list[int]:
    """Union of years that have data for given period across all sub-indicators."""
    cfg = INDICATORS[callback]
    years = set()
    for index_id, _ in cfg["ids"]:
        try:
            data = taldau.fetch_indicator(index_id, period_id)
        except Exception:
            continue
        for label, ds in data["dates"]:
            m = re.search(r"\b(\d{4})\b", label) or re.search(r"\.(\d{4})$", ds)
            if m:
                years.add(int(m.group(1)))
    return sorted(years)


# ──────────────── Data fetch (parallel) ────────────────

def fetch_all(callback: str, period_id: int) -> list[dict]:
    """Fetch each sub-indicator in parallel."""
    cfg = INDICATORS[callback]
    out = [None] * len(cfg["ids"])
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(cfg["ids"])) as ex:
        futures = {
            ex.submit(taldau.fetch_indicator, idx, period_id): i
            for i, (idx, _) in enumerate(cfg["ids"])
        }
        for f in concurrent.futures.as_completed(futures):
            i = futures[f]
            try:
                res = f.result()
                res["display_name"] = cfg["ids"][i][1]
                out[i] = res
            except Exception as e:
                out[i] = {"display_name": cfg["ids"][i][1],
                          "name": "", "unit": "", "dates": [], "values": [],
                          "error": str(e)}
    return out


# ──────────────── Excel writer ────────────────

def build_excel(callback: str, period_id: int, year: int | None,
                file_path: str) -> str:
    """
    Build Excel report. If `year` is None → all years; otherwise filter to that year.
    Returns the file_path on success.
    """
    cfg = INDICATORS[callback]
    period_name = taldau.PERIOD_NAMES.get(period_id, f"period_{period_id}")
    datasets = fetch_all(callback, period_id)

    if year is not None:
        for d in datasets:
            d["dates"], d["values"] = taldau.filter_by_year(d["dates"], d["values"], year)

    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    title = cfg["label"]
    year_part = f" — {year} г." if year else ""
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=1 + len(datasets))
    ws["A1"] = f"{title} ({period_name}){year_part}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = center

    units = " / ".join(sorted({d["unit"] for d in datasets if d["unit"]})) or ""
    if units:
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=1 + len(datasets))
        ws["A2"] = f"Единица измерения: {units}"
        ws["A2"].alignment = center

    header = ["Период"] + [d["display_name"] for d in datasets]
    ws.append([])  # blank row 3
    ws.append(header)
    for cell in ws[4]:
        cell.font = bold
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 35

    all_labels = []
    seen = set()
    for d in datasets:
        for (label, ds) in d["dates"]:
            key = (label, ds)
            if key not in seen:
                seen.add(key)
                all_labels.append(key)

    def _date_key(item):
        label, ds = item
        try:
            return datetime.datetime.strptime(ds, "%d.%m.%Y")
        except Exception:
            return datetime.datetime.min
    all_labels.sort(key=_date_key)

    lookup = []
    for d in datasets:
        lookup.append({k: v for k, v in zip([(l, ds) for l, ds in d["dates"]],
                                            d["values"])})

    for label, ds in all_labels:
        row = [label]
        for li in lookup:
            v = li.get((label, ds))
            row.append(v if v is not None else "")
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(header) + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if c > 1:
                cell.alignment = right
                cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 28
    for i, d in enumerate(datasets, start=2):
        ws.column_dimensions[get_column_letter(i)].width = max(15, len(d["display_name"]) // 2 + 8)

    if any(d["values"] for d in datasets):
        ws_chart = wb.create_sheet("График")
        try:
            x_labels = [lbl for (lbl, _) in all_labels]
            fig, ax = plt.subplots(figsize=(13, 6))
            for d, li in zip(datasets, lookup):
                ys = [li.get((label, ds)) for (label, ds) in all_labels]
                ax.plot(range(len(x_labels)), ys, marker="o", linewidth=1.5,
                        label=d["display_name"][:60])
            step = max(1, len(x_labels) // 20)
            ax.set_xticks(range(0, len(x_labels), step))
            ax.set_xticklabels([x_labels[i] for i in range(0, len(x_labels), step)],
                               rotation=35, ha="right", fontsize=8)
            ax.set_ylabel(units or "")
            ax.set_title(f"{title} ({period_name}){year_part}",
                         fontweight="bold", pad=15)
            ax.legend(fontsize=8, loc="best")
            ax.grid(axis="y", linestyle="--", alpha=0.3)
            plt.figtext(0.01, -0.02, "Источник: Бюро национальной статистики (taldau.stat.gov.kz)",
                        fontsize=8, color="gray")
            plt.figtext(0.85, -0.02, "TENGENOMIKA",
                        fontsize=9, fontweight="bold", color="#003366")
            plt.tight_layout()
            buf = BytesIO()
            plt.savefig(buf, format="png", dpi=180, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            ws_chart.add_image(Image(buf), "A1")
        except Exception:
            pass

    wb.save(file_path)
    return file_path
