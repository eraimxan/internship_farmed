import requests
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter, MonthLocator
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import matplotlib

# Настройки кириллицы
matplotlib.rcParams["font.family"] = "DejaVu Sans"
matplotlib.rcParams["axes.unicode_minus"] = False

BASE_URL = "https://data.nationalbank.kz/api/report/stat-data"  # можно поменять на stat-data-locale
FORM_ID = 1


def _safe_float(x):
    try:
        return float(x)
    except Exception:
        return None


def _extract_group_from_item(item: dict):
    """
    Возвращает (code, name) из item.
    1) Пытаемся через attributes (если они есть)
    2) Если attributes нет — ищем подходящие поля в самом item
    3) Если ничего не нашли — TOTAL
    """

    # 1) attributes (может быть None)
    attrs = item.get("attributes") or []
    if isinstance(attrs, list) and attrs:
        # сначала ищем account_group (как у тебя)
        for a in attrs:
            if a.get("attrCode") == "account_group":
                code = a.get("valueCode")
                name = a.get("valueNameRu") or a.get("valueNameEn") or ""
                if code:
                    return str(code), str(name)

        # если account_group нет — просто возьмём первый атрибут как идентификатор
        a0 = attrs[0]
        code = a0.get("valueCode") or a0.get("valueNameRu") or a0.get("valueNameEn")
        name = a0.get("valueNameRu") or a0.get("valueNameEn") or str(code or "")
        if code:
            return str(code), str(name)

    # 2) fallback по полям самой записи (у разных форм разные ключи)
    # пробуем "кодовые" поля
    code_candidates = [
        "code", "rowCode", "indicatorCode", "accountCode", "groupCode", "itemCode"
    ]
    name_candidates = [
        "nameRu", "rowNameRu", "indicatorNameRu", "accountNameRu", "groupNameRu", "itemNameRu",
        "nameKz", "nameEn"
    ]

    code = None
    for k in code_candidates:
        v = item.get(k)
        if v not in (None, ""):
            code = v
            break

    name = None
    for k in name_candidates:
        v = item.get(k)
        if v not in (None, ""):
            name = v
            break

    # иногда в item есть только "type" — используем как имя (но не как код группы счетов)
    if not name and item.get("type"):
        name = item.get("type")

    if code:
        return str(code), str(name or "")

    # 3) крайний fallback — всё в одну группу
    return "TOTAL", str(name or "Итого")


def load_bvu_balance_data(year: int) -> pd.DataFrame:
    """
    Загружает данные за конкретный год (так стабильнее и быстрее),
    приводит к DataFrame и гарантирует непустой rows (если API вернул хоть что-то).
    """
    year = int(year)
    date_from = f"{year}-01-01"
    date_to = f"{year}-12-31"

    params = {
        "form_id": FORM_ID,
        "date_from": date_from,
        "date_to": date_to,
        # если переключишь на stat-data-locale, добавь "lang": "ru"
    }

    r = requests.get(BASE_URL, params=params, timeout=(10, 60), headers={"Accept": "application/json"})
    r.raise_for_status()
    data = r.json() or []

    rows = []
    for item in data:
        report_date = item.get("reportDate")
        amount = _safe_float(item.get("amount"))

        if not report_date or amount is None:
            continue

        code, name = _extract_group_from_item(item)

        rows.append(
            {
                "Дата": pd.to_datetime(report_date, errors="coerce"),
                "Код": code,
                "Наименование": name,
                "Сумма_тыс_тенге": amount,
            }
        )

    df = pd.DataFrame(rows)
    df = df.dropna(subset=["Дата"])
    return df


def build_excel_bvu_balance_with_chart(filename: str, year: int) -> str:
    """
    Создает Excel строго за выбранный год (year):
    - Лист 1: сводная таблица (строки: группы, колонки: месяцы)
    - Лист 2: график (общая сумма по всем группам по месяцам)
    """

    year = int(year)
    df = load_bvu_balance_data(year)

    if df.empty:
        raise ValueError(f"Нет данных из API за {year} год (проверь form_id={FORM_ID}).")

    # Месяц = первое число месяца
    df["Месяц"] = df["Дата"].dt.to_period("M").dt.to_timestamp()

    # На случай дублей по одному коду/месяцу — суммируем
    grouped = (
        df.groupby(["Код", "Наименование", "Месяц"], as_index=False)["Сумма_тыс_тенге"]
          .sum()
    )

    # Pivot: строки = (Код, Наименование), колонки = Месяц
    pivot = grouped.pivot_table(
        index=["Код", "Наименование"],
        columns="Месяц",
        values="Сумма_тыс_тенге",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()

    # Сортировка по коду (численно, где возможно)
    def _try_int(x: str):
        try:
            return int(x)
        except Exception:
            return 10**12

    pivot = pivot.sort_values(by="Код", key=lambda s: s.map(_try_int))

    month_cols = sorted([c for c in pivot.columns if isinstance(c, pd.Timestamp)])

    # --------- Excel ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводный баланс"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(month_cols))
    ws["A1"] = f"СВОДНЫЙ БАЛАНС ПО БВУ РК ({year} г.)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    header = ["Код", "Наименование"]
    header += [f"на {m.strftime('%d.%m.%Y')} г." for m in month_cols]
    ws.append(header)

    for _, row in pivot.iterrows():
        ws.append([row["Код"], row["Наименование"]] + [float(row[m]) for m in month_cols])

    # Итоговая строка
    total_row_idx = ws.max_row + 1
    ws.cell(row=total_row_idx, column=1, value="ИТОГО").font = Font(bold=True)

    for j in range(3, 3 + len(month_cols)):
        col_letter = get_column_letter(j)
        cell = ws.cell(row=total_row_idx, column=j,
                       value=f"=SUM({col_letter}3:{col_letter}{total_row_idx - 1})")
        cell.font = Font(bold=True)

    # Оформление
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    for j in range(3, 3 + len(month_cols)):
        ws.column_dimensions[get_column_letter(j)].width = 18

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2 + len(month_cols)):
        for c in r:
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in ws[2]:
        c.font = Font(bold=True)

    for r in range(3, total_row_idx + 1):
        for j in range(3, 3 + len(month_cols)):
            ws.cell(row=r, column=j).number_format = "#,##0"

    # --------- График ----------
    total_series = (
        grouped.groupby("Месяц", as_index=False)["Сумма_тыс_тенге"]
               .sum()
               .sort_values("Месяц")
    )

    ws_chart = wb.create_sheet("График")

    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(
        total_series["Месяц"],
        total_series["Сумма_тыс_тенге"],
        linewidth=2.5,
        marker="o",
        label="Итого (тыс. тг.)",
    )

    ax.set_title(f"Сводный баланс по БВУ РК — Итого ({year} г.)", fontsize=12, fontweight="bold", pad=18)
    ax.set_ylabel("тыс. тенге", fontsize=10)
    ax.grid(True, linestyle="--", alpha=0.5)

    ax.xaxis.set_major_locator(MonthLocator(interval=1))
    ax.xaxis.set_major_formatter(DateFormatter("%b.%y"))
    plt.xticks(rotation=45, ha="right")

    ax.legend(loc="upper left", fontsize=9)
    plt.figtext(0.01, -0.02, "Источник: НБ РК", fontsize=9, color="gray")
    plt.figtext(0.87, -0.02, "TENGENOMIKA", fontsize=9, fontweight="bold")

    plt.tight_layout()

    img_data = BytesIO()
    plt.savefig(img_data, format="png", dpi=200, bbox_inches="tight")
    img_data.seek(0)
    plt.close(fig)

    img = XLImage(img_data)
    ws_chart.add_image(img, "A1")

    wb.save(filename)
    return filename
