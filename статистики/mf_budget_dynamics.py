import concurrent.futures
import datetime
import requests
import urllib3
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

API_URL = "https://ows.goszakup.gov.kz/v3/graphql"
HEADERS = {
    "Authorization": "Bearer 8530636543c42453dc78b894f3c32ab6",
    "Content-Type": "application/json",
}

GRAPHQL_QUERY = """
query EmfReportBpr($after: Int, $date: String!) {
  EmfReportBpr(limit: 200, after: $after, filter: { dateReport: $date }) {
    grbId, katFgr, katFgrName,
    klsFpgr, klsFpgrName, pklAbp, pklAbpName,
    spkPrg, spkPrgName,
    adtype, plg, sumrg
  }
}
"""

THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Annual date range: Jan-01 of each year = execution of previous year
def _build_annual_dates():
    current_year = datetime.date.today().year
    return [f"{year}-01-01" for year in range(2016, current_year + 2)]

ANNUAL_DATES  = _build_annual_dates()
ANNUAL_LABELS = [str(int(d[:4]) - 1) for d in ANNUAL_DATES]


def _get_latest_candidates(n=4):
    """Generate n month-start dates going back from last month, excluding annual dates."""
    today = datetime.date.today()
    y, m = today.year, today.month - 1  # skip current (likely incomplete)
    if m == 0:
        m, y = 12, y - 1
    candidates = []
    for _ in range(n):
        s = f"{y}-{m:02d}-01"
        if s not in ANNUAL_DATES:
            candidates.append(s)
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return candidates


# ─────────────────────────── API fetch ───────────────────────────

def _fetch_all(date_str: str) -> list:
    import time
    rows = []
    after = None
    while True:
        payload = {
            "query": GRAPHQL_QUERY,
            "variables": {"after": after, "date": date_str},
        }
        for attempt in range(4):
            try:
                r = requests.post(API_URL, json=payload, headers=HEADERS,
                                  verify=False, timeout=120)
                r.raise_for_status()
                break
            except requests.exceptions.Timeout:
                if attempt == 3:
                    raise
                time.sleep(5 * (attempt + 1))
        data = r.json()
        batch = data.get("data", {}).get("EmfReportBpr", []) or []
        if not batch:
            break
        rows.extend(batch)
        page_info = (data.get("extensions") or {}).get("pageInfo") or {}
        last_id = page_info.get("lastId")
        if not page_info.get("hasNextPage") or not last_id:
            break
        after = last_id
    return rows


def _find_latest_month() -> tuple:
    """Try dynamically generated month candidates; return first one with expense data."""
    for candidate in _get_latest_candidates():
        rows = _fetch_all(candidate)
        if any(_is_expense(r) for r in rows):
            month, year = candidate[5:7], candidate[:4]
            label = f"{month}.{year}"
            return candidate, label, rows
    return None, None, []


# ─────────────────────────── filtering ───────────────────────────

def _safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _is_expense(r: dict) -> bool:
    """grbId=2 expense rows with the standard budget classification filter."""
    grbid  = _safe_int(r.get("grbId"))
    katfgr = _safe_int(r.get("katFgr"))
    adtype = _safe_int(r.get("adtype"))

    if grbid != 2:
        return False

    return (
        (katfgr == 1  and adtype is None) or
        (katfgr == 2) or
        (katfgr == 3) or
        (katfgr == 4) or
        (katfgr == 5  and (adtype is None or adtype not in (3, 4))) or
        (katfgr == 6  and adtype is None) or
        (katfgr == 7  and adtype is None) or
        (katfgr == 8  and adtype is None) or
        (katfgr == 9) or
        (katfgr == 10 and adtype is None) or
        (katfgr == 11 and adtype is None) or
        (katfgr == 12 and adtype is None) or
        (katfgr == 13) or
        (katfgr == 14 and (adtype is None or adtype not in (3, 4))) or
        (katfgr == 15 and adtype is not None and adtype not in (3, 4))
    )


# ─────────────────────────── aggregation ───────────────────────────

def _to_mln(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ".")) / 1_000_000
    except Exception:
        return 0.0


def _aggregate(rows: list, period_count: int, period_idx: int,
               tree: dict, totals: dict) -> None:
    """Accumulate plg + sumrg for a given period index into shared tree."""
    for r in rows:
        if not _is_expense(r):
            continue

        plg   = _to_mln(r.get("plg"))
        exec_ = _to_mln(r.get("sumrg"))

        grb        = _safe_int(r.get("grbId")) or 2
        katFgr     = _safe_int(r.get("katFgr")) or 0
        katFgrName = r.get("katFgrName") or ""
        klsName    = r.get("klsFpgrName") or ""
        pklName    = r.get("pklAbpName") or ""
        spkName    = r.get("spkPrgName") or ""

        # Key includes katFgr number so categories can be sorted correctly later.
        # Totals are computed bottom-up from the tree after all data is loaded
        # to avoid double-counting when the API returns both aggregate and detail rows.
        node = tree[grb][(katFgr, katFgrName)][klsName][pklName][spkName]
        node["plg"][period_idx]  += plg
        node["exec"][period_idx] += exec_


def _compute_totals_from_tree(tree: dict, totals: dict, period_count: int) -> None:
    """Fill totals by summing leaf nodes of the tree (avoids double-counting)."""
    for grb, kat_map in tree.items():
        for kls_map in kat_map.values():
            for pkl_map in kls_map.values():
                for spk_map in pkl_map.values():
                    for spk_node in spk_map.values():
                        for i in range(period_count):
                            totals[grb]["plg"][i]  += spk_node["plg"][i]
                            totals[grb]["exec"][i] += spk_node["exec"][i]


# ─────────────────────────── Excel builder ───────────────────────────

def create_table_in_excel(db_params, file_path, oblast_name=""):
    """
    Builds the dynamics Excel report by pulling data from the API.
    `db_params` is accepted for interface compatibility but ignored.
    """
    # Build date list: annual + optional latest month
    latest_date_str, latest_label, latest_rows = _find_latest_month()

    date_points = ANNUAL_DATES[:]
    labels      = ANNUAL_LABELS[:]

    if latest_date_str:
        date_points.append(latest_date_str)
        labels.append(latest_label)

    period_count = len(labels)

    def new_node():
        return {"plg": [0.0] * period_count, "exec": [0.0] * period_count}

    tree   = defaultdict(lambda: defaultdict(lambda: defaultdict(
             lambda: defaultdict(lambda: defaultdict(new_node)))))
    totals = defaultdict(new_node)

    # Fetch all annual dates in parallel, then aggregate sequentially
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(ANNUAL_DATES)) as executor:
        indexed_futures = [
            (idx, executor.submit(_fetch_all, date_str))
            for idx, date_str in enumerate(ANNUAL_DATES)
        ]

    for idx, future in indexed_futures:
        _aggregate(future.result(), period_count, idx, tree, totals)

    # Latest month (already fetched)
    if latest_date_str:
        _aggregate(latest_rows, period_count, len(ANNUAL_DATES), tree, totals)

    # Compute section totals bottom-up so they always equal the sum of categories
    _compute_totals_from_tree(tree, totals, period_count)

    # ── Build Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    font_h = Font(name="Times New Roman", bold=True, size=9)
    font_i = Font(name="Times New Roman", italic=True, size=7)
    font_n = Font(name="Times New Roman", size=7)

    name_col        = 7
    start_period_col = 8

    ws.freeze_panes = ws.cell(row=8, column=start_period_col)
    ws.column_dimensions["G"].width = 45
    for i in range(2 * period_count):
        ws.column_dimensions[get_column_letter(start_period_col + i)].width = 18

    max_col = start_period_col + 2 * period_count - 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = "Отчет об исполнении бюджета"
    ws["A1"].alignment = CENTER
    ws["A1"].font = Font(name="Times New Roman", bold=True, size=11)

    ws["G2"] = oblast_name
    ws["G2"].alignment = CENTER
    ws["G3"] = "млн. тенге"
    ws["G3"].alignment = CENTER

    r_year = 4
    r_sub  = 5

    for idx, label in enumerate(labels):
        c1 = start_period_col + idx * 2
        c2 = c1 + 1
        ws.merge_cells(start_row=r_year, start_column=c1, end_row=r_year, end_column=c2)
        ws.cell(row=r_year, column=c1).value = label
        ws.cell(row=r_year, column=c1).alignment = CENTER
        ws.cell(row=r_year, column=c1).font = font_h
        ws.cell(row=r_sub, column=c1).value = "Скорректированный бюджет"
        ws.cell(row=r_sub, column=c2).value = "Исполнение бюджета"
        ws.cell(row=r_sub, column=c1).alignment = CENTER
        ws.cell(row=r_sub, column=c2).alignment = CENTER
        ws.cell(row=r_sub, column=c1).font = font_h
        ws.cell(row=r_sub, column=c2).font = font_h

    ws.merge_cells(start_row=r_year, start_column=name_col,
                   end_row=r_sub, end_column=name_col)
    ws.cell(row=r_year, column=name_col).value = "Наименование"
    ws.cell(row=r_year, column=name_col).alignment = CENTER
    ws.cell(row=r_year, column=name_col).font = font_h

    ws.cell(row=7, column=name_col).value = "2"
    for i in range(2 * period_count):
        ws.cell(row=7, column=start_period_col + i).value = str(3 + i)

    for r in range(4, 8):
        for c in range(name_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN

    def write_line(row_idx, text, vals_node, font, indent=""):
        cell = ws.cell(row=row_idx, column=name_col)
        cell.value = (f"{indent}{text}".upper() if font.bold
                      else f"{indent}{text}")
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        for i in range(period_count):
            c_plg  = start_period_col + i * 2
            c_exec = c_plg + 1
            ws.cell(row=row_idx, column=c_plg).value  = round(vals_node["plg"][i],  2)
            ws.cell(row=row_idx, column=c_exec).value = round(vals_node["exec"][i], 2)
        return row_idx + 1

    sections = {2: "II. Затраты"}

    row_idx = 8
    for grbId, title in sections.items():
        row_idx = write_line(row_idx, title, totals[grbId], font_h)

        for (katFgr, katFgrName), kls_map in sorted(tree[grbId].items(), key=lambda x: x[0][0]):
            kat_sum = new_node()
            for pkl_map in kls_map.values():
                for spk_map in pkl_map.values():
                    for spk_node in spk_map.values():
                        for i in range(period_count):
                            kat_sum["plg"][i]  += spk_node["plg"][i]
                            kat_sum["exec"][i] += spk_node["exec"][i]
            row_idx = write_line(row_idx, katFgrName, kat_sum, font_h)

            for klsFpgrName, pkl_map in kls_map.items():
                kls_sum = new_node()
                for spk_map in pkl_map.values():
                    for spk_node in spk_map.values():
                        for i in range(period_count):
                            kls_sum["plg"][i]  += spk_node["plg"][i]
                            kls_sum["exec"][i] += spk_node["exec"][i]
                row_idx = write_line(row_idx, klsFpgrName, kls_sum, font_i)

                for pklAbpName, spk_map in pkl_map.items():
                    pkl_sum = new_node()
                    for spk_node in spk_map.values():
                        for i in range(period_count):
                            pkl_sum["plg"][i]  += spk_node["plg"][i]
                            pkl_sum["exec"][i] += spk_node["exec"][i]
                    row_idx = write_line(row_idx, pklAbpName, pkl_sum, font_n, indent="    ")

                    for spkPrgName, spk_node in spk_map.items():
                        row_idx = write_line(
                            row_idx, f"– {spkPrgName}", spk_node, font_n, indent="      "
                        )
        row_idx += 1

    for r in range(4, row_idx):
        for c in range(name_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN

    wb.save(file_path)
    return file_path
