import asyncio
import aiohttp
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from aiogram.fsm.state import State, StatesGroup
from openpyxl.utils import get_column_letter
import datetime


# FSM для выбора дат (если понадобится в основном боте)
class DateRange(StatesGroup):
    start_year = State()
    start_month = State()
    end_year = State()
    end_month = State()


# ---------------- Функция построения Excel ----------------
def build_excel(data, file_path="report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Депозиты"

    # Заголовки
    ws["A1"] = "Депозиты в депозитных организациях"
    ws["A2"] = "млн. тенге / на конец периода"
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Формируем подписи месяцев
    months = [pd.to_datetime(d["reporting_date"]).strftime("%m.%Y") for d in data]
    ws.append(["Категория"] + months)

    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Табличные строки
    rows = [
        ("Всего депозитов", "total_deposit", True),
        ("    в том числе:", "", False),
        ("в национальной валюте:", "total_deposit_nc", True),
        ("небанковских юридических лиц", "total_deposit_nc_nb_le", False),
        ("физических лиц", "total_deposit_nc_np", False),
        ("в иностранной валюте:", "total_deposit_fc", True),
        ("небанковских юридических лиц", "total_deposit_fc_nb_le", False),
        ("физических лиц", "total_deposit_fc_np", False),
    ]

    # Заполняем таблицу
    for row_title, key, bold in rows:
        values = [d.get(key, "") if key else "" for d in data]
        row = [row_title] + values
        ws.append(row)

        r = ws.max_row
        ws[f"A{r}"].font = Font(bold=bold)
        for c in range(2, len(values) + 2):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(horizontal="right")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '# ##0'

    # Автоширина
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)
    return file_path


# ---------------- Основная функция генерации отчёта ----------------
async def generate_report(start_year, start_month, end_year, end_month, file_name="Депозиты.xlsx"):
    url = "https://nationalbank.kz/ru/depositoryorganizationsdeposits/depozity-v-depozitnyh-organizaciyah-/records"

    async with aiohttp.ClientSession() as session:
        async with session.get(url) as resp:
            records = await resp.json()

    # фильтрация по датам
    start_date = datetime.date(start_year, start_month, 1)
    end_date = (datetime.date(end_year, end_month, 28).replace(day=1) +
                datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)

    filtered = []
    for r in records:
        d = pd.to_datetime(r["reporting_date"]).date()
        if start_date <= d <= end_date:
            filtered.append(r)

    if not filtered:
        return None

    file_path = build_excel(filtered, file_name)
    return file_path
