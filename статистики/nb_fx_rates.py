# fx_rates_33.py
import requests
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_TEMPLATE = (
    "https://data.nationalbank.kz/api/report/stat-data-locale"
    "?form_id=33&date_from={start}&date_to={end}&lang=ru"
)


MONTHS_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}
# Порядок столбцов с кварталами после каждого квартала
MONTH_COLS = [
    1, 2, 3, "Q1",
    4, 5, 6, "Q2",
    7, 8, 9, "Q3",
    10, 11, 12, "Q4",
    "Y",
]

# Какие валюты подсвечивать красным, как на скрине
HIGHLIGHT_CODES = {"USD", "EUR", "RUB", "CNY"}


def _fetch_year_data(year: int):
    start = f"{year}-02-01"
    end = f"{year + 1}-01-01"  # до 1 января следующего года
    url = API_TEMPLATE.format(start=start, end=end)
    r = requests.get(url, timeout=(10, 30))
    r.raise_for_status()
    data = r.json()
    # Если вернулся объект с "data", вытащим список
    if isinstance(data, dict):
        data = data.get("data") or data.get("records") or []
    return data


def _normalize_records(data, year: int):
    """
    Собираем значения по каждой валюте.
    Возвращает:
      rates[(code, name)][key] = value, где key ∈ {1..12, "Q1".."Q4", "Y"}
    """
    rates = defaultdict(dict)  # (code, name) -> {month|quarter|Y: value}

    for item in data:
        report_date = item.get("reportDate") or item.get("reporting_date")
        amount = item.get("amount")
        attrs = item.get("attributes", [])

        if amount is None or not report_date:
            continue

        code = name = periodicity = None
        period_code = None  # для квартала, может быть "1".."4"
        for a in attrs:
            if a.get("attrCode") == "currency":
                code = a.get("valueCode")
                name = a.get("valueNameRu")
            elif a.get("attrCode") == "periodicity":
                periodicity = a.get("valueCode")
            elif a.get("attrCode") in ("period", "quarter"):  # подхватим номер квартала
                period_code = a.get("valueCode")

        if not code:
            continue

        # Сдвиг на -1 месяц: 2025-11-01 -> фактически октябрь 2025; 
        # аналогично для квартальных/годовых дат.
        d = pd.to_datetime(report_date) - pd.DateOffset(months=1)

        try:
            val = float(amount)
        except Exception:
            continue

        # Записываем в нужный «ключ» строки и только для запрошенного года
        if periodicity == "monthly":
            if d.year != year:
                continue
            month = int(d.month)
            rates[(code, name)][month] = val

        elif periodicity == "quarterly":
            if d.year != year:
                continue
            # Определим квартал по сдвинутой дате (надёжно даже без period_code)
            qnum = (int(d.month) - 1) // 3 + 1  # 1..4
            rates[(code, name)][f"Q{qnum}"] = val

        elif periodicity == "yearly":
            # Годовая точка после сдвига относится к фактическому году d.year
            if d.year != year:
                continue
            rates[(code, name)]["Y"] = val

        else:
            # другие периодичности игнорируем
            continue

    return rates


def _quarter_avg(row_dict, m1, m2, m3):
    vals = [row_dict.get(m) for m in (m1, m2, m3) if row_dict.get(m) is not None]
    return round(sum(vals) / len(vals), 2) if vals else None

def _format_val(val: float):
    if val is None:
        return ""
    return f"{val:,.3f}".replace(",", "X").replace(".", ",").replace("X", " ")


def build_fx_rates_excel(year: int, filename: str) -> str:
    """
    Строит Excel «Официальные обменные курсы иностранных валют в {year} году»
    по API form_id=33.
    """
    data = _fetch_year_data(year)
    rates = _normalize_records(data, year)

    # Сортируем строки по русскому названию валюты
    keys_sorted = sorted(rates.keys(), key=lambda x: (x[1] or "", x[0] or ""))

    # --- Готовим книгу ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}"

    # Заголовок
    title = f"Официальные обменные курсы иностранных валют в {year} году"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(MONTH_COLS))
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Шапка таблицы (2 строки: названия столбцов)
    ws.cell(row=2, column=1, value="Наименование").font = Font(bold=True)
    ws.cell(row=2, column=2, value="Код").font = Font(bold=True)

    col = 3
    for mc in MONTH_COLS:
        if isinstance(mc, int):
            ws.cell(row=2, column=col, value=MONTHS_RU[mc]).font = Font(bold=True)
        elif mc in {"Q1", "Q2", "Q3", "Q4"}:
            # "Q1" -> "1 кв." и т.д.
            q_label = {"Q1": "1 кв.", "Q2": "2 кв.", "Q3": "3 кв.", "Q4": "4 кв."}[mc]
            ws.cell(row=2, column=col, value=q_label).font = Font(bold=True)
        else:
            ws.cell(row=2, column=col, value="Год").font = Font(bold=True)

        col += 1

    # Тонкие границы/формат
    thin = Side(border_style="thin", color="000000")

    # Данные
    row_idx = 3
    for (code, name) in keys_sorted:
        row_map = rates[(code, name)]

        ws.cell(row=row_idx, column=1, value=name or "")
        ws.cell(row=row_idx, column=2, value=code or "")

        # Подсветка важных валют
        is_highlight = (code in HIGHLIGHT_CODES)

        # Месяцы + кварталы
        col_idx = 3
        cache_vals = {}
        for mc in MONTH_COLS:
            val = row_map.get(mc)
            val_str = _format_val(val)
            c = ws.cell(row=row_idx, column=col_idx, value=val_str)

            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if is_highlight:
                c.font = Font(color="FF0000")
            col_idx += 1


        # Границы/шрифты для первых двух колонок
        for c_idx in (1, 2):
            c = ws.cell(row=row_idx, column=c_idx)
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            c.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
            if is_highlight:
                c.font = Font(color="FF0000")

        row_idx += 1

    # Ширина колонок
    ws.column_dimensions[get_column_letter(1)].width = 34  # Наименование валюты
    ws.column_dimensions[get_column_letter(2)].width = 7    # Код
    for i in range(3, 3 + len(MONTH_COLS)):
        ws.column_dimensions[get_column_letter(i)].width = 10.5

    # Заморозка шапки
    ws.freeze_panes = "C3"

    # Центрирование шапки
    for col in range(1, 2 + len(MONTH_COLS) + 1):
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col).border = Border(top=thin, bottom=thin, left=thin, right=thin)

    wb.save(filename)
    return filename