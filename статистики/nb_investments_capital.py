import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ======================= Настройки =======================
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'ERI',
    'user': 'postgres',
    'password': '20041015R'
}

REGIONS = [
    "АКМОЛИНСКАЯ ОБЛАСТЬ",
    "АКТЮБИНСКАЯ ОБЛАСТЬ",
    "АЛМАТИНСКАЯ ОБЛАСТЬ",
    "АТЫРАУСКАЯ ОБЛАСТЬ",
    "ВОСТОЧНО-КАЗАХСТАНСКАЯ ОБЛАСТЬ",
    "ЖАМБЫЛСКАЯ ОБЛАСТЬ",
    "ЗАПАДНО-КАЗАХСТАНСКАЯ ОБЛАСТЬ",
    "КАРАГАНДИНСКАЯ ОБЛАСТЬ",
    "КОСТАНАЙСКАЯ ОБЛАСТЬ",
    "КЫЗЫЛОРДИНСКАЯ ОБЛАСТЬ",
    "МАНГИСТАУСКАЯ ОБЛАСТЬ",
    "ОБЛАСТЬ АБАЙ",
    "ОБЛАСТЬ ЖЕТІСУ",
    "ОБЛАСТЬ ҰЛЫТАУ",
    "ПАВЛОДАРСКАЯ ОБЛАСТЬ",
    "СЕВЕРО-КАЗАХСТАНСКАЯ ОБЛАСТЬ",
    "ТУРКЕСТАНСКАЯ ОБЛАСТЬ",
    "ЮЖНО-КАЗАХСТАНСКАЯ ОБЛАСТЬ",
    "Г.АЛМАТЫ",
    "Г.АСТАНА",
    "Г.ШЫМКЕНТ",
    "РЕСПУБЛИКА КАЗАХСТАН"
]


# ======================= Основная функция =======================
def build_excel_investments(filename="Инвестиции_в_основной_капитал.xlsx"):
    # --- Подключаемся к базе и получаем данные ---
    query = """
        SELECT region AS "Регион", year AS "Год", value / 1000000 AS "Значение"
        FROM investments_data
        WHERE region IN %s
        ORDER BY region, year;
    """

    with psycopg2.connect(**DB_CONFIG) as conn:
        df = pd.read_sql_query(query, conn, params=(tuple(REGIONS),))

    if df.empty:
        raise ValueError("❌ Нет данных в таблице investments_data для указанных регионов")

    # --- Сводная таблица ---
    pivot = df.pivot_table(index="Регион", columns="Год", values="Значение", aggfunc="sum").fillna(0)
    pivot = pivot.loc[[r for r in REGIONS if r in pivot.index]]  # сохраняем нужный порядок

    # --- Создание Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Инвестиции"

    # --- Заголовок ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(pivot.columns) + 1)
    ws["A1"] = "Инвестиции в основной капитал"
    ws["A1"].font = Font(size=12, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # --- Единицы измерения ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(pivot.columns) + 1)
    ws["A2"] = "(млн тенге)"
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- Заголовки ---
    ws.append(["Регион / Годы"] + [str(y) for y in pivot.columns])

    thin = Side(border_style="thin", color="000000")

    # --- Данные ---
    for idx, (region, row) in enumerate(pivot.iterrows(), start=4):
        ws.append([region] + [round(v) for v in row.tolist()])

    # --- Форматирование ---
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 40
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(filename)
    print(f"✅ Файл сохранён: {filename}")
    return filename

