import requests
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter, MonthLocator
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import matplotlib

# Настройки кириллицы
matplotlib.rcParams['font.family'] = 'DejaVu Sans'
matplotlib.rcParams['axes.unicode_minus'] = False

API_URL = "https://data.nationalbank.kz/api/report/stat-data-locale?form_id=28&date_from=2015-01-01&date_to=2050-01-01&lang=ru"


def load_usd_data():
    """Загружает и подготавливает данные по валюте USD"""
    r = requests.get(API_URL, timeout=(10, 30))
    r.raise_for_status()
    data = r.json()

    rows = []
    for item in data:
        date = item.get("reportDate")
        amount = item.get("amount")
        attrs = item.get("attributes", [])

        currency = None
        region = None
        for a in attrs:
            if a.get("attrCode") == "currency":
                currency = a.get("valueCode")
            if a.get("attrCode") == "region":
                region = a.get("valueNameRu")

        # Берём только USD и месячные данные
        if currency == "USD" and date and amount is not None:
            d = pd.to_datetime(date)

            # ⚠️ Сдвигаем месяц на -1 (например, 2024-03-01 → февраль 2024)
            shifted = d - pd.DateOffset(months=1)
            rows.append({
                "Дата_отчета": d,
                "Фактический_месяц": shifted,
                "Регион": region,
                "Сумма": float(amount)
            })

    df = pd.DataFrame(rows)
    return df


def build_usd_excel_with_chart(filename: str, year: int):
    """Создаёт Excel с графиком нетто продаж USD по выбранный год"""
    df = load_usd_data()

    # Группировка по фактическому месяцу
    df["Год"] = df["Фактический_месяц"].dt.year
    df["Месяц"] = df["Фактический_месяц"].dt.month

    # Фильтр по выбранному году
    df = df[df["Год"] <= year]

    grouped = df.groupby(["Год", "Месяц"], as_index=False)["Сумма"].sum()

    # Названия месяцев по-русски
    MONTHS_RU = {
        1: "январь", 2: "февраль", 3: "март", 4: "апрель",
        5: "май", 6: "июнь", 7: "июль", 8: "август",
        9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
    }
    grouped["Месяц"] = grouped["Месяц"].map(MONTHS_RU)
    grouped["Период"] = grouped["Месяц"] + " " + grouped["Год"].astype(str)
    grouped["Дата"] = pd.to_datetime(grouped["Год"].astype(str) + "-" +
                                     grouped["Период"].apply(lambda x: str(list(MONTHS_RU.values()).index(x.split()[0]) + 1)) + "-01")

    # === Создаём Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "USD по Республике"

    ws.merge_cells("A1:C1")
    ws["A1"] = f"Нетто продажи валюты (USD) — Всего по Республике (до {year} г.)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(["Год", "Месяц (факт.)", "USD, млн тенге"])
    for _, row in grouped.iterrows():
        ws.append([row["Год"], row["Месяц"], round(row["Сумма"], 1)])

    # Оформление таблицы
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
        for c in row:
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # === Второй лист с графиком ===
    ws_chart = wb.create_sheet("График USD")

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(grouped["Дата"], grouped["Сумма"], color="#1236c2", linewidth=2.5, marker="o", label="USD (млн тенге)")

    ax.set_title(f"Нетто продажи валюты USD — Всего по Республике (до {year} г.)",
                 fontsize=12, fontweight="bold", pad=20)
    ax.set_ylabel("млн тенге", fontsize=10)
    ax.grid(True, linestyle="--", alpha=0.5)

    for spine in ax.spines.values():
        spine.set_linewidth(1.2)
        spine.set_color("#000000")

    # --- Настройка оси X ---
    ax.xaxis.set_major_locator(MonthLocator(interval=2))
    ax.xaxis.set_major_formatter(DateFormatter('%b.%y'))
    plt.xticks(rotation=45, ha="right")

    # Переводим месяца на русский
    plt.draw()
    labels = [lbl.get_text() for lbl in ax.get_xticklabels()]
    MONTHS_EN_RU = {
        'Jan': 'янв', 'Feb': 'фев', 'Mar': 'мар', 'Apr': 'апр', 'May': 'май',
        'Jun': 'июн', 'Jul': 'июл', 'Aug': 'авг', 'Sep': 'сен', 'Oct': 'окт',
        'Nov': 'ноя', 'Dec': 'дек'
    }
    new_labels = []
    for lbl in labels:
        if '.' in lbl:
            m, y = lbl.split('.')
            new_labels.append(f"{MONTHS_EN_RU.get(m, m)}.{y}")
        else:
            new_labels.append(lbl)
    ax.set_xticklabels(new_labels, rotation=45, ha="right")

    # Подписи
    ax.legend(loc="upper left", fontsize=9)
    plt.figtext(0.01, -0.02, "Источник: по данным НБ РК", fontsize=9, color="gray")
    plt.figtext(0.87, -0.02, "TENGENOMIKA", fontsize=9, color="#1236c2", fontweight="bold")

    plt.tight_layout()

    # Добавляем график в Excel
    img_data = BytesIO()
    plt.savefig(img_data, format="png", dpi=200, bbox_inches="tight")
    img_data.seek(0)
    plt.close()

    img = Image(img_data)
    ws_chart.add_image(img, "A1")

    wb.save(filename)
    return filename
