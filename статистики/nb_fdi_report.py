def build_excel_pii(filename="ПИИ_по_годам.xlsx"):
    """Создаёт отчёт по прямым иностранным инвестициям (валовые и чистые ПИИ)."""
    import requests, pandas as pd, matplotlib.pyplot as plt
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    API_GROSS = "https://data.nationalbank.kz/api/report/stat-data-locale?form_id=287&date_from=2006-01-01&date_to=2050-01-01&lang=ru"
    API_NET = "https://data.nationalbank.kz/api/report/stat-data-locale?form_id=310&date_from=2006-01-01&date_to=2050-01-01&lang=ru"

    def load_pii(api_url):
        r = requests.get(api_url, timeout=(10, 30))
        data = r.json()
        rows = []
        for d in data:
            date = d.get("reportDate")
            amount = d.get("amount")
            if not date or amount is None:
                continue
            year = int(date[:4])
            month = int(date[5:7])
            if month == 1:
                year -= 1
            rows.append({"Год": year, "Месяц": month, "Сумма": float(amount)})
        return pd.DataFrame(rows)

    gross_df = load_pii(API_GROSS)
    net_df = load_pii(API_NET)

    latest_row = gross_df.sort_values(["Год", "Месяц"]).iloc[-1]
    latest_year, latest_month = int(latest_row["Год"]), int(latest_row["Месяц"])
    latest_label = f"{latest_year} ({latest_month}М)" if latest_month < 12 else str(latest_year)

    gross_yearly = gross_df.groupby("Год", as_index=False)["Сумма"].sum()
    net_yearly = net_df.groupby("Год", as_index=False)["Сумма"].sum()

    years = sorted(set(gross_yearly["Год"]) | set(net_yearly["Год"]))
    table = pd.DataFrame({"Год": years})
    table["Валовые ПИИ"] = table["Год"].map(dict(zip(gross_yearly["Год"], gross_yearly["Сумма"])))
    table["Чистые ПИИ"] = table["Год"].map(dict(zip(net_yearly["Год"], net_yearly["Сумма"])))
    table = table.fillna(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "ПИИ"
    ws.append(["Показатель"] + [str(y) for y in table["Год"]])
    ws.append(["Валовые ПИИ"] + [round(v, 1) for v in table["Валовые ПИИ"]])
    ws.append(["Чистые ПИИ"] + [round(v, 1) for v in table["Чистые ПИИ"]])

    # стиль
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 22

    # ==================== График ====================
    ws_chart = wb.create_sheet("График")
    plt.figure(figsize=(11, 6))
    plt.rcParams['font.family'] = 'DejaVu Sans'
    years_str = [str(y) for y in table["Год"]]
    x_positions = range(len(years_str))
    # --- Линии ---
    plt.plot(
        x_positions, table["Валовые ПИИ"],
        color="#0a1a5b", linewidth=3, marker='o', markersize=6, label="Валовые ПИИ"
    )
    plt.plot(
        x_positions, table["Чистые ПИИ"],
        color="#1ba64f", linewidth=3, marker='o', markersize=6, label="Чистые ПИИ"
    )

    # --- Подписи чисел (через год и немного выше линии) ---
    for i, (x, y) in enumerate(zip(x_positions, table["Валовые ПИИ"])):
        if i % 2 == 0 and y > 0:
            plt.text(
                x, y + max(table["Валовые ПИИ"]) * 0.035, f"{y:,.1f}",
                ha='center', va='bottom', fontsize=9,
                color="#000000", 
                fontweight="bold"
            )

    for i, (x, y) in enumerate(zip(x_positions, table["Чистые ПИИ"])):
        if i % 2 == 0 and y > 0:
            plt.text(
                x, y + max(table["Чистые ПИИ"]) * 0.035, f"{y:,.1f}",
                ha='center', va='bottom', fontsize=9,
                color="#000000", 
                fontweight="bold"
            )

    # --- Оформление ---
    plt.title("Прямые иностранные инвестиции в Казахстан", fontsize=14, weight="bold", pad=20)
    plt.grid(True, color="#d9d9d9", linestyle="-", linewidth=0.7, alpha=0.6)
    plt.xlabel("")
    plt.ylabel("")
    plt.xticks(x_positions, years_str, rotation=30, fontsize=9)
    plt.yticks(fontsize=9)
    plt.ylim(bottom=0)

    plt.legend(loc="upper left", fontsize=9, frameon=False)

    # --- Подписи внизу ---
    plt.text(-0.3, -3000, "По данным НБ РК", fontsize=9, color="#555555")
    plt.text(len(years_str) - 3, -3000, "TENGENOMIKA", fontsize=10, color="#0a1a5b", weight="bold")

    # --- Сохранение ---
    plt.tight_layout()
    img_data = BytesIO()
    plt.savefig(img_data, format="png", dpi=300, bbox_inches="tight")
    plt.close()
    img_data.seek(0)

    img = Image(img_data)
    ws_chart.add_image(img, "B2")

    wb.save(filename)
    return filename
