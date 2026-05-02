import pandas as pd
import datetime
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


# ======================= Классификация кодов по разделам =======================

SECTION_MAP = {
    "I. Доходы, связанные с получением вознаграждения": range(4050, 4500),
    "II. Доходы, не связанные с получением вознаграждения": range(4500, 5000),
    "IV. Расходы, связанные с выплатой вознаграждения": range(5020, 5450),
    "V. Расходы, не связанные с выплатой вознаграждением": range(5450, 5999),
    "VII. Подоходный налог": range(5999, 6000),
}


# ======================= Основная функция =======================

def build_excel_banks_stats(records, filename="Статистика_по_банкам.xlsx"):
    rows = []
    for r in records:
        if not r.get("attributes"):
            continue
        attr = r["attributes"][0]
        code = attr.get("valueCode")
        if not code or not code.isdigit():
            continue
        rows.append({
            "Код": int(code),
            "Наименование": attr.get("valueNameRu"),
            "Дата": pd.to_datetime(r["reportDate"]).date(),
            "Значение": r.get("amount", 0)
        })

    df = pd.DataFrame(rows)
    if df.empty:
        raise ValueError("Нет данных для формирования отчёта")

    # --- 12 месяцев ---
    months = sorted(df["Дата"].unique())
    if len(months) > 12:
        months = months[-12:]
        df = df[df["Дата"].isin(months)]

    df["Дата"] = df["Дата"].apply(lambda d: d.strftime("%m.%Y"))

    pivot = df.pivot_table(
        index=["Код", "Наименование"],
        columns="Дата",
        values="Значение",
        aggfunc="sum"
    ).fillna(0).sort_index()

    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика по банкам"

    # --- Шапка ---
    ws.merge_cells("A1:O1")
    ws["A1"] = "Сводный отчет о доходах и расходах по банкам второго уровня Республики Казахстан"
    ws["A1"].font = Font(size=12, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:O2")
    ws["A2"] = "(тыс. тенге)"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])
    ws.append(["№ группы счетов", "Наименование доходов и расходов", "Строка"] + list(pivot.columns))

    thin = Side(border_style="thin", color="000000")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for section, code_range in SECTION_MAP.items():
        ws.append([section])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column)
        ws.cell(ws.max_row, 1).font = bold_font
        ws.cell(ws.max_row, 1).alignment = Alignment(horizontal="left")

        section_df = pivot.loc[pivot.index.get_level_values("Код").isin(code_range)]

        for (code, name), row in section_df.iterrows():
            ws.append([code, name, code] + list(row.values))

        if not section_df.empty:
            ws.append([f"Итого {section.split('.')[1].strip()}:", "", ""] +
                      [section_df[c].sum() for c in pivot.columns])
            for cell in ws[ws.max_row]:
                cell.font = bold_font

        ws.append([])

    # --- Итоги ---
    def section_sum(code_prefix):
        df_section = pivot[pivot.index.get_level_values("Код").astype(str).str.startswith(str(code_prefix))]
        return df_section[pivot.columns].sum().tolist() if not df_section.empty else [0]*len(pivot.columns)

    I_sum = section_sum(40)
    II_sum = section_sum(45)
    IV_sum = section_sum(50)
    V_sum = section_sum(54)
    VII_sum = section_sum(59)

    VIII = [i + j - k - l for i, j, k, l in zip(I_sum, II_sum, IV_sum, V_sum)]
    IX = [viii - vii for viii, vii in zip(VIII, VII_sum)]
    perc_income = [i - j for i, j in zip(I_sum, IV_sum)]
    nonperc_income = [i - j for i, j in zip(II_sum, V_sum)]

    # --- % прирост ---
    def growth(series):
        g = [""]
        for i in range(1, len(series)):
            prev = series[i-1]
            curr = series[i]
            if prev == 0:
                g.append("")
            else:
                g.append(round((curr - prev) / prev * 100, 1))
        return g

    perc_growth = growth(perc_income)
    nonperc_growth = growth(nonperc_income)
    total_growth = growth(IX)

    ws.append([])
    ws.append(["VIII", "Превышение текущих доходов над расходами (до налога)", ""] + VIII)
    ws.append(["VII", "Подоходный налог (КПН)", ""] + VII_sum)
    ws.append(["IX", "Превышение текущих доходов над расходами (после налога)", ""] + IX)
    ws.append(["", "Чистые процентные доходы", ""] + perc_income)
    ws.append(["", "% рост чистых процентных доходов", ""] + perc_growth)
    ws.append(["", "Чистые непроцентные доходы", ""] + nonperc_income)
    ws.append(["", "% рост чистых непроцентных доходов", ""] + nonperc_growth)
    ws.append(["", "Совокупная чистая прибыль", ""] + IX)
    ws.append(["", "% рост совокупной прибыли", ""] + total_growth)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 85
    ws.column_dimensions["C"].width = 10
    for col in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ==================== ГРАФИК ====================

    # --- Формируем месяцы по фактическим данным ---
    available_months = list(pivot.columns)
    month_labels = []
    month_names = {
        "01": "Янв", "02": "Фев", "03": "Мар", "04": "Апр", "05": "Май", "06": "Июн",
        "07": "Июл", "08": "Авг", "09": "Сен", "10": "Окт", "11": "Ноя", "12": "Дек"
    }

    for m in available_months:
        try:
            mm, yy = m.split(".")
            label = month_names.get(mm, mm)
        except Exception:
            label = m
        month_labels.append(label)

    num_points = len(available_months)
    x = range(num_points)

    # --- Данные ---
    y1 = [round(v / 1_000_000_000, 1) for v in perc_income[:num_points]]      # млрд тг
    y2 = [round(v / 1_000_000_000, 1) for v in nonperc_income[:num_points]]
    y3 = [round(v / 1_000_000_000, 1) for v in IX[:num_points]]

    # --- Построение ---
    import matplotlib.pyplot as plt
    from io import BytesIO

    fig, ax = plt.subplots(figsize=(10, 6))
    bar_width = 0.35

    # Синие и красные столбцы
    ax.bar([i - bar_width/2 for i in x], y1, width=bar_width, color="#002D72", label="Чистые процентные доходы (млрд тг)")
    ax.bar([i + bar_width/2 for i in x], y2, width=bar_width, color="#D82626", label="Чистые непроцентные доходы (млрд тг)")

    # Голубая линия — совокупная прибыль
    ax.plot(x, y3, color="#3CA8E0", marker="o", linewidth=2.5, label="Совокупная чистая прибыль (млрд тг)")

    # --- Подписи значений над столбцами ---
    for i, v in enumerate(y1):
        ax.text(i - 0.25, v + (0.05 if v >= 0 else -0.15), f"{int(v)}", ha='center', va='bottom' if v >= 0 else 'top', fontsize=8, color="#002D72", fontweight='bold')
    for i, v in enumerate(y2):
        ax.text(i + 0.25, v + (0.05 if v >= 0 else -0.15), f"{int(v)}", ha='center', va='bottom' if v >= 0 else 'top', fontsize=8, color="#D82626", fontweight='bold')

    for i, v in enumerate(y3):
        ax.text(i, v + 0.15, f"{int(v)}", ha='center', fontsize=8, color="#3CA8E0", fontweight='bold')

    # --- Стилизация ---
    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels(month_labels, fontsize=9)
    ax.set_ylabel("млрд тенге", fontsize=9)
    ax.set_title("Динамика доходов банковского сектора Казахстана", fontsize=11, fontweight="bold", pad=20)
    ax.legend(loc="upper left", fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.3)
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')

    # --- Логотип и подпись источника ---
    plt.figtext(0.01, -0.03, "По данным НБ РК", fontsize=8, color="gray")
    plt.figtext(0.85, -0.03, "TENGENOMIKA", fontsize=9, fontweight='bold', color="#002D72")

    plt.tight_layout()
    img_buf = BytesIO()
    plt.savefig(img_buf, format="png", dpi=300, bbox_inches="tight")
    plt.close(fig)
    img_buf.seek(0)

    # --- Вставка в Excel ---
    ws_chart = wb.create_sheet("График доходов")
    img = Image(img_buf)
    ws_chart.add_image(img, "A1")

    wb.save(filename)
    return filename