import psycopg2
from collections import OrderedDict, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
import openpyxl


# -------------------------- SQL загрузка --------------------------

def fetch_rows(conn, kato_like: str, datereport: str):
    """Выборка по региону и дате — с корректной агрегацией"""
    if datereport is None:
        datereport = get_latest_datereport(conn)
    q = """
        SELECT
            grbId,
            katFgr,
            katFgrName,
            klsFpgrName,
            pklAbpName,
            spkPrgName,
            SUM(sumrg) AS sumrg,
            SUM(plg) AS plg,
            SUM(utv) AS utv,
            SUM(utch) AS utch,
            SUM(plgo) AS plgo,
            SUM(plgp) AS plgp,
            MAX(kato) AS kato
        FROM emfreportbpr_v9
        WHERE datereport = %s
          AND CAST(kato AS TEXT) LIKE %s
          AND (
                (grbid = 1 AND (
                    (katfgr = 1 AND strid = 1) OR
                    (katfgr = 2 AND strid = 2 AND adtype IS NULL) OR
                    (katfgr = 3 AND strid = 3) OR
                    (katfgr = 4 AND strid = 4 AND adtype NOT IN (3,4))
                ))
                OR
                (grbid = 2 AND (
                    (katfgr = 1 AND adtype IS NULL) 
                    OR (katfgr = 2) 
                    OR (katfgr = 3) 
                    OR (katfgr = 4)
                    OR (katfgr = 5 AND (adtype IS NULL or adtype NOT IN (3, 4))) 
                    OR (katfgr = 6 AND adtype IS NULL) 
                    OR (katfgr = 7 AND adtype IS NULL) 
                    OR (katfgr = 8 AND adtype IS NULL) 
                    OR (katfgr = 9)
                    OR (katfgr = 10 AND adtype IS NULL) 
                    OR (katfgr = 11 AND adtype IS NULL) 
                    OR (katfgr = 12 AND adtype IS NULL)
                    OR (katfgr = 13)
                    OR (katfgr = 14 AND (adtype IS NULL or adtype NOT IN (3, 4))) 
                    OR (katfgr = 15 AND adtype NOT IN (3,4))
                ))
                OR grbid IN (3,4,5,6)
          )
        GROUP BY
            grbId, katFgr, katFgrName, klsFpgrName, pklAbpName, spkPrgName
        ORDER BY
            grbId, katFgr, katFgrName, klsFpgrName, pklAbpName, spkPrgName
    """
    cur = conn.cursor()
    cur.execute(q, (datereport, kato_like))
    rows = cur.fetchall()
    cur.close()
    return rows


def get_latest_datereport(conn):
    """Возвращает самую свежую дату datereport из базы"""
    cur = conn.cursor()
    cur.execute("SELECT MAX(datereport) FROM emfreportbpr_v9;")
    latest = cur.fetchone()[0]
    cur.close()
    return latest

# --------------------- Иерархия и агрегация ---------------------

def normalize_number(x):
    if x is None:
        return 0.0
    try:
        return float(str(x).replace(",", ".")) / 1000.0
    except:
        return 0.0


def build_hierarchy(rows, grb_id):
    """Строим структуру строк: katFgrName → pklAbpName (для доходов) или katFgrName → klsFpgrName (для затрат)"""
    out = []
    sec_title = "I. Доходы" if grb_id == 1 else "II. Затраты"
    out.append((("SEC", grb_id), sec_title, "sec"))

    fgr_groups = OrderedDict()
    for r in rows:
        if r[0] != grb_id:
            continue
        _, katFgr, katFgrName, klsFpgrName, pklAbpName, *_ = r

        # 👇 выбираем, какое поле использовать
        sub_name = pklAbpName if grb_id == 1 else klsFpgrName
        if sub_name is None:
            continue

        fgr = fgr_groups.setdefault((katFgr, katFgrName), [])
        if sub_name not in fgr:
            fgr.append(sub_name)

    for (katFgr, katFgrName), sub_list in fgr_groups.items():
        out.append((("FGR", grb_id, katFgr, katFgrName), katFgrName.upper(), "fgr"))
        for name in sub_list:
            out.append((("PKL", grb_id, katFgr, katFgrName, name), f"   {name}", "pkl"))
    return out


def compute_values(rows, field_index, grb_id):
    """Агрегация по grb_id, katFgrName и pklAbpName / klsFpgrName"""
    prg_map = defaultdict(float)
    for r in rows:
        if r[0] != grb_id:
            continue
        sec, katFgr, katFgrName, klsFpgrName, pklAbpName = r[0], r[1], r[2], r[3], r[4]
        sub_name = pklAbpName if grb_id == 1 else klsFpgrName
        if sub_name is None:
            continue
        val = normalize_number(r[field_index])
        key = ("PKL", sec, katFgr, katFgrName, sub_name)
        prg_map[key] += val

    fgr_map = defaultdict(float)
    sec_map = defaultdict(float)

    for key, v in prg_map.items():
        _, sec, katFgr, katFgrName, _ = key
        fgrk = ("FGR", sec, katFgr, katFgrName)
        fgr_map[fgrk] += v
        seck = ("SEC", sec)
        sec_map[seck] += v

    values = {}
    values.update(prg_map)
    values.update(fgr_map)
    values.update(sec_map)
    return values


#Общая иерархия
def build_full_hierarchy(conn, regions, grb_id):
    """Создает объединённый список всех pklAbpName из всех регионов и годов"""
    all_rows = []
    latest_date = get_latest_datereport(conn)
    for kato_like, _ in regions:
        for date in ("2024-01-01", "2025-01-01", str(latest_date)):
            all_rows.extend(fetch_rows(conn, kato_like, date))
    return build_hierarchy(all_rows, grb_id)

# -------------------------- Запись в Excel --------------------------

def write_summary_workbook(conn, output_path: str):
    REGIONS = [
        ("10%", "Абайская область"),
        ("11%", "Акмолинская область"),
        ("15%", "Актюбинская область"),
        ("19%", "Алматинская область"),
        ("23%", "Атырауская область"),
        ("27%", "Западно-Казахстанская область"),
        ("31%", "Жамбылская область"),
        ("33%", "область Жетісу"),
        ("35%", "Карагандинская область"),
        ("39%", "Костанайская область"),
        ("43%", "Кызылординская область"),
        ("47%", "Мангистауская область"),
        ("51%", "Южно-Казахстанская область"),
        ("55%", "Павлодарская область"),
        ("59%", "Северо-Казахстанская область"),
        ("61%", "Туркестанская область"),
        ("62%", "область Ұлытау"),
        ("63%", "Восточно-Казахстанская область"),
        ("71%", "г.Астана"),
        ("75%", "г.Алматы"),
        ("79%", "г.Шымкент"),
    ]

    COL_LABELS = ["2023 Исп.", "2024 Исп.", "2025 Исп.", "2025 Утв.", "2025 Уточн.", "2025 Скорр."]
    IDX = {"sumrg": 6, "plg": 7, "utv": 8, "utch": 9}

    wb = Workbook()
    # 🔹 Удаляем стандартный пустой лист
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # цикл по листам (Доходы / Затраты)
    for sheet_name, grb_id in [("Доходы", 1), ("Затраты", 2)]:
        print(f"📄 Формируем лист: {sheet_name}...")
        ws = wb.create_sheet(sheet_name)

        # Формируем эталонную иерархию
        row_scheme = build_full_hierarchy(conn, REGIONS, grb_id)
        
        latest_date = get_latest_datereport(conn)
        
        # Собираем все данные по регионам и годам
        val_maps = {}
        for kato_like, _ in REGIONS:
            r2023 = fetch_rows(conn, kato_like, "2024-01-01")
            val_maps[(kato_like, "2023")] = compute_values(r2023, IDX["sumrg"], grb_id)
            r2024 = fetch_rows(conn, kato_like, "2025-01-01")
            val_maps[(kato_like, "2024")] = compute_values(r2024, IDX["sumrg"], grb_id)
            r2025 = fetch_rows(conn, kato_like, str(latest_date))
            val_maps[(kato_like, "2025_sumrg")] = compute_values(r2025, IDX["sumrg"], grb_id)
            val_maps[(kato_like, "2025_utv")]   = compute_values(r2025, IDX["utv"], grb_id)
            val_maps[(kato_like, "2025_utch")]  = compute_values(r2025, IDX["utch"], grb_id)
            val_maps[(kato_like, "2025_plg")]   = compute_values(r2025, IDX["plg"], grb_id)

        # ---------- Оформление шапки ----------
        ws["A1"] = "Наименования"
        ws.merge_cells("A1:A2")
        ws["A1"].font = Font(name="Calibri", bold=True, size=10)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

        start_col = 2
        for _, region_name in REGIONS:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 6 - 1)
            cell = ws.cell(row=1, column=start_col, value=region_name)
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for j, text in enumerate(COL_LABELS):
                subcell = ws.cell(row=2, column=start_col + j, value=text)
                subcell.font = Font(name="Calibri", bold=True, size=9)
                subcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            start_col += 6

        # ---------- Тело таблицы ----------
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_sec = Font(name="Calibri", bold=True, size=9)
        font_fgr = Font(name="Calibri", bold=True, size=9)
        font_pkl = Font(name="Calibri", size=8)

        thin = Side(style="thin")
        thick = Side(style="medium")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_sec = Font(name="Calibri", bold=True, size=9)
        font_fgr = Font(name="Calibri", bold=True, size=9)
        font_pkl = Font(name="Calibri", size=8)

        row = 3
        for key, label, style in row_scheme:
            if key[1] != grb_id:
                continue

            cellA = ws.cell(row=row, column=1, value=label)
            if key[0] == "SEC":
                # Общий итог: Доходы / Затраты
                cellA.font = Font(name="Calibri", bold=True, size=10)
                ws.row_dimensions[row].height = 22
            elif style == "fgr":
                # katFgrName (группа)
                cellA.font = Font(name="Calibri", bold=True, size=9)
            else:
                # pklAbpName (обычные строки)
                cellA.font = font_pkl

            cellA.alignment = Alignment(vertical="center", wrap_text=True)
            cellA.border = border

            col = 2
            for kato_like, _ in REGIONS:
                vals = [
                    val_maps.get((kato_like, "2023"), {}).get(key, 0.0),
                    val_maps.get((kato_like, "2024"), {}).get(key, 0.0),
                    val_maps.get((kato_like, "2025_sumrg"), {}).get(key, 0.0),
                    val_maps.get((kato_like, "2025_utv"), {}).get(key, 0.0),
                    val_maps.get((kato_like, "2025_utch"), {}).get(key, 0.0),
                    val_maps.get((kato_like, "2025_plg"), {}).get(key, 0.0),
                ]
                for v in vals:
                    c = ws.cell(row=row, column=col, value=round(v, 3))
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.border = border
                    col += 1
            row += 1

        # ---------- Размеры колонок ----------
        ws.column_dimensions["A"].width = 45
        for col_idx in range(2, ws.max_column + 1):
            letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[letter].width = 13

    # ---------- Сохраняем ----------
    wb.save(output_path)
