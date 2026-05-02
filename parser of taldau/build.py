"""
build_kazakhstan.py — единый скрипт
  Читает ALL_KAZAKHSTAN_FULL.csv (годовые 2019-2024)
         MONTHLY_FINAL.csv       (месячные Янв 2025 – Янв-Фев 2026)
  Объединяет по ключу (oblast, region, depth, text)
  Дубли в FULL (один текст, два id с разными годами) — схлопываются в одну строку
  Строит parent_id через (text, depth-1) lookup
  DFS-сортировка иерархии
  Создаёт KAZAKHSTAN_FINAL.xlsx
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

FULL_CSV    = "ALL_KAZAKHSTAN_FULL.csv"
MONTHLY_CSV = "MONTHLY_FINAL.csv"
OUTPUT_XLSX = "KAZAKHSTAN_FINAL.xlsx"

YEAR_MAP = {
    "y122019": "2019",
    "y122020": "2020",
    "y122021": "2021",
    "y122022": "2022",
    "y122023": "2023",
    "y122024": "2024",
}

MONTHLY_MAP = {
    "y012025": "Янв 2025",
    "y022025": "Янв-Фев 2025",
    "y032025": "Янв-Мар 2025",
    "y042025": "Янв-Апр 2025",
    "y052025": "Янв-Май 2025",
    "y062025": "Янв-Июн 2025",
    "y072025": "Янв-Июл 2025",
    "y082025": "Янв-Авг 2025",
    "y092025": "Янв-Сен 2025",
    "y102025": "Янв-Окт 2025",
    "y112025": "Янв-Ноя 2025",
    "y122025": "2025 (год)",
    "y012026": "Янв 2026",
    "y022026": "Янв-Фев 2026",
}

LEVEL_COLORS = {
    0: ("1F3864", "FFFFFF", True,  11),
    1: ("2E75B6", "FFFFFF", True,  10),
    2: ("BDD7EE", "1F3864", True,   9),
    3: ("DDEBF7", "1F3864", False,  9),
    4: ("EEF4FB", "2E3A4E", False,  9),
    5: ("F5F9FD", "2E3A4E", False,  9),
    6: ("FFFFFF", "2E3A4E", False,  9),
}

HDR_NAV     = "1F3864"
HDR_KEY     = "1F4E79"
HDR_ANNUAL  = "2E75B6"
HDR_MONTHLY = "14375E"

BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

KEY = ["oblast", "region", "depth", "text"]

def safe_title(name):
    for ch in r"\/*?:[]'":
        name = name.replace(ch, "")
    return name[:31].strip()

def format_val(val):
    s = str(val).strip()
    if s.lower() in ("x", "х"):
        return "x"
    if s.lower() in ("nan", "none", "nat", ""):
        return None
    try:
        f = float(s.replace(" ", "").replace(",", "."))
        return None if np.isnan(f) else int(round(f / 1000))
    except (ValueError, TypeError):
        return None

def apply_style(cell, depth, h_align="left"):
    bg, fg, bold, size = LEVEL_COLORS.get(min(depth, 6), ("FFFFFF", "000000", False, 9))
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal=h_align, vertical="center")

def apply_header(cell, bg):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Чтение и дедупликация FULL ─────────────────────────────────────────────────
print(f"Читаю {FULL_CSV}...")
df_full = pd.read_csv(FULL_CSV, encoding="utf-8-sig", low_memory=False)
year_cols = [c for c in YEAR_MAP if c in df_full.columns]
df_full = df_full[KEY + ["id", "parent_text"] + year_cols]

# Дубли: одна позиция разбита на 2 строки с разными id и непересекающимися годами
# Схлопываем: берём первое непустое значение по каждому году, id — меньший (оригинальный)
dups_mask = df_full.duplicated(subset=KEY, keep=False)
n_dups = dups_mask.sum()
if n_dups > 0:
    print(f"  Схлопываю {n_dups} дублирующихся строк ({n_dups//2} позиций)...")
    dup_df = df_full[dups_mask].copy()
    seen = {}
    for _, row in dup_df.iterrows():
        k = tuple(row[c] for c in KEY)
        if k not in seen:
            seen[k] = row.copy()
        else:
            for col in year_cols:
                if pd.isna(seen[k][col]) and pd.notna(row[col]):
                    seen[k][col] = row[col]
            if pd.notna(row["id"]) and pd.notna(seen[k]["id"]):
                seen[k]["id"] = min(seen[k]["id"], row["id"])
    deduped = pd.DataFrame(list(seen.values()))
    df_full = pd.concat([df_full[~dups_mask], deduped], ignore_index=True)

print(f"  Строк: {len(df_full):,}  Областей: {df_full['oblast'].nunique()}  Районов: {df_full['region'].nunique()}")

print(f"Читаю {MONTHLY_CSV}...")
df_monthly = pd.read_csv(MONTHLY_CSV, encoding="utf-8-sig", low_memory=False)
monthly_cols = [c for c in MONTHLY_MAP if c in df_monthly.columns]
df_monthly = df_monthly[KEY + ["id", "parent_text"] + monthly_cols]
print(f"  Строк: {len(df_monthly):,}  Областей: {df_monthly['oblast'].nunique()}  Районов: {df_monthly['region'].nunique()}")

# ── Объединение ────────────────────────────────────────────────────────────────
print("Объединяю датасеты...")
df = pd.merge(
    df_full,
    df_monthly[KEY + ["id", "parent_text"] + monthly_cols],
    on=KEY,
    how="outer",
    suffixes=("_full", "_monthly"),
)
df["id"]          = df["id_full"].combine_first(df["id_monthly"])
df["parent_text"] = df["parent_text_full"].combine_first(df["parent_text_monthly"])
df.drop(columns=["id_full", "id_monthly", "parent_text_full", "parent_text_monthly"], inplace=True)
df["depth"] = df["depth"].fillna(0).astype(int)
df["id"]    = df["id"].astype("Int64")
print(f"  Итого строк после объединения: {len(df):,}")

all_data_cols = year_cols + monthly_cols

# ── Построение parent_id ───────────────────────────────────────────────────────
# Ключ (text.strip(), depth-1) — корректен при дублирующихся названиях узлов
print("Строю parent_id...")

def add_parent_id(group):
    group = group.copy()
    dtid = {}
    for _, row in group.iterrows():
        dtid[(str(row["text"]).strip(), int(row["depth"]))] = row["id"]
    def get_pid(row):
        if pd.isna(row["parent_text"]):
            return None
        return dtid.get((str(row["parent_text"]).strip(), int(row["depth"]) - 1))
    group["parent_id"] = group.apply(get_pid, axis=1)
    return group

parts = []
for (ob, reg), grp in df.groupby(["oblast", "region"], sort=False):
    parts.append(add_parent_id(grp))
df = pd.concat(parts, ignore_index=True)

# ── DFS-сортировка ─────────────────────────────────────────────────────────────
def dfs_sort(group):
    rows_by_id = defaultdict(list)
    children   = defaultdict(list)
    for _, row in group.iterrows():
        nid = int(row["id"]) if pd.notna(row["id"]) else None
        pid = int(row["parent_id"]) if pd.notna(row.get("parent_id")) else None
        if nid is not None:
            rows_by_id[nid].append(row)
            children[pid].append(nid)
    all_ids = set(rows_by_id.keys())
    roots = sorted(
        nid for nid in all_ids
        if (int(rows_by_id[nid][0]["parent_id"]) if pd.notna(rows_by_id[nid][0].get("parent_id")) else None)
        not in all_ids
    )
    result, visited = [], set()
    def dfs(nid):
        if nid in visited:
            return
        visited.add(nid)
        result.extend(rows_by_id[nid])
        for child in sorted(children.get(nid, [])):
            dfs(child)
    for root in roots:
        dfs(root)
    for nid in sorted(all_ids - visited):
        result.extend(rows_by_id[nid])
    return pd.DataFrame(result)

# ── Запись листа области ───────────────────────────────────────────────────────
def write_oblast_sheet(ws, oblast_df):
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    headers = (["ID", "Parent ID", "Уровень", "Район", "Наименование (вид деятельности)"]
               + [YEAR_MAP[c] for c in year_cols]
               + [MONTHLY_MAP[c] for c in monthly_cols])
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        if ci <= 3:                    apply_header(cell, HDR_NAV)
        elif ci <= 5:                  apply_header(cell, HDR_KEY)
        elif ci <= 5 + len(year_cols): apply_header(cell, HDR_ANNUAL)
        else:                          apply_header(cell, HDR_MONTHLY)
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "F2"

    excel_row = 2
    total_rows = 0

    for (ob, reg), grp in oblast_df.groupby(["oblast", "region"], sort=False):
        try:
            sorted_grp = dfs_sort(grp)
        except Exception:
            sorted_grp = grp

        for _, row in sorted_grp.iterrows():
            depth = min(int(row.get("depth", 0)), 6)
            nid   = int(row["id"]) if pd.notna(row["id"]) else None
            pid   = int(row["parent_id"]) if pd.notna(row.get("parent_id")) else None

            c = ws.cell(excel_row, 1, nid);    apply_style(c, depth, "right")
            c = ws.cell(excel_row, 2, pid);    apply_style(c, depth, "right")
            c = ws.cell(excel_row, 3, depth);  apply_style(c, depth, "center")
            c = ws.cell(excel_row, 4, str(reg)); apply_style(c, depth, "left")
            c = ws.cell(excel_row, 5, "    " * depth + str(row.get("text", "")).strip())
            apply_style(c, depth, "left")

            for ci, col in enumerate(all_data_cols, 6):
                val = format_val(row.get(col))
                c   = ws.cell(excel_row, ci)
                apply_style(c, depth, "right")
                if val == "x":
                    _, fg, _, size = LEVEL_COLORS.get(min(depth, 6), ("FFFFFF","000000",False,9))
                    c.value = "x"
                    c.font  = Font(name="Arial", color=fg, italic=True, size=size)
                elif val is not None:
                    c.value = val
                    c.number_format = "#,##0"

            ws.row_dimensions[excel_row].height = 15
            if depth > 0:
                ws.row_dimensions[excel_row].outline_level = min(depth, 7)
                ws.row_dimensions[excel_row].hidden = False

            excel_row  += 1
            total_rows += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 60
    for i in range(len(all_data_cols)):
        ws.column_dimensions[get_column_letter(6 + i)].width = 16

    back = ws.cell(excel_row + 1, 1, "← Навигация")
    back.hyperlink = "#'Навигация'!A1"
    back.font      = Font(name="Arial", color="2E75B6", underline="single", bold=True, size=9)
    back.alignment = Alignment(vertical="center")
    return total_rows

# ── Навигационный лист ─────────────────────────────────────────────────────────
def write_nav_sheet(ws, info_map):
    ws.title = "Навигация"
    ws.sheet_view.showGridLines = False

    t = ws.cell(1, 1, "Инвестиции в основной капитал — Казахстан")
    t.font = Font(name="Arial", bold=True, color="1F4E79", size=15)
    t.alignment = Alignment(vertical="center")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 40

    first_col = YEAR_MAP[year_cols[0]] if year_cols else "—"
    last_col  = MONTHLY_MAP[monthly_cols[-1]] if monthly_cols else "—"
    sub = ws.cell(2, 1, f"Данные: {first_col}  →  {last_col}  |  тыс. тенге")
    sub.font = Font(name="Arial", italic=True, color="808080", size=10)
    ws.row_dimensions[2].height = 20

    for ci, h in enumerate(["Область / Город", "Районов", "Строк", "Перейти →"], 1):
        c = ws.cell(3, ci, h)
        c.fill = PatternFill("solid", fgColor=HDR_NAV)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 26

    alt = ["EBF3FB", "FFFFFF"]
    for i, (oblast, info) in enumerate(info_map.items(), 4):
        fill = PatternFill("solid", fgColor=alt[i % 2])
        c = ws.cell(i, 1, oblast)
        c.fill = fill; c.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
        c.border = BORDER; c.alignment = Alignment(vertical="center")
        c = ws.cell(i, 2, info["districts"])
        c.fill = fill; c.font = Font(name="Arial", size=10); c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(i, 3, info["rows"])
        c.fill = fill; c.font = Font(name="Arial", size=10); c.border = BORDER
        c.number_format = "#,##0"; c.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(i, 4, f"→ {info['sheet']}")
        c.hyperlink = f"#{info['sheet']}!A1"
        c.fill = fill
        c.font = Font(name="Arial", color="1F4E79", bold=True, underline="single", size=10)
        c.border = BORDER; c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40

# ── Главный цикл ───────────────────────────────────────────────────────────────
all_oblasts = list(dict.fromkeys(
    list(df_full["oblast"].unique()) + list(df_monthly["oblast"].unique())
))
print(f"\nСоздаю Excel ({len(all_oblasts)} областей/городов)...")

wb = openpyxl.Workbook()
wb.remove(wb.active)
nav_ws   = wb.create_sheet("Навигация")
info_map = {}

for oblast in all_oblasts:
    oblast_df   = df[df["oblast"] == oblast]
    sheet_title = safe_title(oblast)
    ws          = wb.create_sheet(title=sheet_title)
    n_rows      = write_oblast_sheet(ws, oblast_df)
    districts   = oblast_df["region"].nunique()
    info_map[oblast] = {"sheet": sheet_title, "districts": districts, "rows": n_rows}
    print(f"  ✓ {oblast:<45} {districts:>3} р-нов  {n_rows:>6} строк")

write_nav_sheet(nav_ws, info_map)

print(f"\nСохраняю {OUTPUT_XLSX}...")
wb.save(OUTPUT_XLSX)

total = sum(v["rows"] for v in info_map.values())
print(f"\n{'='*60}")
print(f" ГОТОВО!")
print(f"   Листов:       {len(info_map) + 1}")
print(f"   Строк данных: {total:,}")
print(f"   Файл:         {OUTPUT_XLSX}")
print(f"{'='*60}")