"""
Конвертирует akmola_clean.csv (данные из API) в формат XLS с taldau:
- Иерархические отступы (4 пробела × уровень)
- Сортировка: сначала по региону, потом по иерархии
- Значения в тысячах тенге (делим на 1000)
- x и NaN сохраняются как есть

Запуск:
    pip install pandas openpyxl xlrd
    python convert_to_xls_format.py

Результат: taldau_formatted.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "akmola_clean.csv"
OUTPUT_FILE = "taldau_formatted.xlsx"
YEAR_COLS   = ["2019", "2020", "2021", "2022", "2023", "2024"]
INDENT      = "    "  # 4 пробела на уровень (как в оригинале)


def format_value(val):
    """Конвертирует значение: делит на 1000, сохраняет x и пустые."""
    if val == "x" or val == "х":
        return "x"
    try:
        v = float(val)
        if np.isnan(v):
            return None
        return round(v / 1000)   # тенге → тысячи тенге
    except (ValueError, TypeError):
        return None


def build_formatted(df: pd.DataFrame) -> pd.DataFrame:
    """Строит DataFrame с отступами в названиях и правильными значениями."""
    rows = []
    for _, row in df.iterrows():
        level = int(row["Уровень"])
        indent = INDENT * level
        name = indent + str(row["Наименование"]).strip()

        new_row = {
            "Регион":        row["Регион"],
            "Наименование":  name,
            "Уровень":       level,
        }
        for col in YEAR_COLS:
            new_row[col] = format_value(row[col])

        rows.append(new_row)
    return pd.DataFrame(rows)


def write_excel(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    # ── Цвета по уровням (как в taldau) ─────────────────────────
    LEVEL_COLORS = {
        0: "1F4E79",   # тёмно-синий  — Всего
        1: "2E75B6",   # синий        — отрасль
        2: "9DC3E6",   # светло-синий — подотрасль
        3: "D6E4F0",   # очень светлый
        4: "EBF3FB",   # почти белый
        5: "FFFFFF",   # белый
    }
    LEVEL_FONTS = {
        0: Font(bold=True, color="FFFFFF", size=10),
        1: Font(bold=True, color="FFFFFF", size=10),
        2: Font(bold=True, color="1F4E79", size=10),
        3: Font(bold=False, color="000000", size=9),
        4: Font(bold=False, color="000000", size=9),
        5: Font(bold=False, color="000000", size=9),
    }

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Заголовок ────────────────────────────────────────────────
    headers = ["Регион", "Наименование (вид деятельности)"] + YEAR_COLS
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"  # заморозить первые 2 колонки и строку заголовка

    # ── Данные ───────────────────────────────────────────────────
    regions = df["Регион"].unique()

    current_row = 2
    for region in regions:
        region_df = df[df["Регион"] == region]

        for _, row in region_df.iterrows():
            level = int(row["Уровень"])
            fill_color = LEVEL_COLORS.get(level, "FFFFFF")
            font = LEVEL_FONTS.get(level, Font(size=9))
            fill = PatternFill("solid", fgColor=fill_color)

            # Регион
            c = ws.cell(row=current_row, column=1, value=row["Регион"])
            c.font = font
            c.fill = fill
            c.border = border
            c.alignment = Alignment(vertical="center")

            # Наименование (с отступом)
            c = ws.cell(row=current_row, column=2, value=row["Наименование"])
            c.font = font
            c.fill = fill
            c.border = border
            c.alignment = Alignment(vertical="center")

            # Годы
            for col_idx, year in enumerate(YEAR_COLS, 3):
                val = row[year]
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.font = font
                c.fill = fill
                c.border = border
                c.alignment = Alignment(horizontal="right", vertical="center")
                # Формат числа с разделителями тысяч
                if isinstance(val, (int, float)) and not np.isnan(val) if isinstance(val, float) else isinstance(val, int):
                    c.number_format = "#,##0"

            ws.row_dimensions[current_row].height = 15
            current_row += 1

    # ── Ширина колонок ───────────────────────────────────────────
    ws.column_dimensions["A"].width = 25   # Регион
    ws.column_dimensions["B"].width = 65   # Наименование
    for col_idx in range(3, 3 + len(YEAR_COLS)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(output_path)
    print(f"💾 Сохранено: {output_path}")


def main():
    print("Читаю данные...")
    df = pd.read_csv(INPUT_FILE, encoding="utf-8-sig")
    print(f"  Строк: {len(df)}, Регионов: {df['Регион'].nunique()}")

    print("Форматирую...")
    formatted = build_formatted(df)

    print("Сохраняю в Excel...")
    write_excel(formatted, OUTPUT_FILE)

    print(f"\n✅ Готово!")
    print(f"   Строк:    {len(formatted)}")
    print(f"   Регионов: {formatted['Регион'].nunique()}")
    print(f"   Файл:     {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
