"""
Объединяет два CSV-датасета в один Excel-файл KAZAKHSTAN_MERGED.xlsx

Источники:
  - ALL_KAZAKHSTAN_FINAL.csv  → годовые данные   (2019–2024)
  - output/MONTHLY_FINAL.csv  → помесячные данные (Янв 2025 → Янв-Фев 2026)

Структура выходного файла:
  - Лист "Навигация" с гиперссылками
  - Отдельный лист на каждую область:
    Район | Наименование | 2019 | 2020 | ... | 2024 | Янв 2025 | Янв-Фев 2025 | ...

Запуск:
    python to_excel_merged.py
    python to_excel_merged.py --final data/FINAL.csv --monthly data/MONTHLY.csv --output out.xlsx
"""

import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CLI-аргументы ──────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--final",   default="ALL_KAZAKHSTAN_FULL.csv")
parser.add_argument("--monthly", default="MONTHLY_FINAL.csv")
parser.add_argument("--output",  default="KAZAKHSTAN.xlsx")
args = parser.parse_args()

FINAL_FILE   = args.final
MONTHLY_FILE = args.monthly
OUTPUT_FILE  = args.output

INDENT = "    "

# ── Маппинг колонок ────────────────────────────────────────────────────────────

# Годовые: колонка в CSV → заголовок в Excel
YEAR_MAP = {
    "y122019": "2019",
    "y122020": "2020",
    "y122021": "2021",
    "y122022": "2022",
    "y122023": "2023",
    "y122024": "2024",
}

# Помесячные: колонка в CSV → заголовок в Excel
MONTHLY_MAP = {
    "y012025": "Янв 2025",
    "y022025": "Янв-Фев 2025",
    "y032025": "Янв-Мар 2025",
    "y042025": "Янв-Апр 2025",
    "y052025": "Янв-Май 2025",
    "y062025": "Янв-Июн 2025",
    "y072025": "Янв-Июл 2025",
    "y082025": "Янв-Авг 2025",
    "y092025": "Янв-Сен 2025",
    "y102025": "Янв-Окт 2025",
    "y112025": "Янв-Ноя 2025",
    "y122025": "Янв-Дек 2025",
    "y012026": "Янв 2026",
    "y022026": "Янв-Фев 2026",
}

# Цвета заголовков
HDR_KEY     = "1F4E79"   # Район / Наименование
HDR_ANNUAL  = "2E75B6"   # Годовые колонки
HDR_MONTHLY = "14375E"   # Месячные колонки

# Стили уровней иерархии строк
LEVEL_STYLE = {
    0: ("1F4E79", "FFFFFF", True),
    1: ("2E75B6", "FFFFFF", True),
    2: ("9DC3E6", "1F4E79", True),
    3: ("D6E4F0", "000000", False),
    4: ("EBF3FB", "000000", False),
    5: ("FFFFFF", "000000", False),
}

BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ── Вспомогательные функции ────────────────────────────────────────────────────

def format_val(val):
    """Тенге → тысячи тенге. 'x'/'х' оставляем как есть."""
    s = str(val).strip()
    if s.lower() in ("x", "х"):
        return "x"
    if s.lower() in ("nan", "none", "nat", ""):
        return None
    try:
        f = float(s.replace(" ", "").replace(",", "."))
        return None if np.isnan(f) else int(round(f / 1000))
    except (ValueError, TypeError):
        return None


def safe_title(name: str) -> str:
    for ch in r"\/*?:[]":
        name = name.replace(ch, "")
    return name[:31]


def style_header(cell, bg: str):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color="FFFFFF", size=9)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_cell(cell, level: int, h_align="left"):
    bg, fg, bold = LEVEL_STYLE.get(level, ("FFFFFF", "000000", False))
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, color=fg, size=9)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal=h_align, vertical="center")


# ── Чтение CSV ─────────────────────────────────────────────────────────────────

def load_final(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, encoding="utf-8-sig")
    # Оставляем только нужные колонки
    keep = ["oblast", "region", "depth", "text"] + [c for c in YEAR_MAP if c in df.columns]
    missing = [c for c in YEAR_MAP if c not in df.columns]
    if missing:
        print(f"  ⚠️  Годовые колонки не найдены в CSV: {missing}")
    return df[keep]


def load_monthly(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, encoding="utf-8-sig")
    keep = ["oblast", "region", "depth", "text"] + [c for c in MONTHLY_MAP if c in df.columns]
    missing = [c for c in MONTHLY_MAP if c not in df.columns]
    if missing:
        print(f"  ⚠️  Месячные колонки не найдены в CSV: {missing}")
    return df[keep]


# ── Запись листа области ───────────────────────────────────────────────────────

def write_oblast_sheet(ws, final_oblast: pd.DataFrame, monthly_oblast: pd.DataFrame):
    """
    Объединяет годовые и месячные данные по ключу (region, depth, text)
    и записывает на лист ws.
    """
    key = ["region", "depth", "text"]

    # Активные годовые и месячные колонки (только те, что есть в данных)
    annual_cols  = [c for c in YEAR_MAP   if c in final_oblast.columns]
    monthly_cols = [c for c in MONTHLY_MAP if c in monthly_oblast.columns]

    merged = pd.merge(
        final_oblast,
        monthly_oblast[key + monthly_cols],
        on=key,
        how="outer"
    )
    # Восстанавливаем порядок по оригинальным данным
    merged = merged.sort_values(by=key).reset_index(drop=True)

    # ── Строка заголовков ──
    header_annual  = [YEAR_MAP[c]    for c in annual_cols]
    header_monthly = [MONTHLY_MAP[c] for c in monthly_cols]
    headers = ["Район", "Наименование (вид деятельности)"] + header_annual + header_monthly

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        if ci <= 2:
            bg = HDR_KEY
        elif ci <= 2 + len(annual_cols):
            bg = HDR_ANNUAL
        else:
            bg = HDR_MONTHLY
        style_header(cell, bg)

    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "C2"

    # ── Строки данных ──
    all_data_cols = annual_cols + monthly_cols
    ri = 2
    for _, row in merged.iterrows():
        level = min(int(row.get("depth", 0)), 5)
        name  = INDENT * level + str(row.get("text", "")).strip()

        c = ws.cell(row=ri, column=1, value=str(row.get("region", "")))
        style_cell(c, level)

        c = ws.cell(row=ri, column=2, value=name)
        style_cell(c, level)

        for ci, col in enumerate(all_data_cols, 3):
            val = format_val(row.get(col, None))
            c = ws.cell(row=ri, column=ci, value=val)
            style_cell(c, level, h_align="right")
            if isinstance(val, int):
                c.number_format = "#,##0"

        ws.row_dimensions[ri].height = 14
        ri += 1

    # ── Ширина колонок ──
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 65
    for ci in range(3, 3 + len(all_data_cols)):
        ltr = get_column_letter(ci)
        ws.column_dimensions[ltr].width = 15 if ci <= 2 + len(annual_cols) else 18

    # ── Ссылка «Назад» ──
    back = ws.cell(row=ri + 1, column=1, value="← Назад к навигации")
    back.hyperlink = "#'Навигация'!A1"
    back.font      = Font(color="2E75B6", underline="single", bold=True, size=9)
    back.alignment = Alignment(vertical="center")

    return ri - 2  # кол-во строк данных


# ── Навигационный лист ─────────────────────────────────────────────────────────

def write_nav_sheet(ws, info_map: dict):
    ws.title = "Навигация"
    ws.sheet_view.showGridLines = False

    t = ws.cell(row=1, column=1,
                value="Инвестиции в основной капитал — Казахстан")
    t.font = Font(bold=True, color="1F4E79", size=14)
    t.alignment = Alignment(vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 38

    first_annual  = list(YEAR_MAP.values())[0]
    last_monthly  = list(MONTHLY_MAP.values())[-1]
    sub = ws.cell(row=2, column=1,
                  value=f"Данные: {first_annual}  →  {last_monthly}")
    sub.font = Font(italic=True, color="808080", size=10)
    ws.row_dimensions[2].height = 20

    for ci, h in enumerate(["Область / Город", "Районов", "Строк", "Перейти →"], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.border    = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    alt = ["EBF3FB", "FFFFFF"]
    for i, (oblast, info) in enumerate(info_map.items(), 4):
        fill = PatternFill("solid", fgColor=alt[i % 2])

        c = ws.cell(row=i, column=1, value=oblast)
        c.fill = fill; c.font = Font(bold=True, color="1F4E79", size=10)
        c.border = BORDER; c.alignment = Alignment(vertical="center")

        c = ws.cell(row=i, column=2, value=info["districts"])
        c.fill = fill; c.font = Font(size=10); c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=i, column=3, value=info["rows"])
        c.fill = fill; c.font = Font(size=10); c.border = BORDER
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=i, column=4, value=f"→ {info['sheet']}")
        c.hyperlink = f"#{info['sheet']}!A1"
        c.fill = fill
        c.font = Font(color="1F4E79", bold=True, underline="single", size=10)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 38


# ── Главная функция ────────────────────────────────────────────────────────────

def main():
    print(f"📂 Читаю годовые данные: {FINAL_FILE}")
    df_final = load_final(FINAL_FILE)
    print(f"   Строк: {len(df_final)}  Областей: {df_final['oblast'].nunique()}  "
          f"Районов: {df_final['region'].nunique()}")

    print(f"\n📂 Читаю помесячные данные: {MONTHLY_FILE}")
    df_monthly = load_monthly(MONTHLY_FILE)
    print(f"   Строк: {len(df_monthly)}  Областей: {df_monthly['oblast'].nunique()}  "
          f"Районов: {df_monthly['region'].nunique()}")

    # Все уникальные области из обоих датасетов, порядок сохраняем
    all_oblasts = list(dict.fromkeys(
        list(df_final["oblast"].unique()) +
        list(df_monthly["oblast"].unique())
    ))
    print(f"\n📊 Создаю листы ({len(all_oblasts)} областей)...")

    wb = Workbook()
    wb.remove(wb.active)
    nav_ws   = wb.create_sheet("Навигация")
    info_map = {}

    for oblast in all_oblasts:
        oblast_final   = df_final  [df_final  ["oblast"] == oblast] if oblast in df_final  ["oblast"].values else pd.DataFrame()
        oblast_monthly = df_monthly[df_monthly["oblast"] == oblast] if oblast in df_monthly["oblast"].values else pd.DataFrame()

        # Гарантируем нужные колонки даже для пустых датафреймов
        key = ["oblast", "region", "depth", "text"]
        if oblast_final.empty:
            oblast_final = pd.DataFrame(columns=key + list(YEAR_MAP.keys()))
        if oblast_monthly.empty:
            oblast_monthly = pd.DataFrame(columns=key + list(MONTHLY_MAP.keys()))

        sheet_title = safe_title(oblast)
        ws = wb.create_sheet(title=sheet_title)
        n_rows = write_oblast_sheet(ws, oblast_final, oblast_monthly)

        districts = len(set(
            list(oblast_final ["region"].dropna()) +
            list(oblast_monthly["region"].dropna())
        ))
        info_map[oblast] = {"sheet": sheet_title, "districts": districts, "rows": n_rows}
        print(f"   ✓ {oblast:<42} {districts:>3} р-нов  {n_rows:>6} строк")

    write_nav_sheet(nav_ws, info_map)

    print(f"\n💾 Сохраняю {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    total = sum(v["rows"] for v in info_map.values())
    print(f"✅ Готово! Листов: {len(info_map) + 1}  Строк данных: {total}")


if __name__ == "__main__":
    main()