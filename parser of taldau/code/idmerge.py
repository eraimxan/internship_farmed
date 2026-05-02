"""
add_id_columns.py — добавляет ID и Parent_ID в KAZAKHSTAN_MERGED_Patched.xlsx
Depth определяется по ведущим пробелам (indent // 4), покрытие 100%.
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_XLSX  = "KAZAKHSTAN.xlsx"
INPUT_CSV   = "ALL_KAZAKHSTAN_FULL.csv"
OUTPUT_XLSX = "KAZAKHSTAN_WITH_IDS.xlsx"

print("Строим иерархию id из CSV...")
df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig", low_memory=False)

def compute_parent_ids(group):
    rows = group.reset_index(drop=True)
    parent_ids = [None] * len(rows)
    for i in range(len(rows)):
        d = rows.at[i, "depth"]
        if d > 0:
            for j in range(i - 1, -1, -1):
                if rows.at[j, "depth"] == d - 1:
                    parent_ids[i] = int(rows.at[j, "id"])
                    break
    rows["parent_id"] = parent_ids
    return rows

groups = []
for (oblast, region), grp in df.groupby(["oblast", "region"], sort=False):
    groups.append(compute_parent_ids(grp))
df_full = pd.concat(groups, ignore_index=True)
df_full["text_norm"] = df_full["text"].str.strip()

# Lookup 1: точный (oblast, region, text_norm, depth)
exact_lookup = {}
for _, row in df_full.iterrows():
    key = (str(row["oblast"]), str(row["region"]), row["text_norm"], int(row["depth"]))
    if key not in exact_lookup:
        pid = int(row["parent_id"]) if pd.notna(row["parent_id"]) else None
        exact_lookup[key] = (int(row["id"]), pid)

# Lookup 2: глобальный fallback (text_norm, depth)
global_lookup = {}
for _, row in df_full.iterrows():
    key = (row["text_norm"], int(row["depth"]))
    if key not in global_lookup:
        pid = int(row["parent_id"]) if pd.notna(row["parent_id"]) else None
        global_lookup[key] = (int(row["id"]), pid)

print(f"  Точный lookup: {len(exact_lookup):,} | Глобальный: {len(global_lookup):,}")

FONT_COLOR_BY_FILL = {
    "1F4E79": "FFFFFF",
    "2E75B6": "FFFFFF",
    "9DC3E6": "1F4E79",
    "D6E4F0": "000000",
}

def get_fill_hex(cell):
    try:
        if cell.fill.fill_type == "solid":
            rgb = cell.fill.fgColor.rgb
            return rgb[-6:] if len(rgb) >= 6 else "FFFFFF"
    except:
        pass
    return "FFFFFF"

def make_cell_style(fill_hex, bold, size=9):
    fc = FONT_COLOR_BY_FILL.get(fill_hex, "000000")
    return (Font(name="Arial", size=size, bold=bold, color=fc),
            PatternFill("solid", fgColor=fill_hex),
            Alignment(horizontal="center", vertical="center"))

HDR_FONT  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

print(f"\nОткрываем {INPUT_XLSX}...")
wb = load_workbook(INPUT_XLSX)

stats = {"exact": 0, "fallback": 0, "missing": 0, "total": 0}

for sheet_name in wb.sheetnames:
    if sheet_name == "Навигация":
        continue

    ws = wb[sheet_name]
    oblast = sheet_name
    print(f"  [{sheet_name}] {ws.max_row} строк...", end=" ", flush=True)

    ws.insert_cols(1, 2)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 11
    ws.freeze_panes = "E2"

    for col, label in ((1, "ID"), (2, "Parent ID")):
        c = ws.cell(row=1, column=col, value=label)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN

    current_region = None
    sh_exact = sh_fallback = sh_miss = 0

    for row_idx in range(2, ws.max_row + 1):
        region_cell = ws.cell(row=row_idx, column=3)
        name_cell   = ws.cell(row=row_idx, column=4)

        if region_cell.value:
            current_region = str(region_cell.value).strip()

        name_val = str(name_cell.value) if name_cell.value else ""
        if not name_val or name_val == "None":
            continue

        stripped = name_val.strip()
        indent   = len(name_val) - len(stripped)
        depth    = indent // 4

        row_id = row_pid = None
        k1 = (oblast, current_region, stripped, depth)
        k2 = (stripped, depth)

        if k1 in exact_lookup:
            row_id, row_pid = exact_lookup[k1]
            sh_exact += 1
        elif k2 in global_lookup:
            row_id, row_pid = global_lookup[k2]
            sh_fallback += 1
        else:
            sh_miss += 1

        fill_hex = get_fill_hex(region_cell)
        is_bold  = bool(region_cell.font.bold) if region_cell.font else False
        fnt_sz   = float(region_cell.font.size) if (region_cell.font and region_cell.font.size) else 9.0

        fnt, fll, aln = make_cell_style(fill_hex, is_bold, fnt_sz)

        id_cell = ws.cell(row=row_idx, column=1, value=row_id)
        id_cell.font = fnt; id_cell.fill = fll; id_cell.alignment = aln

        pid_cell = ws.cell(row=row_idx, column=2, value=row_pid)
        pid_cell.font = fnt; pid_cell.fill = fll; pid_cell.alignment = aln

    stats["exact"]    += sh_exact
    stats["fallback"] += sh_fallback
    stats["missing"]  += sh_miss
    stats["total"]    += sh_exact + sh_fallback + sh_miss
    print(f"exact={sh_exact} fallback={sh_fallback} miss={sh_miss}")

print(f"\nСохраняем {OUTPUT_XLSX}...")
wb.save(OUTPUT_XLSX)

covered = stats["exact"] + stats["fallback"]
print(f"\n{'='*60}")
print(f" ГОТОВО!")
print(f"   Строк обработано:    {stats['total']:,}")
print(f"   Точных совпадений:   {stats['exact']:,}")
print(f"   Глобальный fallback: {stats['fallback']:,}")
print(f"   Без id (пусто):      {stats['missing']:,}")
print(f"   Покрытие:            {covered/stats['total']*100:.1f}%")
print(f"   Файл:                {OUTPUT_XLSX}")
print(f"{'='*60}")