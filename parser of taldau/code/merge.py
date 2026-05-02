"""
Объединяет KAZAKHSTAN_FINAL.xlsx (годовые 2019–2024)
         и MONTHLY_KAZAKHSTAN.xlsx (помесячные 2025–2026)
→ KAZAKHSTAN_MERGED.xlsx

Запуск:
    python merge_kazakhstan.py
    python merge_kazakhstan.py --final path/to/final.xlsx --monthly path/to/monthly.xlsx --output out.xlsx
"""

import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Аргументы CLI ─────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--final",   default="output/KAZAKHSTAN_FINAL.xlsx")
parser.add_argument("--monthly", default="MONTHLY_KAZAKHSTAN.xlsx")
parser.add_argument("--output",  default="KAZAKHSTAN_MERGED.xlsx")
args = parser.parse_args()

FINAL_FILE   = args.final
MONTHLY_FILE = args.monthly
OUTPUT_FILE  = args.output

# ── Стили ─────────────────────────────────────────────────────────────────────
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

LEVEL_STYLE = {
    0: ("1F4E79", "FFFFFF", True),
    1: ("2E75B6", "FFFFFF", True),
    2: ("9DC3E6", "1F4E79", True),
    3: ("D6E4F0", "000000", False),
    4: ("EBF3FB", "000000", False),
    5: ("FFFFFF", "000000", False),
}

# Цвета заголовков: ключевые | годовые | месячные
HDR_KEY     = "1F4E79"
HDR_ANNUAL  = "2E75B6"
HDR_MONTHLY = "14375E"


def safe_title(name: str) -> str:
    for ch in r"\/*?:[]":
        name = name.replace(ch, "")
    return name[:31]


def style_header(cell, bg: str, wrap=True):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color="FFFFFF", size=9)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def style_data(cell, level: int, h_align="left"):
    bg, fg, bold = LEVEL_STYLE.get(level, ("FFFFFF", "000000", False))
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, color=fg, size=9)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=False)


# ── Чтение данных ─────────────────────────────────────────────────────────────
def read_sheet(path: str, sheet: str) -> pd.DataFrame:
    """Читает лист, первая строка = заголовок."""
    return pd.read_excel(path, sheet_name=sheet, header=0, dtype=str)


def get_period_cols(df: pd.DataFrame, skip=2) -> list[str]:
    """Возвращает имена колонок начиная с позиции skip (после ключевых)."""
    return list(df.columns[skip:])


# ── Запись листа области ──────────────────────────────────────────────────────
def write_oblast_sheet(ws, df_final: pd.DataFrame, df_monthly: pd.DataFrame):
    key_cols = ["Район", "Наименование (вид деятельности)"]

    annual_cols  = get_period_cols(df_final)
    monthly_cols = get_period_cols(df_monthly)

    # Объединяем по ключевым колонкам
    merged = pd.merge(
        df_final,
        df_monthly[key_cols + monthly_cols],
        on=key_cols,
        how="outer"
    )

    all_data_cols = annual_cols + monthly_cols
    headers = key_cols + all_data_cols

    # Строка заголовков
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        if ci <= 2:
            bg = HDR_KEY
        elif ci <= 2 + len(annual_cols):
            bg = HDR_ANNUAL
        else:
            bg = HDR_MONTHLY
        style_header(cell, bg)
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "C2"

    # Строки данных
    for ri, (_, row) in enumerate(merged.iterrows(), 2):
        # Определяем уровень по отступам в названии
        name_raw = str(row.get("Наименование (вид деятельности)", "") or "")
        level = len(name_raw) - len(name_raw.lstrip())
        level = min(level // 4, 5)

        cell = ws.cell(row=ri, column=1, value=row.get("Район", ""))
        style_data(cell, level)

        cell = ws.cell(row=ri, column=2, value=name_raw)
        style_data(cell, level)

        for ci, col in enumerate(all_data_cols, 3):
            raw = row.get(col, None)
            val = _parse_val(raw)
            cell = ws.cell(row=ri, column=ci, value=val)
            style_data(cell, level, h_align="right")
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        ws.row_dimensions[ri].height = 14

    # Ширина колонок
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 62
    for ci in range(3, 3 + len(all_data_cols)):
        ws.column_dimensions[get_column_letter(ci)].width = 17

    # Ссылка назад
    back = ws.cell(row=len(merged) + 3, column=1, value="← Назад к навигации")
    back.hyperlink = "#'Навигация'!A1"
    back.font      = Font(color="2E75B6", underline="single", bold=True, size=9)
    back.alignment = Alignment(vertical="center")

    return len(merged)


def _parse_val(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if s.lower() in ("x", "х", "nan", "none", "", "nat"):
        return None
    try:
        f = float(s.replace(" ", "").replace(",", "."))
        return None if np.isnan(f) else int(round(f))
    except (ValueError, TypeError):
        return s or None


# ── Навигационный лист ────────────────────────────────────────────────────────
def write_nav_sheet(ws, info_map: dict, period_range: tuple):
    ws.sheet_view.showGridLines = False

    title_val = "Инвестиции в основной капитал — Казахстан"
    t = ws.cell(row=1, column=1, value=title_val)
    t.font = Font(bold=True, color="1F4E79", size=13)
    t.alignment = Alignment(vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 36

    first, last = period_range
    sub = ws.cell(row=2, column=1, value=f"Данные: {first}  →  {last}")
    sub.font = Font(italic=True, color="808080", size=10)
    ws.row_dimensions[2].height = 20

    for ci, h in enumerate(["Область / Город", "Районов", "Строк", "Перейти →"], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    alt = ["EBF3FB", "FFFFFF"]
    for i, (oblast, info) in enumerate(info_map.items(), 4):
        fill = PatternFill("solid", fgColor=alt[i % 2])

        c = ws.cell(row=i, column=1, value=oblast)
        c.fill = fill; c.font = Font(bold=True, color="1F4E79", size=10)
        c.border = BORDER; c.alignment = Alignment(vertical="center")

        c = ws.cell(row=i, column=2, value=info["districts"])
        c.fill = fill; c.font = Font(size=10); c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=i, column=3, value=info["rows"])
        c.fill = fill; c.font = Font(size=10); c.border = BORDER
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=i, column=4, value=f"→ {info['sheet']}")
        c.hyperlink = f"#{info['sheet']}!A1"
        c.fill = fill
        c.font = Font(color="1F4E79", bold=True, underline="single", size=10)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 38


# ── Главная функция ────────────────────────────────────────────────────────────
def main():
    print(f"📂 Читаю {FINAL_FILE}...")
    xf_final = pd.ExcelFile(FINAL_FILE)

    print(f"📂 Читаю {MONTHLY_FILE}...")
    xf_monthly = pd.ExcelFile(MONTHLY_FILE)

    sheets_final   = xf_final.sheet_names
    sheets_monthly = xf_monthly.sheet_names

    oblast_sheets = [s for s in sheets_final if s != "Навигация"]

    # Определяем диапазон периодов для навигации
    sample_final   = read_sheet(FINAL_FILE,   oblast_sheets[0])
    sample_monthly = read_sheet(MONTHLY_FILE, oblast_sheets[0])
    first_period   = get_period_cols(sample_final)[0]
    last_period    = get_period_cols(sample_monthly)[-1]

    wb = Workbook()
    wb.remove(wb.active)
    nav_ws = wb.create_sheet("Навигация")
    info_map = {}

    print(f"\n📊 Создаю листы ({len(oblast_sheets)} областей)...")
    for oblast in oblast_sheets:
        df_f = read_sheet(FINAL_FILE,   oblast) if oblast in sheets_final   else pd.DataFrame()
        df_m = read_sheet(MONTHLY_FILE, oblast) if oblast in sheets_monthly else pd.DataFrame()

        if df_f.empty and df_m.empty:
            print(f"   ⚠️  Пропускаю {oblast} — нет данных ни в одном файле")
            continue

        # Если один файл не содержит лист — создаём пустой с нужными колонками
        key_cols = ["Район", "Наименование (вид деятельности)"]
        if df_f.empty:
            df_f = pd.DataFrame(columns=key_cols)
        if df_m.empty:
            df_m = pd.DataFrame(columns=key_cols)

        sheet_title = safe_title(oblast)
        ws = wb.create_sheet(title=sheet_title)
        n_rows = write_oblast_sheet(ws, df_f, df_m)

        # Кол-во уникальных районов из объединённых данных
        merged_districts = set(
            list(df_f.get("Район", pd.Series(dtype=str)).dropna()) +
            list(df_m.get("Район", pd.Series(dtype=str)).dropna())
        )
        info_map[oblast] = {
            "sheet":     sheet_title,
            "districts": len(merged_districts),
            "rows":      n_rows,
        }
        print(f"   ✓ {oblast:<42} {info_map[oblast]['districts']:>3} р-нов  {n_rows:>5} строк")

    write_nav_sheet(nav_ws, info_map, (first_period, last_period))

    print(f"\n💾 Сохраняю {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    total_rows = sum(v["rows"] for v in info_map.values())
    print(f"✅ Готово! Листов: {len(info_map)+1}  Строк данных: {total_rows}")


if __name__ == "__main__":
    main()