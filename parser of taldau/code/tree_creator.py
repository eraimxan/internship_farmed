import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

LEVEL_COLORS = {
    0: ("1F3864", "FFFFFF"),
    1: ("2E75B6", "FFFFFF"),
    2: ("BDD7EE", "1F3864"),
    3: ("DDEBF7", "1F3864"),
    4: ("EEF4FB", "2E3A4E"),
    5: ("FFFFFF", "2E3A4E"),
}

INDENT_PER_LEVEL = "    "

def compute_depths(df):
    id_to_parent = {}
    for _, row in df.iterrows():
        if pd.notna(row['ID']) and pd.notna(row['Parent ID']):
            id_to_parent[int(row['ID'])] = int(row['Parent ID'])
    depths = {}
    for node_id in df['ID'].dropna().astype(int):
        depth = 0
        current = node_id
        visited = set()
        while current in id_to_parent and current not in visited:
            visited.add(current)
            current = id_to_parent[current]
            depth += 1
        depths[node_id] = depth
    return depths

def sort_by_tree(df):
    id_to_parent = {}
    children_of = defaultdict(list)
    for _, row in df.iterrows():
        node_id = int(row['ID']) if pd.notna(row['ID']) else None
        parent_id = int(row['Parent ID']) if pd.notna(row['Parent ID']) else None
        if node_id is not None:
            id_to_parent[node_id] = parent_id
            if parent_id is not None:
                children_of[parent_id].append(node_id)
    all_ids = set(df['ID'].dropna().astype(int))
    roots = [i for i in all_ids if id_to_parent.get(i) is None]
    result_ids = []
    def dfs(node_id):
        result_ids.append(node_id)
        for child in children_of.get(node_id, []):
            dfs(child)
    for root in roots:
        dfs(root)
    id_to_rows = defaultdict(list)
    for idx, row in df.iterrows():
        if pd.notna(row['ID']):
            id_to_rows[int(row['ID'])].append(row)
    ordered_rows = []
    for node_id in result_ids:
        ordered_rows.extend(id_to_rows[node_id])
    return pd.DataFrame(ordered_rows)

xl = pd.read_excel('KAZAKHSTAN_WITH_IDS.xlsx', sheet_name=None)
sheet_names = list(xl.keys())

wb = openpyxl.Workbook()
wb.remove(wb.active)

data_cols = ['2019','2020','2021','2022','2023','2024',
             'Янв 2025','Янв-Фев 2025','Янв-Мар 2025','Янв-Апр 2025',
             'Янв-Май 2025','Янв-Июн 2025','Янв-Июл 2025','Янв-Авг 2025',
             'Янв-Сен 2025','Янв-Окт 2025','Янв-Ноя 2025','Янв-Дек 2025',
             'Янв 2026','Янв-Фев 2026']

def write_row(ws, excel_row, row, depth, data_cols):
    node_id  = int(row['ID']) if pd.notna(row['ID']) else None
    parent_id = int(row['Parent ID']) if pd.notna(row['Parent ID']) else None
    bg_color, fg_color = LEVEL_COLORS.get(min(depth, 5), ("FFFFFF", "000000"))
    fill = PatternFill("solid", start_color=bg_color)
    font_bold = depth <= 1

    def sc(col, val, num_fmt=None, halign=None):
        c = ws.cell(excel_row, col, val)
        c.fill = fill
        c.font = Font(color=fg_color, bold=font_bold, size=9)
        if halign:
            c.alignment = Alignment(horizontal=halign, vertical='center')
        else:
            c.alignment = Alignment(vertical='center')
        if num_fmt:
            c.number_format = num_fmt
        return c

    sc(1, node_id,   halign='right')
    sc(2, parent_id, halign='right')
    sc(3, depth,     halign='center')
    sc(4, str(row['Район']) if pd.notna(row['Район']) else '')
    
    name = str(row['Наименование (вид деятельности)']) if pd.notna(row['Наименование (вид деятельности)']) else ''
    sc(5, INDENT_PER_LEVEL * depth + name)

    for i, col_name in enumerate(data_cols):
        val = row.get(col_name)
        c = ws.cell(excel_row, 6 + i)
        c.fill = fill
        c.alignment = Alignment(horizontal='right', vertical='center')
        if pd.isna(val) or val == '' or val is None:
            c.value = None
            c.font = Font(color=fg_color, size=9)
        elif str(val).strip() == 'x':
            c.value = 'x'
            c.font = Font(color=fg_color, italic=True, size=9)
        else:
            try:
                c.value = int(float(str(val)))
                c.number_format = '#,##0'
                c.font = Font(color=fg_color, bold=font_bold, size=9)
            except:
                c.value = val
                c.font = Font(color=fg_color, size=9)

    ws.row_dimensions[excel_row].height = 16

for sheet_name in sheet_names:
    df_orig = xl[sheet_name]

    if sheet_name == 'Навигация':
        ws = wb.create_sheet('Навигация')
        for r_idx, row in enumerate(df_orig.itertuples(index=False), 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(r_idx, c_idx, val)
        continue

    print(f"Обрабатываем: {sheet_name}")

    ws = wb.create_sheet(sheet_name)
    # Группировка: кнопка СВЕРХУ от группы (стандарт Excel)
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # Заголовок
    headers = ['ID','Parent ID','Уровень','Район','Наименование (вид деятельности)'] + data_cols
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, c_idx, h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'F2'

    current_row = 2

    for district, group in df_orig.groupby('Район', sort=False):
        depths = compute_depths(group)
        try:
            sorted_group = sort_by_tree(group)
        except:
            sorted_group = group

        # Записываем строки и собираем позиции по уровням
        row_info = []  # (excel_row, depth)

        for _, row in sorted_group.iterrows():
            node_id = int(row['ID']) if pd.notna(row['ID']) else None
            depth = depths.get(node_id, 0) if node_id else 0
            write_row(ws, current_row, row, depth, data_cols)
            row_info.append((current_row, depth))
            current_row += 1

        # Теперь выставляем outline_level правильно:
        # Для каждого родительского узла находим диапазон его детей
        # и группируем их как единый блок
        #
        # Логика: строка с depth=D группируется на уровне D-1
        # (т.е. depth=1 → outline 0, depth=2 → outline 1, ...)
        # Это означает что при нажатии на кнопку уровня N
        # скрываются все строки с outline >= N
        
        for excel_row, depth in row_info:
            if depth > 0:
                # outline_level = depth (1-based, максимум 7 в Excel)
                ws.row_dimensions[excel_row].outline_level = min(depth, 7)
                ws.row_dimensions[excel_row].hidden = False

    # Ширина колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 55
    for i in range(len(data_cols)):
        ws.column_dimensions[get_column_letter(6 + i)].width = 14

print("Сохраняем...")
wb.save('KAZAKHSTAN_TREE_V2.xlsx')
print("Готово!")