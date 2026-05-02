"""
build_excel.py — Казахстан: Инвестиции в основной капитал
Объединяет ALL_KAZAKHSTAN_PATCHED.csv (годовые) и MONTHLY_FINAL.csv (месячные)
в форматированный Excel-файл по образцу KAZAKHSTAN_FINAL_PATCHED.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER

# ─── Цвета (из образца) ──────────────────────────────────────────────────────
C_DARK_BLUE   = "1F4E79"   # заголовок, Всего строка
C_MID_BLUE    = "2E75B6"   # depth=1 (секции)
C_LIGHT_BLUE  = "9DC3E6"   # depth=2
C_PALE_BLUE   = "D6E4F0"   # depth=3
C_LIGHTEST    = "EBF3FB"   # depth=4+, нечётные
C_WHITE       = "FFFFFF"   # depth=4+, чётные
C_NAV_EVEN    = "EBF3FB"
C_WHITE_FONT  = "FFFFFF"
C_DARK_FONT   = "000000"
C_BLUE_FONT   = "1F4E79"
C_GRAY_FONT   = "808080"

# Месяцы для заголовков
MONTH_NAMES = {
    "01":"Янв","02":"Фев","03":"Мар","04":"Апр","05":"Май","06":"Июн",
    "07":"Июл","08":"Авг","09":"Сен","10":"Окт","11":"Ноя","12":"Дек"
}

ANNUAL_COL_MAP = {
    "y122019": "2019", "y122020": "2020", "y122021": "2021",
    "y122022": "2022", "y122023": "2023", "y122024": "2024",
}

MONTHLY_COLS = [
    "y012025","y022025","y032025","y042025","y052025","y062025",
    "y072025","y082025","y092025","y102025","y112025","y122025",
    "y012026","y022026",
]

def col_header(code):
    """y012025 → Янв 2025"""
    m, y = code[1:3], code[3:]
    return f"{MONTH_NAMES[m]} {y}"

def safe_int(v):
    """Преобразует значение в int, деля на 1000 (тенге → тыс.тенге). None/NaN → None."""
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return None
        f = float(str(v))
        if np.isnan(f):
            return None
        return round(f / 1000)
    except:
        return None

def safe_int_monthly(v):
    """Monthly уже в тенге — делим на 1000."""
    return safe_int(v)

def depth_style(depth, row_idx):
    """Возвращает (fill_hex, font_hex, bold) по глубине строки."""
    if depth == 0:
        return C_DARK_BLUE, C_WHITE_FONT, True
    if depth == 1:
        return C_MID_BLUE, C_WHITE_FONT, True
    if depth == 2:
        return C_LIGHT_BLUE, C_BLUE_FONT, True
    if depth == 3:
        return C_PALE_BLUE, C_DARK_FONT, False
    # depth 4+: чередование
    return (C_LIGHTEST if row_idx % 2 == 0 else C_WHITE), C_DARK_FONT, False

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(hex_color, bold=False, size=9):
    return Font(name="Arial", color=hex_color, bold=bold, size=size)

def border_thin():
    s = Side(style="thin", color="BDD7EE")
    return Border(bottom=s)

def write_sheet(wb, oblast_name, df_oblast):
    """Записывает один лист области."""
    ws = wb.create_sheet(title=oblast_name)

    # Колонки: A=Район, B=Наименование, C..H=годы, I..V=месяцы
    all_data_cols = list(ANNUAL_COL_MAP.keys()) + MONTHLY_COLS
    header_labels = (
        ["Район", "Наименование (вид деятельности)"]
        + list(ANNUAL_COL_MAP.values())
        + [col_header(c) for c in MONTHLY_COLS]
    )
    n_cols = len(header_labels)

    # Ширины колонок
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 65
    for i in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # ── Заголовочная строка ──────────────────────────────────────────
    hdr_fill = make_fill(C_DARK_BLUE)
    hdr_font = make_font(C_WHITE_FONT, bold=True, size=10)
    for ci, label in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"

    # ── Данные ──────────────────────────────────────────────────────
    regions = df_oblast["region"].unique()
    excel_row = 2

    for region in regions:
        df_reg = df_oblast[df_oblast["region"] == region].reset_index(drop=True)
        row_counter = 0  # для чередования цвета

        for _, row in df_reg.iterrows():
            depth = int(row.get("depth", 4))
            fill_hex, font_hex, bold = depth_style(depth, row_counter)
            fill = make_fill(fill_hex)
            fnt  = make_font(font_hex, bold=bold, size=9)

            # Отступ по глубине
            indent = "    " * depth
            name_val = f"{indent}{row['text']}" if pd.notna(row.get("text")) else ""

            # Колонка A — Район
            ca = ws.cell(row=excel_row, column=1, value=region)
            ca.font = fnt; ca.fill = fill
            ca.alignment = Alignment(horizontal="left", vertical="center")

            # Колонка B — Наименование
            cb = ws.cell(row=excel_row, column=2, value=name_val)
            cb.font = fnt; cb.fill = fill
            cb.alignment = Alignment(horizontal="left", vertical="center")

            # Числовые колонки
            col_idx = 3
            for col_key in all_data_cols:
                raw = row.get(col_key)
                val = safe_int(raw)
                cc = ws.cell(row=excel_row, column=col_idx, value=val)
                cc.font = fnt; cc.fill = fill
                cc.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    cc.number_format = '#,##0'
                col_idx += 1

            ws.row_dimensions[excel_row].height = 15
            excel_row += 1
            row_counter += 1

    return excel_row - 2  # кол-во строк данных


def write_nav_sheet(wb, oblast_stats):
    """Навигационный лист."""
    ws = wb.active
    ws.title = "Навигация"

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 35

    # Заголовки
    t1 = ws.cell(row=1, column=1, value="Инвестиции в основной капитал — Казахстан")
    t1.font = Font(name="Arial", bold=True, color=C_BLUE_FONT, size=14)

    t2 = ws.cell(row=2, column=1, value="Нажмите на ссылку для перехода к данным области")
    t2.font = Font(name="Arial", color=C_GRAY_FONT, size=10)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # Шапка таблицы
    hdr_fill = make_fill(C_DARK_BLUE)
    for ci, label in enumerate(["Область / Город", "Районов", "Строк (данные)", "Перейти →"], 1):
        c = ws.cell(row=3, column=ci, value=label)
        c.font = make_font(C_WHITE_FONT, bold=True, size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # Строки областей
    for i, (oblast, stats) in enumerate(oblast_stats.items()):
        r = i + 4
        fill_hex = C_NAV_EVEN if i % 2 == 0 else C_WHITE
        fill = make_fill(fill_hex)

        ca = ws.cell(row=r, column=1, value=oblast)
        ca.font = make_font(C_BLUE_FONT, bold=True, size=10)
        ca.fill = fill
        ca.alignment = Alignment(horizontal="left", vertical="center")

        cb = ws.cell(row=r, column=2, value=stats["regions"])
        cb.font = make_font(C_DARK_FONT, bold=False, size=10)
        cb.fill = fill
        cb.alignment = Alignment(horizontal="center", vertical="center")

        cc = ws.cell(row=r, column=3, value=stats["rows"])
        cc.font = make_font(C_DARK_FONT, bold=False, size=10)
        cc.fill = fill
        cc.alignment = Alignment(horizontal="center", vertical="center")

        link_text = f"→ {oblast}"
        cd = ws.cell(row=r, column=4, value=link_text)
        cd.font = make_font(C_BLUE_FONT, bold=True, size=10)
        cd.fill = fill
        cd.hyperlink = f"#{oblast}!A1"
        cd.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[r].height = 18

    # Кнопка "назад" — добавим на каждый лист
    return ws


# ─── Загрузка данных ─────────────────────────────────────────────────────────

print("Загружаем данные...")
df_a = pd.read_csv(
    "ALL_KAZAKHSTAN_PATCHED.csv",
    encoding="utf-8-sig", low_memory=False
)
df_m = pd.read_csv(
    "MONTHLY_FINAL.csv",
    encoding="utf-8-sig", low_memory=False
)

# Приводим числовые колонки
for c in ANNUAL_COL_MAP.keys():
    if c in df_a.columns:
        df_a[c] = pd.to_numeric(df_a[c], errors="coerce")
for c in MONTHLY_COLS:
    if c in df_m.columns:
        df_m[c] = pd.to_numeric(df_m[c], errors="coerce")

# ─── Объединение по (oblast, region, id) ─────────────────────────────────────

print("Объединяем датасеты...")
merge_key = ["oblast", "region", "id"]

# Берём из annual: структурные поля + годовые данные
cols_from_annual = merge_key + ["text", "depth", "parent_text", "leaf"] + list(ANNUAL_COL_MAP.keys())
df_annual = df_a[cols_from_annual].copy()
df_annual["id"] = df_annual["id"].astype(str)

# Берём из monthly: только числовые колонки (те что есть)
m_available = [c for c in MONTHLY_COLS if c in df_m.columns]
cols_from_monthly = merge_key + m_available
df_monthly = df_m[cols_from_monthly].copy()
df_monthly["id"] = df_monthly["id"].astype(str)

# Мерж — LEFT JOIN: берём все строки из annual, добавляем monthly если есть
df = df_annual.merge(df_monthly, on=merge_key, how="left")

# Заполняем отсутствующие monthly колонки None
for c in MONTHLY_COLS:
    if c not in df.columns:
        df[c] = None

# Сортируем: по области, сохраняем оригинальный порядок строк внутри области
df = df.sort_values(["oblast"]).reset_index(drop=True)

print(f"Итого строк: {len(df)}, уникальных регионов: {df['region'].nunique()}")

# ─── Порядок областей ────────────────────────────────────────────────────────

OBLAST_ORDER = [
    "АЛМАТИНСКАЯ ОБЛАСТЬ", "АКМОЛИНСКАЯ ОБЛАСТЬ", "АКТЮБИНСКАЯ ОБЛАСТЬ",
    "Г.АСТАНА", "КАРАГАНДИНСКАЯ ОБЛАСТЬ", "Г.ШЫМКЕНТ", "Г.АЛМАТЫ",
    "ПАВЛОДАРСКАЯ ОБЛАСТЬ", "ТУРКЕСТАНСКАЯ ОБЛАСТЬ", "ОБЛАСТЬ ҰЛЫТАУ",
    "ВОСТОЧНО-КАЗАХСТАНСКАЯ ОБЛАСТЬ", "ЗАПАДНО-КАЗАХСТАНСКАЯ ОБЛАСТЬ",
    "СЕВЕРО-КАЗАХСТАНСКАЯ ОБЛАСТЬ", "ОБЛАСТЬ АБАЙ", "МАНГИСТАУСКАЯ ОБЛАСТЬ",
    "АТЫРАУСКАЯ ОБЛАСТЬ", "КОСТАНАЙСКАЯ ОБЛАСТЬ", "КЫЗЫЛОРДИНСКАЯ ОБЛАСТЬ",
    "ОБЛАСТЬ ЖЕТІСУ", "ЖАМБЫЛСКАЯ ОБЛАСТЬ",
]

# Добавим любые области которые есть в данных но не в списке
extra = [o for o in df["oblast"].unique() if o not in OBLAST_ORDER]
OBLAST_ORDER = OBLAST_ORDER + extra

# ─── Создаём книгу ───────────────────────────────────────────────────────────

print("Создаём Excel-файл...")
wb = Workbook()

# Сначала запишем листы (навигация будет первой)
oblast_stats = {}

for oblast in OBLAST_ORDER:
    df_obl = df[df["oblast"] == oblast]
    if df_obl.empty:
        continue

    n_regions = df_obl["region"].nunique()
    n_rows    = len(df_obl)

    oblast_stats[oblast] = {"regions": n_regions, "rows": n_rows}
    write_sheet(wb, oblast, df_obl)
    print(f"  ✓ {oblast}: {n_regions} районов, {n_rows} строк")

# Навигационный лист — создаём первым (перемещаем в начало)
nav_ws = write_nav_sheet(wb, oblast_stats)

# Перемещаем Навигацию в начало
wb.move_sheet("Навигация", offset=-len(wb.sheetnames) + 1)

# ─── Кнопка «← Навигация» на каждом листе ───────────────────────────────────

for ws in wb.worksheets:
    if ws.title == "Навигация":
        continue
    # Добавляем гиперссылку в ячейку выше заголовка (строка 0 нельзя, добавим в шапку H1/последнюю)
    last_col = ws.max_column
    nav_cell = ws.cell(row=1, column=last_col + 1)
    nav_cell.value = "← Навигация"
    nav_cell.hyperlink = "#Навигация!A1"
    nav_cell.font = Font(name="Arial", color=C_WHITE_FONT, bold=True, size=9)
    nav_cell.fill = make_fill("2E75B6")
    nav_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 14

# ─── Сохранение ─────────────────────────────────────────────────────────────

out_path = "KAZAKHSTAN_INVESTMENTS.xlsx"
wb.save(out_path)
print(f"\n✅ Файл сохранён: {out_path}")
print(f"   Листов: {len(wb.sheetnames)}")