"""
Конвертирует ALL_KAZAKHSTAN_FINAL.csv → KAZAKHSTAN_FINAL.xlsx
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "ALL_KAZAKHSTAN_PATCHED.csv"
OUTPUT_FILE = "KAZAKHSTAN_FINAL_PATCHED.xlsx"
INDENT      = "    "

YEAR_MAP = {
    "2019": "y122019",
    "2020": "y122020",
    "2021": "y122021",
    "2022": "y122022",
    "2023": "y122023",
    "2024": "y122024",
}
YEAR_COLS = list(YEAR_MAP.keys())

LEVEL_STYLE = {
    0: ("1F4E79", "FFFFFF", True),
    1: ("2E75B6", "FFFFFF", True),
    2: ("9DC3E6", "1F4E79", True),
    3: ("D6E4F0", "000000", False),
    4: ("EBF3FB", "000000", False),
    5: ("FFFFFF", "000000", False),
}

BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def format_val(val):
    s = str(val).strip()
    if s.lower() in ("x", "х"):
        return "x"
    try:
        f = float(val)
        if np.isnan(f):
            return None
        return int(round(f / 1000))
    except (ValueError, TypeError):
        return None

def safe_title(name: str) -> str:
    for ch in r"\/*?:[]":
        name = name.replace(ch, "")
    return name[:31]

def style_cell(cell, level: int, h_align="left"):
    bg, fg, bold = LEVEL_STYLE.get(level, ("FFFFFF", "000000", False))
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, color=fg, size=9)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal=h_align, vertical="center")

def write_oblast_sheet(ws, oblast_df):
    headers = ["Район", "Наименование (вид деятельности)"] + YEAR_COLS
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.border    = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"

    ri = 2
    for district in oblast_df["region"].unique():
        dist_df = oblast_df[oblast_df["region"] == district]
        for _, row in dist_df.iterrows():
            level = int(row["depth"])
            name  = INDENT * level + str(row["text"]).strip()

            c = ws.cell(row=ri, column=1, value=district)
            style_cell(c, level)

            c = ws.cell(row=ri, column=2, value=name)
            style_cell(c, level)

            for ci, year in enumerate(YEAR_COLS, 3):
                raw = row.get(YEAR_MAP[year], None)
                val = format_val(raw)
                c = ws.cell(row=ri, column=ci, value=val)
                style_cell(c, level, h_align="right")
                if isinstance(val, int):
                    c.number_format = "#,##0"

            ws.row_dimensions[ri].height = 14
            ri += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 65
    for ci in range(3, 3 + len(YEAR_COLS)):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    back = ws.cell(row=ri + 1, column=1, value="← Назад к навигации")
    back.hyperlink  = "#'Навигация'!A1"
    back.font       = Font(color="2E75B6", underline="single", bold=True, size=9)
    back.alignment  = Alignment(vertical="center")

    return ri - 2

def write_nav_sheet(ws, info_map):
    ws.title = "Навигация"
    ws.sheet_view.showGridLines = False

    t = ws.cell(row=1, column=1, value="Инвестиции в основной капитал — Казахстан")
    t.font = Font(bold=True, color="1F4E79", size=14)
    t.alignment = Alignment(vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 38

    sub = ws.cell(row=2, column=1, value="Нажмите на ссылку для перехода к данным области")
    sub.font = Font(italic=True, color="808080", size=10)
    ws.row_dimensions[2].height = 20

    for ci, h in enumerate(["Область / Город", "Районов", "Строк", "Перейти →"], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.border    = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    alt = ["EBF3FB", "FFFFFF"]
    for i, (oblast, info) in enumerate(info_map.items(), 4):
        bg   = alt[i % 2]
        fill = PatternFill("solid", fgColor=bg)

        c = ws.cell(row=i, column=1, value=oblast)
        c.fill = fill; c.font = Font(bold=True, color="1F4E79", size=10)
        c.border = BORDER; c.alignment = Alignment(vertical="center")

        c = ws.cell(row=i, column=2, value=info["districts"])
        c.fill = fill; c.font = Font(size=10); c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=i, column=3, value=info["rows"])
        c.fill = fill; c.font = Font(size=10); c.border = BORDER
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="center", vertical="center")

        sheet_title = info["sheet"]
        c = ws.cell(row=i, column=4, value=f"→ {sheet_title}")
        c.hyperlink = f"#{sheet_title}!A1"
        c.fill = fill
        c.font = Font(color="1F4E79", bold=True, underline="single", size=10)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 38

def main():
    print("📂 Читаю данные...")
    df = pd.read_csv(INPUT_FILE, encoding="utf-8-sig")
    print(f"   Строк: {len(df)}, Областей: {df['oblast'].nunique()}, Районов: {df['region'].nunique()}")

    wb = Workbook()
    wb.remove(wb.active)
    nav_ws = wb.create_sheet("Навигация")

    info_map = {}
    print("\n📊 Создаю листы...")

    for oblast in df["oblast"].unique():
        oblast_df   = df[df["oblast"] == oblast]
        sheet_title = safe_title(oblast)
        ws = wb.create_sheet(title=sheet_title)
        n_rows = write_oblast_sheet(ws, oblast_df)
        info_map[oblast] = {
            "sheet":     sheet_title,
            "districts": int(oblast_df["region"].nunique()),
            "rows":      n_rows,
        }
        print(f"   ✓ {oblast:<42} {info_map[oblast]['districts']:>3} р-нов  {n_rows:>6} строк")

    write_nav_sheet(nav_ws, info_map)

    print(f"\n💾 Сохраняю {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print(f"✅ Готово! Листов: {len(info_map)+1}, Строк: {len(df)}")

if __name__ == "__main__":
    main()