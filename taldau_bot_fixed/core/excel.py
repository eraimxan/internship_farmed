"""
core/excel.py — Построение Excel-файла из мастер-данных.

Создаёт многолистовой файл с навигацией, иерархическим
группированием строк и стилями по уровням глубины.
"""

import logging
import re
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import LEVEL_COLORS, OUTPUT_DIR
from core.periods import period_sort_key, parse_period_key, period_label, is_yearly_key

logger = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Цвета заголовков
HDR_NAV     = "1F3864"
HDR_KEY     = "1F4E79"
HDR_ANNUAL  = "2E75B6"
HDR_MONTHLY = "14375E"


# ─────────────────────────────────────────────
# ФОРМАТИРОВАНИЕ ЯЧЕЕК
# ─────────────────────────────────────────────

def _fmt_val(val) -> int | str | None:
    """Преобразует значение в тысячи тенге (int) или None/str."""
    s = str(val).strip()
    if s.lower() in ("x", "х"):
        return "x"
    if s.lower() in ("nan", "none", "nat", ""):
        return None
    try:
        f = float(s.replace(" ", "").replace(",", "."))
        return None if np.isnan(f) else int(round(f / 1000))
    except Exception:
        return None


def _apply_style(cell, depth: int, h_align: str = "left") -> None:
    bg, fg, bold, size = LEVEL_COLORS.get(min(depth, 6), ("FFFFFF", "000000", False, 9))
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=h_align, vertical="center")


def _apply_header(cell, bg: str) -> None:
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _safe_title(name: str) -> str:
    for ch in r"\/*?:[]'":
        name = name.replace(ch, "")
    return name[:31].strip()


# ─────────────────────────────────────────────
# СОРТИРОВКА СТРОК: DFS ПО ДЕРЕВУ
# ─────────────────────────────────────────────

def _add_parent_id(group: pd.DataFrame) -> pd.DataFrame:
    group = group.copy()
    dtid: dict = {}
    for _, row in group.iterrows():
        key = (str(row["text"]).strip(), int(row["depth"]))
        dtid[key] = row["id"]

    def get_pid(row):
        if pd.isna(row.get("parent_text")):
            return None
        return dtid.get((str(row["parent_text"]).strip(), int(row["depth"]) - 1))

    group["parent_id"] = group.apply(get_pid, axis=1)
    return group


def _dfs_sort(group: pd.DataFrame) -> pd.DataFrame:
    rows_by_id: dict = defaultdict(list)
    children: dict = defaultdict(list)

    for _, row in group.iterrows():
        nid = int(row["id"]) if pd.notna(row["id"]) else None
        pid = int(row["parent_id"]) if pd.notna(row.get("parent_id")) else None
        if nid is not None:
            rows_by_id[nid].append(row)
            children[pid].append(nid)

    all_ids = set(rows_by_id.keys())
    roots = sorted(
        nid for nid in all_ids
        if (
            int(rows_by_id[nid][0]["parent_id"])
            if pd.notna(rows_by_id[nid][0].get("parent_id"))
            else None
        ) not in all_ids
    )

    result: list = []
    visited: set = set()

    def dfs(nid: int) -> None:
        if nid in visited:
            return
        visited.add(nid)
        result.extend(rows_by_id[nid])
        for child in sorted(children.get(nid, [])):
            dfs(child)

    for root in roots:
        dfs(root)
    for nid in sorted(all_ids - visited):
        result.extend(rows_by_id[nid])

    return pd.DataFrame(result)


# ─────────────────────────────────────────────
# ПОСТРОЕНИЕ EXCEL
# ─────────────────────────────────────────────

def build_excel(df: pd.DataFrame, period_map: dict[str, str], output_path: str) -> dict:
    """
    Строит Excel-файл со всеми данными.

    Колонки делятся на:
      - Годовые (y12XXXX) — синий заголовок
      - Месячные — тёмно-синий заголовок

    Returns:
        info_map: dict {oblast_name: {"sheet": ..., "districts": ..., "rows": ...}}
    """
    data_cols = [c for c in period_map if c in df.columns]

    # Разделяем на годовые и месячные для раскраски заголовков
    yearly_cols = [c for c in data_cols if is_yearly_key(c)]
    monthly_cols = [c for c in data_cols if not is_yearly_key(c)]

    # Добавляем parent_id к каждой группе oblast/region
    parts = []
    for (ob, reg), grp in df.groupby(["oblast", "region"], sort=False):
        parts.append(_add_parent_id(grp))
    df = pd.concat(parts, ignore_index=True)
    df["depth"] = df["depth"].fillna(0).astype(int)
    df["id"] = pd.to_numeric(df["id"], errors="coerce").astype("Int64")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    nav_ws = wb.create_sheet("Навигация")
    info_map: dict = {}

    year_col_count = len(yearly_cols)

    for oblast in list(dict.fromkeys(df["oblast"].dropna().unique())):
        oblast_df = df[df["oblast"] == oblast]
        sheet_title = _safe_title(oblast)
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False

        headers = ["ID", "Parent ID", "Уровень", "Район", "Наименование (вид деятельности)"] + [
            period_map[c] for c in data_cols
        ]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            if ci <= 3:
                bg = HDR_NAV
            elif ci <= 5:
                bg = HDR_KEY
            elif ci <= 5 + year_col_count:
                bg = HDR_ANNUAL
            else:
                bg = HDR_MONTHLY
            _apply_header(cell, bg)

        ws.row_dimensions[1].height = 42
        ws.freeze_panes = "F2"

        excel_row, n_rows = 2, 0
        for (ob, reg), grp in oblast_df.groupby(["oblast", "region"], sort=False):
            try:
                sorted_grp = _dfs_sort(grp)
            except Exception:
                sorted_grp = grp

            for _, row in sorted_grp.iterrows():
                depth = min(int(row.get("depth", 0)), 6)
                nid = int(row["id"]) if pd.notna(row["id"]) else None
                pid = int(row["parent_id"]) if pd.notna(row.get("parent_id")) else None

                ws.cell(excel_row, 1, nid);   _apply_style(ws.cell(excel_row, 1), depth, "right")
                ws.cell(excel_row, 2, pid);   _apply_style(ws.cell(excel_row, 2), depth, "right")
                ws.cell(excel_row, 3, depth); _apply_style(ws.cell(excel_row, 3), depth, "center")
                ws.cell(excel_row, 4, str(reg)); _apply_style(ws.cell(excel_row, 4), depth, "left")
                ws.cell(excel_row, 5, "    " * depth + str(row.get("text", "")).strip())
                _apply_style(ws.cell(excel_row, 5), depth, "left")

                for ci, col in enumerate(data_cols, 6):
                    val = _fmt_val(row.get(col))
                    c = ws.cell(excel_row, ci)
                    _apply_style(c, depth, "right")
                    if val == "x":
                        _, fg, _, size = LEVEL_COLORS.get(min(depth, 6), ("FFFFFF", "000000", False, 9))
                        c.value = "x"
                        c.font = Font(name="Arial", color=fg, italic=True, size=size)
                    elif val is not None:
                        c.value = val
                        c.number_format = "#,##0"

                ws.row_dimensions[excel_row].height = 15
                if depth > 0:
                    ws.row_dimensions[excel_row].outline_level = min(depth, 7)
                    ws.row_dimensions[excel_row].hidden = False

                excel_row += 1
                n_rows += 1

        # Ширины колонок
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 7
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 60
        for i in range(len(data_cols)):
            ws.column_dimensions[get_column_letter(6 + i)].width = 16

        # Ссылка на навигацию
        back = ws.cell(excel_row + 1, 1, "← Навигация")
        back.hyperlink = "#'Навигация'!A1"
        back.font = Font(name="Arial", color="2E75B6", underline="single", bold=True, size=9)
        back.alignment = Alignment(vertical="center")

        info_map[oblast] = {
            "sheet": sheet_title,
            "districts": oblast_df["region"].nunique(),
            "rows": n_rows,
        }

    # ── Навигационный лист ──────────────────────────────────────
    nav_ws.title = "Навигация"
    nav_ws.sheet_view.showGridLines = False

    t = nav_ws.cell(1, 1, "Инвестиции в основной капитал — Казахстан")
    t.font = Font(name="Arial", bold=True, color="1F4E79", size=15)
    t.alignment = Alignment(vertical="center")
    nav_ws.merge_cells("A1:E1")
    nav_ws.row_dimensions[1].height = 40

    labels = list(period_map.values())
    sub = nav_ws.cell(
        2, 1,
        f"Данные: {labels[0] if labels else '—'}  →  {labels[-1] if labels else '—'}  |  тыс. тенге"
    )
    sub.font = Font(name="Arial", italic=True, color="808080", size=10)
    nav_ws.row_dimensions[2].height = 20

    for ci, h in enumerate(["Область / Город", "Районов", "Строк", "Перейти →"], 1):
        c = nav_ws.cell(3, ci, h)
        c.fill = PatternFill("solid", fgColor=HDR_NAV)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    nav_ws.row_dimensions[3].height = 26

    alt = ["EBF3FB", "FFFFFF"]
    for i, (oblast, info) in enumerate(info_map.items(), 4):
        fill = PatternFill("solid", fgColor=alt[i % 2])
        c = nav_ws.cell(i, 1, oblast)
        c.fill = fill
        c.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical="center")

        c = nav_ws.cell(i, 2, info["districts"])
        c.fill = fill; c.font = Font(name="Arial", size=10); c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = nav_ws.cell(i, 3, info["rows"])
        c.fill = fill; c.font = Font(name="Arial", size=10); c.border = THIN_BORDER
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = nav_ws.cell(i, 4, f"→ {info['sheet']}")
        c.hyperlink = f"#{info['sheet']}!A1"
        c.fill = fill
        c.font = Font(name="Arial", color="1F4E79", bold=True, underline="single", size=10)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        nav_ws.row_dimensions[i].height = 22

    nav_ws.column_dimensions["A"].width = 40
    nav_ws.column_dimensions["B"].width = 12
    nav_ws.column_dimensions["C"].width = 14
    nav_ws.column_dimensions["D"].width = 40

    wb.save(output_path)
    logger.info(f"Excel сохранён: {output_path}")
    return info_map


def generate_excel_filename(period_map: dict[str, str]) -> str:
    """Генерирует имя файла на основе диапазона периодов."""
    labels = list(period_map.values())
    if not labels:
        return "Kazakhstan_data.xlsx"
    first = labels[0].replace(" ", "_").replace("(", "").replace(")", "")
    last = labels[-1].replace(" ", "_").replace("(", "").replace(")", "")
    return f"Kazakhstan_{first}_{last}.xlsx"
