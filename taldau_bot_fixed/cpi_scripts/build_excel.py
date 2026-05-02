"""
cpi_scripts/build_excel.py — Build Excel file from CPI data CSV.

Format mirrors core/excel.py:
  • Navigation sheet with hyperlinks to each oblast
  • One sheet per oblast
  • Hierarchical row grouping (outline levels)
  • Color-coded depth levels
  • Two column groups: "К соотв. периоду прошлого года" and "К предыдущему году"

Usage:
    python build_excel.py [--csv output/cpi_data.csv] [--out output/CPI_Kazakhstan.xlsx]
"""

import sys
import re
import json
import argparse
import logging
from pathlib import Path
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from cpi_config import (
    DATA_CSV, EXCEL_OUT, LEVEL_COLORS,
    TARGET_PERIODS, PERIOD_LABELS,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────────
THIN = Border(
    left=Side(style="thin",   color="D0D0D0"),
    right=Side(style="thin",  color="D0D0D0"),
    top=Side(style="thin",    color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
HDR_NAV       = "1F3864"
HDR_KEY       = "1F4E79"
HDR_YOY       = "2E75B6"   # year-on-year group header
HDR_PREV      = "14375E"   # prev-year group header
HDR_PERIOD_YOY  = "4472C4"  # individual period columns for yoy
HDR_PERIOD_PREV = "1F3864"  # individual period columns for prev_year


def _safe_title(name: str) -> str:
    for ch in r"\/*?:[]'":
        name = name.replace(ch, "")
    return name[:31].strip()


def _hdr(cell, bg: str, wrap: bool = True) -> None:
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.border = THIN
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _style(cell, depth: int, h_align: str = "left") -> None:
    bg, fg, bold, size = LEVEL_COLORS.get(min(depth, 6), ("FFFFFF", "000000", False, 9))
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.border = THIN
    cell.alignment = Alignment(horizontal=h_align, vertical="center")


# ─────────────────────────────────────────────
# PERIOD KEY HELPERS
# ─────────────────────────────────────────────

def _period_sort_key(key: str) -> tuple:
    """
    Sort period keys chronologically.
    Keys look like: y12025 (Q1 2025), y22024 (Q2 2024), y102023 (Oct 2023), etc.
    """
    m = re.match(r"^y(\d+)$", key)
    if not m:
        return (0, 0)
    num = int(m.group(1))
    # Determine if quarterly (1-4 digit quarter) or monthly (6-digit yMMYYYY)
    s = m.group(1)
    if len(s) <= 2:
        # Format: y{Q}{YYYY} e.g. y12025 → Q=1, Y=2025... but y12025 is ambiguous
        # Try: last 4 digits = year, first digit(s) = quarter/month
        year = int(s[-4:]) if len(s) >= 4 else 0
        period = int(s[:-4]) if len(s) > 4 else int(s)
        return (year, period)
    elif len(s) == 6:
        # Monthly: MMYYYY
        month = int(s[:2])
        year  = int(s[2:])
        return (year, month)
    else:
        # Fallback: numeric sort
        return (num, 0)


def _period_label(key: str, period_name_map: dict) -> str:
    """Human-readable period label from key and optional name map."""
    if key in period_name_map:
        return period_name_map[key]
    m = re.match(r"^y(\d+)$", key)
    if not m:
        return key
    s = m.group(1)
    if len(s) == 6:
        month, year = int(s[:2]), int(s[2:])
        q = (month - 1) // 3 + 1
        return f"Q{q} {year}"
    elif len(s) == 5:
        q, year = int(s[0]), int(s[1:])
        return f"Q{q} {year}"
    elif len(s) <= 2:
        # Could be quarter + truncated year — show as-is
        return key
    return key


# ─────────────────────────────────────────────
# PARENT ID LINKING (for outline grouping)
# ─────────────────────────────────────────────

def _build_parent_ids(group: pd.DataFrame) -> pd.DataFrame:
    group = group.copy()
    text_to_id: dict = {}
    for _, row in group.iterrows():
        k = (str(row.get("text", "")).strip(), int(row.get("depth", 0)))
        text_to_id[k] = row.get("id")

    def _pid(row):
        pt = row.get("parent_text")
        if pd.isna(pt) or str(pt).strip() == "":
            return None
        d = int(row.get("depth", 0)) - 1
        return text_to_id.get((str(pt).strip(), d))

    group["_parent_id"] = group.apply(_pid, axis=1)
    return group


def _dfs_sort(group: pd.DataFrame) -> pd.DataFrame:
    rows_by_id  = defaultdict(list)
    children    = defaultdict(list)

    for _, row in group.iterrows():
        nid = row.get("id")
        pid = row.get("_parent_id")
        if pd.notna(nid):
            nid = int(float(nid))
            rows_by_id[nid].append(row)
            pid_int = int(float(pid)) if pd.notna(pid) else None
            children[pid_int].append(nid)

    all_ids = set(rows_by_id.keys())
    roots = sorted(
        nid for nid in all_ids
        if (
            int(float(rows_by_id[nid][0].get("_parent_id")))
            if pd.notna(rows_by_id[nid][0].get("_parent_id"))
            else None
        ) not in all_ids
    )

    result: list = []
    visited: set = set()

    def dfs(nid):
        if nid in visited:
            return
        visited.add(nid)
        result.extend(rows_by_id[nid])
        for child in sorted(children.get(nid, [])):
            dfs(child)

    for root in roots:
        dfs(root)
    for nid in sorted(all_ids - visited):
        result.extend(rows_by_id[nid])

    return pd.DataFrame(result) if result else group


# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────

def build_excel(df: pd.DataFrame, params_path: Path | None, out_path: Path) -> None:
    """Build multi-sheet Excel workbook from CPI DataFrame."""

    # Load period name map if available
    period_name_map: dict[str, str] = {}
    if params_path and params_path.exists():
        try:
            p = json.loads(params_path.read_text(encoding="utf-8"))
            dl = p.get("date_list", [])
            nl = p.get("period_name_list", [])
            for k, n in zip(dl, nl):
                period_name_map[f"y{k}"] = n
        except Exception:
            pass

    # Identify period columns
    all_period_cols = sorted(
        [c for c in df.columns if re.match(r"^y\d+$", c)],
        key=_period_sort_key,
    )
    if not all_period_cols:
        log.error("No period columns found in DataFrame!")
        return

    # Split DataFrame by comparison type
    comp_keys = [k for k in TARGET_PERIODS if k in df["comparison_type"].unique()]
    if not comp_keys:
        comp_keys = df["comparison_type"].dropna().unique().tolist()

    # Period cols that actually have data (by comparison type)
    def live_cols(comp_key):
        sub = df[df["comparison_type"] == comp_key]
        return [
            c for c in all_period_cols
            if c in sub.columns and sub[c].notna().any()
        ]

    cols_by_type = {k: live_cols(k) for k in comp_keys}

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    nav_ws = wb.create_sheet("Навигация")
    info_map: dict = {}

    # Support both "oblast" and "region" columns (CPI data has oblast=region)
    oblast_col = "oblast" if "oblast" in df.columns else "region"
    oblasts = list(dict.fromkeys(df[oblast_col].dropna().tolist()))

    for oblast in oblasts:
        ob_df = df[df[oblast_col] == oblast].copy()
        sheet_title = _safe_title(oblast)
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False

        # ── Build column layout ────────────────────────────────────
        # Columns: [ID | Parent ID | Level | District | Category] + [yoy cols...] + [prev_year cols...]
        fixed_cols  = ["ID", "Parent ID", "Уровень", "Район", "Наименование (товар/услуга)"]
        n_fixed     = len(fixed_cols)

        col_headers: list[tuple[str, str]] = []  # (label, bg_color)
        col_data_map: list[tuple[str, str]] = []  # (comparison_type_key, period_col)

        for ck in comp_keys:
            for pc in cols_by_type.get(ck, []):
                lbl = _period_label(pc, period_name_map)
                bg  = HDR_PERIOD_YOY if ck == "yoy" else HDR_PERIOD_PREV
                col_headers.append((lbl, bg))
                col_data_map.append((ck, pc))

        total_cols = n_fixed + len(col_headers)

        # ── Row 1: merged group headers ────────────────────────────
        col_cur = n_fixed + 1
        for ck in comp_keys:
            n = len(cols_by_type.get(ck, []))
            if n == 0:
                continue
            grp_bg = HDR_YOY if ck == "yoy" else HDR_PREV
            grp_lbl = PERIOD_LABELS.get(ck, ck)
            start_col = col_cur
            end_col   = col_cur + n - 1
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1,   end_column=end_col,
            )
            c = ws.cell(1, start_col, grp_lbl)
            _hdr(c, grp_bg)
            col_cur += n

        # ── Row 2: fixed + individual period headers ───────────────
        for ci, h in enumerate(fixed_cols, 1):
            bg = HDR_NAV if ci <= 3 else HDR_KEY
            c = ws.cell(2, ci, h)
            _hdr(c, bg)
        for ci, (lbl, bg) in enumerate(col_headers, n_fixed + 1):
            c = ws.cell(2, ci, lbl)
            _hdr(c, bg)

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 42
        ws.freeze_panes = "F3"

        # ── Data rows ─────────────────────────────────────────────
        # Pivot: for each (region, category), put both comparison types on one row
        # by building a dict keyed on (region, depth, text, id)
        key_cols = ["region", "depth", "parent_text", "text", "id"]
        pivot: dict[tuple, dict] = {}

        for ck in comp_keys:
            sub = ob_df[ob_df["comparison_type"] == ck]
            for _, row in sub.iterrows():
                k = (
                    row.get("region", ""),
                    int(float(row.get("depth", 0))),
                    str(row.get("parent_text", "")),
                    str(row.get("text", "")),
                    row.get("id"),
                )
                if k not in pivot:
                    pivot[k] = {
                        "region":      row.get("region", ""),
                        "depth":       int(float(row.get("depth", 0))),
                        "parent_text": str(row.get("parent_text", "")),
                        "text":        str(row.get("text", "")),
                        "id":          row.get("id"),
                    }
                for pc in cols_by_type.get(ck, []):
                    cell_key = f"{ck}__{pc}"
                    pivot[k][cell_key] = row.get(pc)

        # Sort within each (region, depth, text) group
        rows_list = list(pivot.values())

        # Group by region then sort by depth within region
        region_order = list(dict.fromkeys(ob_df["region"].tolist()))
        rows_by_region: dict[str, list] = defaultdict(list)
        for r in rows_list:
            rows_by_region[r["region"]].append(r)

        excel_row = 3
        n_data_rows = 0

        for reg in region_order:
            reg_rows = rows_by_region.get(reg, [])
            if not reg_rows:
                continue
            # DFS order: each parent immediately followed by its children (like the website tree)
            reg_df = pd.DataFrame(reg_rows)
            try:
                reg_df = _build_parent_ids(reg_df)
                reg_df = _dfs_sort(reg_df)
            except Exception:
                pass

            for _, r in reg_df.iterrows():
                depth = min(int(r.get("depth", 0)), 6)
                nid   = r.get("id")
                nid_v = int(float(nid)) if pd.notna(nid) and str(nid).strip() not in ("", "nan") else None
                pid   = r.get("_parent_id")
                pid_v = int(float(pid)) if pd.notna(pid) and str(pid).strip() not in ("", "nan") else None

                ws.cell(excel_row, 1, nid_v);           _style(ws.cell(excel_row, 1), depth, "right")
                ws.cell(excel_row, 2, pid_v);           _style(ws.cell(excel_row, 2), depth, "right")
                ws.cell(excel_row, 3, depth);           _style(ws.cell(excel_row, 3), depth, "center")
                ws.cell(excel_row, 4, reg);             _style(ws.cell(excel_row, 4), depth, "left")
                ws.cell(excel_row, 5, "    " * depth + str(r["text"]))
                _style(ws.cell(excel_row, 5), depth, "left")

                for ci, (ck, pc) in enumerate(col_data_map, n_fixed + 1):
                    val = r.get(f"{ck}__{pc}")
                    c = ws.cell(excel_row, ci)
                    _style(c, depth, "right")
                    if val == "x":
                        _, fg, _, size = LEVEL_COLORS.get(min(depth, 6), ("FFFFFF", "000000", False, 9))
                        c.value = "x"
                        c.font = Font(name="Arial", color=fg, italic=True, size=size)
                    elif val is not None:
                        try:
                            fv = float(val)
                            c.value = round(fv, 2)
                            c.number_format = "#,##0.00"
                        except Exception:
                            c.value = val

                ws.row_dimensions[excel_row].height = 15
                if depth > 0:
                    ws.row_dimensions[excel_row].outline_level = min(depth, 7)
                    ws.row_dimensions[excel_row].hidden = False

                excel_row  += 1
                n_data_rows += 1

        # ── Column widths ──────────────────────────────────────────
        ws.column_dimensions["A"].width = 10   # ID
        ws.column_dimensions["B"].width = 10   # Parent ID
        ws.column_dimensions["C"].width = 7    # Level
        ws.column_dimensions["D"].width = 28   # District
        ws.column_dimensions["E"].width = 54   # Category
        for i in range(len(col_headers)):
            ws.column_dimensions[get_column_letter(n_fixed + 1 + i)].width = 13

        # ── Back link ─────────────────────────────────────────────
        back = ws.cell(excel_row + 1, 1, "← Навигация")
        back.hyperlink = "#'Навигация'!A1"
        back.font = Font(name="Arial", color="2E75B6", underline="single", bold=True, size=9)
        back.alignment = Alignment(vertical="center")

        info_map[oblast] = {
            "sheet":    sheet_title,
            "regions":  ob_df["region"].nunique(),
            "rows":     n_data_rows,
        }
        log.info("Sheet '%s': %d data rows", sheet_title, n_data_rows)

    # ── Navigation sheet ──────────────────────────────────────────
    nav_ws.title = "Навигация"
    nav_ws.sheet_view.showGridLines = False

    t = nav_ws.cell(1, 1, "Индексы потребительских цен — Казахстан")
    t.font = Font(name="Arial", bold=True, color="1F4E79", size=15)
    t.alignment = Alignment(vertical="center")
    nav_ws.merge_cells("A1:E1")
    nav_ws.row_dimensions[1].height = 40

    comp_labels = " | ".join(PERIOD_LABELS.get(k, k) for k in comp_keys)
    sub = nav_ws.cell(2, 1, f"Периоды: {comp_labels}  |  единица: %")
    sub.font = Font(name="Arial", italic=True, color="808080", size=10)
    nav_ws.row_dimensions[2].height = 20

    for ci, h in enumerate(["Область / Город", "Районов", "Строк", "Перейти →"], 1):
        c = nav_ws.cell(3, ci, h)
        c.fill = PatternFill("solid", fgColor=HDR_NAV)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")
    nav_ws.row_dimensions[3].height = 26

    alt = ["EBF3FB", "FFFFFF"]
    for i, (oblast, info) in enumerate(info_map.items(), 4):
        fill = PatternFill("solid", fgColor=alt[i % 2])

        c = nav_ws.cell(i, 1, oblast)
        c.fill = fill
        c.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
        c.border = THIN
        c.alignment = Alignment(vertical="center")

        c = nav_ws.cell(i, 2, info["regions"])
        c.fill = fill; c.font = Font(name="Arial", size=10); c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = nav_ws.cell(i, 3, info["rows"])
        c.fill = fill; c.font = Font(name="Arial", size=10); c.border = THIN
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = nav_ws.cell(i, 4, f"→ {info['sheet']}")
        c.hyperlink = f"#{info['sheet']}!A1"
        c.fill = fill
        c.font = Font(name="Arial", color="1F4E79", bold=True, underline="single", size=10)
        c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")
        nav_ws.row_dimensions[i].height = 22

    nav_ws.column_dimensions["A"].width = 40
    nav_ws.column_dimensions["B"].width = 12
    nav_ws.column_dimensions["C"].width = 14
    nav_ws.column_dimensions["D"].width = 40

    wb.save(str(out_path))
    log.info("Excel saved: %s", out_path)


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build CPI Excel from CSV")
    parser.add_argument("--csv",    default=str(DATA_CSV),   help="Input CSV file")
    parser.add_argument("--out",    default=str(EXCEL_OUT),  help="Output Excel file")
    parser.add_argument("--params", default="",              help="cpi_params.json path")
    args = parser.parse_args()

    csv_path = Path(args.csv)
    out_path = Path(args.out)

    if not csv_path.exists():
        log.error("CSV not found: %s — run fetch.py first", csv_path)
        sys.exit(1)

    df = pd.read_csv(csv_path, encoding="utf-8-sig", low_memory=False)
    log.info("Loaded %d rows from %s", len(df), csv_path)

    params_path = Path(args.params) if args.params else csv_path.parent / "cpi_params.json"

    build_excel(df, params_path if params_path.exists() else None, out_path)


if __name__ == "__main__":
    main()
