import json
import sys
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "pharma_data.json"
OUTPUT_FILE = "pharma_register.xlsx"

COLUMNS = [
    "№",
    "Торговое наименование",
    "МНН (международное непатентованное наименование)",
    "Лекарственная форма",
    "Форма выпуска",
    "Вид лекарственного препарата",
    "Производитель",
    "Страна производителя",
    "Адрес производственной площадки",
    "Номер регистрационного удостоверения",
    "Дата регистрации",
    "Срок действия",
    "Статус регистрации",
    "Страна регистрации",
    "АТХ код / группа",
    "Рецептурный отпуск",
]

DRUG_KIND_CODES = {
    "01": "Оригинальный",
    "02": "Воспроизведённый",
    "03": "Биоаналог",
    "04": "Гибридный",
    "05": "Хорошо изученный",
    "06": "Комбинированный",
    "07": "Иной",
}

UNIT_CODES = {
    "H87": "шт.", "796": "шт.", "PCS": "шт.",
    "CMQ": "см3", "MLT": "мл", "111": "мл",
    "GRM": "г", "KGM": "кг",
    "MGM": "мг", "161": "мг",
    "MTQ": "м3", "MMQ": "мм3", "113": "мкл",
    "163": "л", "E27": "МЕ", "A99": "МЕ",
}

SECONDARY_PKG_CODES = {
    "151": "пачка картонная",
    "149": "пачка",
    "072": "контурная ячейковая упаковка",
    "222": "пакет",
    "221": "пакетик",
    "070": "флакон",
    "155": "тюбик-капельница",
    "109": "шприц",
    "130": "ампула",
    "128": "флакон с капельницей",
    "125": "банка",
    "132": "контейнер",
    "049": "туба",
    "179": "пенал",
    "192": "мешок",
    "172": "ланцет",
    "163": "пакет фольгированный",
    "178": "саше",
}


def format_release_form(pkg_list: list) -> str:
    lines = []
    for pkg in pkg_list:
        pd_list = pkg.get("packageDetails") or []
        if not pd_list:
            continue
        pd = pd_list[0]

        form = pd.get("dosageFormName", "")

        # Volume / quantity with unit
        measures = pd.get("packageMeasure") or []
        vol_str = ""
        if measures:
            m = measures[0]
            val = m.get("value", "")
            unit = UNIT_CODES.get(m.get("measurementUnitCode", ""), m.get("measurementUnitCode", ""))
            if val:
                vol_str = f"по {val} {unit}".strip()

        # Primary packaging — shorten packageMaterialText to first clause
        mat_text = pd.get("packageMaterialText", "")
        if mat_text:
            primary = mat_text.split(",")[0].strip().lower()
        else:
            primary = ""

        qty = pd.get("componentPackageQuantity", "")

        # Secondary packaging
        sec_code = pkg.get("drugSecondaryPackageKindCode", "")
        secondary = SECONDARY_PKG_CODES.get(sec_code, f"упаковка {sec_code}" if sec_code else "")

        parts = []
        if form:
            parts.append(form)
        if vol_str:
            parts.append(vol_str)
        if primary:
            parts.append(f'в первичной упаковке «{primary}»')
        if qty:
            parts.append(f"по {qty} шт.")
        if secondary:
            parts.append(f'во вторичной упаковке «{secondary}»')

        lines.append(" ".join(parts))
    # Remove duplicate lines while preserving order
    seen = dict.fromkeys(lines)
    return "\n".join(seen)


def parse_date(val) -> str:
    if not val:
        return ""
    if isinstance(val, dict):
        val = val.get("$date", "")
    if not val:
        return ""
    try:
        dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return str(val)


def format_address(addr_list: list) -> str:
    if not addr_list:
        return ""
    addr = addr_list[0]
    parts = [
        addr.get("unifiedCountryCode", {}).get("value", ""),
        addr.get("postCode", ""),
        addr.get("regionName", ""),
        addr.get("cityName", ""),
        addr.get("streetName", ""),
        addr.get("buildingNumberId", ""),
    ]
    return ", ".join(p for p in parts if p)


def extract_row(rec: dict, index: int) -> list:
    dd = rec.get("drugDetails", {}) or {}
    dcd_list = rec.get("drugCountryRegistrationDetails", []) or []
    dcd = dcd_list[0] if dcd_list else {}

    # Trade name
    trade_name = dcd.get("drugTradeName", "")

    # INN — active substance from drugNameDetails
    drug_names = dd.get("drugNameDetails", []) or []
    inn = "; ".join(d.get("drugName", "") for d in drug_names if d.get("drugName"))

    # Dosage form
    dosage_form = (dd.get("dosageFormDetails") or {}).get("dosageFormName", "")

    # Manufacturer
    mah_list = dd.get("manufacturingAuthorizationHolderDetails", []) or []
    if mah_list:
        mah = mah_list[0]
        manufacturer = mah.get("businessEntityName", "")
        mfr_country = (mah.get("unifiedCountryCode") or {}).get("value", "")
        mfr_address = format_address(mah.get("subjectAddressDetails", []))
    else:
        manufacturer = ""
        mfr_country = ""
        mfr_address = ""

    # Registration certificate
    cert = dcd.get("drugRegistrationCertificateDetails") or {}
    cert_id = cert.get("registrationCertificateId", "")
    reg_date = parse_date(cert.get("docCreationDate"))
    validity_date = parse_date(cert.get("docValidityDate") or cert.get("docExtensionDate"))
    reg_status = cert.get("registrationStatusName", "")

    # Country of registration
    reg_country = (dcd.get("unifiedCountryCode") or {}).get("value", "")

    # Release form (форма выпуска)
    pkg_list = dd.get("packageFormDetails") or []
    release_form = format_release_form(pkg_list)

    # Drug kind (вид лекарственного препарата)
    kind_code = dd.get("registrationDrugKindCode", "")
    kind_name = dd.get("registrationDrugKindName", "") or DRUG_KIND_CODES.get(kind_code, kind_code)

    # ATC
    atc = dd.get("atcName", "")

    # Prescription (descriptionText often holds "по рецепту" / "без рецепта")
    prescription = dd.get("descriptionText", "")

    return [
        index,
        trade_name,
        inn,
        dosage_form,
        release_form,
        kind_name,
        manufacturer,
        mfr_country,
        mfr_address,
        cert_id,
        reg_date,
        validity_date,
        reg_status,
        reg_country,
        atc,
        prescription,
    ]


def style_header(ws, num_cols: int):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border


def style_data_rows(ws, num_rows: int, num_cols: int):
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    thin = Side(style="thin", color="D0D7E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(size=9, name="Calibri")

    for row in range(2, num_rows + 2):
        fill = alt_fill if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill


def set_column_widths(ws):
    widths = [5, 30, 35, 25, 55, 22, 35, 12, 40, 30, 14, 14, 20, 12, 30, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    print(f"Loading {INPUT_FILE}...", flush=True)
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"Loaded {len(data)} records.", flush=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр лекарственных средств"

    # Header row
    ws.row_dimensions[1].height = 40
    for col, name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows
    print("Writing rows...", flush=True)
    for i, rec in enumerate(data, start=1):
        row_data = extract_row(rec, i)
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=i + 1, column=col, value=val)
        if i % 1000 == 0:
            print(f"  {i}/{len(data)}...", flush=True)

    # Styles
    print("Applying styles...", flush=True)
    style_header(ws, len(COLUMNS))
    style_data_rows(ws, len(data), len(COLUMNS))
    set_column_widths(ws)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    print(f"Saving {OUTPUT_FILE}...", flush=True)
    wb.save(OUTPUT_FILE)
    print(f"Done! Saved to {OUTPUT_FILE}", flush=True)


if __name__ == "__main__":
    main()
