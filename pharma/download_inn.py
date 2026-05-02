"""
Download "Справочник международных непатентованных наименований лекарственных средств"
(Dictionary 1016) from nsi.eaeunion.org and export to Excel.

API docs: https://nsi.eaeunion.org/portal/info/api
Endpoint: GET https://nsi.eaeunion.org/api/v1/dictionaries/1016/elements
"""

import requests
import time
import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DICTIONARY_CODE = "1016"
DATE_ACTUAL = "2026-04-24"
BASE_URL = f"https://nsi.eaeunion.org/api/v1/dictionaries/{DICTIONARY_CODE}/elements"
BATCH_SIZE = 1500  # API max per request
OUTPUT_FILE = "inn_register.xlsx"

COLUMNS = [
    ("№", 5),
    ("Код МНН", 10),
    ("МНН (русское наименование)", 40),
    ("МНН (английское наименование)", 40),
    ("МНН (латинское наименование)", 40),
    ("Дата начала действия", 20),
    ("Дата окончания действия", 20),
]


def fetch_batch(offset: int, retries: int = 3) -> list:
    params = {
        "dateActual": DATE_ACTUAL,
        "offset": offset,
        "limit": BATCH_SIZE,
        "code": "Code",
        "sortOrder": "asc",
    }
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(BASE_URL, params=params, timeout=60)
            resp.raise_for_status()
            return resp.json().get("elements", [])
        except requests.exceptions.Timeout:
            print(f"  Timeout (attempt {attempt}/{retries}), retrying...", flush=True)
            time.sleep(5 * attempt)
        except requests.exceptions.RequestException as e:
            print(f"  Error (attempt {attempt}/{retries}): {e}", flush=True)
            time.sleep(5 * attempt)
    return []


def download_all() -> list:
    all_elements = []
    offset = 0
    batch_num = 0

    print(f"Downloading INN dictionary (code {DICTIONARY_CODE}), date={DATE_ACTUAL}...")
    print(f"Batch size: {BATCH_SIZE}")
    print("-" * 60)

    while True:
        batch_num += 1
        print(f"Batch {batch_num}: offset {offset}...", end=" ", flush=True)
        batch = fetch_batch(offset)
        if not batch:
            print("Empty — done." if batch_num > 1 else "Empty response — check API.")
            break
        all_elements.extend(batch)
        print(f"got {len(batch)}, total so far: {len(all_elements)}", flush=True)
        if len(batch) < BATCH_SIZE:
            print("Last batch received.")
            break
        offset += BATCH_SIZE
        time.sleep(0.3)

    print("-" * 60)
    print(f"Total records downloaded: {len(all_elements)}")
    return all_elements


def parse_date(dt_str: str) -> str:
    if not dt_str:
        return ""
    try:
        # "2018-06-13T00:00:00.000Z" → "13.06.2018"
        return dt_str[:10].split("-")[::-1][0] + "." + dt_str[5:7] + "." + dt_str[:4]
    except Exception:
        return dt_str[:10] if len(dt_str) >= 10 else dt_str


def build_excel(elements: list):
    print(f"\nBuilding Excel workbook with {len(elements)} rows...")

    wb = Workbook()
    ws = wb.active
    ws.title = "МНН лекарственных средств"

    # ── Header ──────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    thin_white = Side(style="thin", color="FFFFFF")
    header_border = Border(
        left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
    )

    ws.row_dimensions[1].height = 36
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Data rows ────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    thin_gray = Side(style="thin", color="D0D7E0")
    data_border = Border(
        left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
    )
    data_font = Font(size=9, name="Calibri")

    for row_idx, elem in enumerate(elements, start=1):
        excel_row = row_idx + 1
        data = elem.get("data", {})

        date_from = parse_date(elem.get("dateTimeFrom", ""))
        date_to_raw = elem.get("dateTimeTo", "")
        # "2100-01-01" means "indefinite" — show as blank
        date_to = "" if date_to_raw.startswith("2100") else parse_date(date_to_raw)

        row_values = [
            row_idx,
            data.get("Code", ""),
            data.get("Value_RUS", ""),
            data.get("Value_ENG", ""),
            data.get("Value_LAT", ""),
            date_from,
            date_to,
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = data_border
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if fill:
                cell.fill = fill

        if row_idx % 2000 == 0:
            print(f"  Written {row_idx}/{len(elements)}...", flush=True)

    # ── Metadata sheet ───────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("О справочнике")
    meta_rows = [
        ("Справочник", "Международные непатентованные наименования лекарственных средств"),
        ("Код справочника (НСИ ЕАЭС)", DICTIONARY_CODE),
        ("Дата актуальности данных", DATE_ACTUAL),
        ("Дата выгрузки", str(date.today())),
        ("Источник", f"https://nsi.eaeunion.org/portal/{DICTIONARY_CODE}"),
        ("API endpoint", BASE_URL),
        ("Всего записей", len(elements)),
    ]
    bold = Font(bold=True, size=10, name="Calibri")
    normal = Font(size=10, name="Calibri")
    for r, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(row=r, column=1, value=k).font = bold
        ws_meta.cell(row=r, column=2, value=v).font = normal
    ws_meta.column_dimensions["A"].width = 35
    ws_meta.column_dimensions["B"].width = 60

    print(f"Saving to {OUTPUT_FILE}...", flush=True)
    wb.save(OUTPUT_FILE)
    print(f"Done! Saved: {OUTPUT_FILE}")


def main():
    elements = download_all()
    if not elements:
        print("No data downloaded — exiting.", file=sys.stderr)
        sys.exit(1)
    build_excel(elements)


if __name__ == "__main__":
    main()
