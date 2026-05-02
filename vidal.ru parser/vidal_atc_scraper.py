import requests
import time
import json
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_URL    = "https://www.vidal.ru"
OUTPUT      = "vidal_atc.xlsx"
CHECKPOINT  = "vidal_atc_checkpoint.json"
DELAY       = 2.0    # seconds between requests
MAX_RETRIES = 5
RETRY_WAIT  = 10     # seconds to wait on timeout before retry

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
}

LEVEL_STYLES = {
    1: {"fill": "1F3864", "font_color": "FFFFFF", "bold": True,  "size": 12, "indent": 0},
    2: {"fill": "4472C4", "font_color": "FFFFFF", "bold": True,  "size": 11, "indent": 1},
    3: {"fill": "D9E1F2", "font_color": "000000", "bold": True,  "size": 10, "indent": 2},
    4: {"fill": "EBF0FA", "font_color": "000000", "bold": False, "size": 10, "indent": 3},
    5: {"fill": "FFFFFF", "font_color": "000000", "bold": False, "size": 10, "indent": 4},
}


def atc_level(code: str) -> int:
    return {1: 1, 3: 2, 4: 3, 5: 4, 7: 5}.get(len(code), 0)


def fetch_with_retry(url: str) -> requests.Response:
    """Fetch URL with retry on connection errors / timeouts."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            time.sleep(DELAY)
            r = requests.get(url, headers=HEADERS, timeout=45)
            r.encoding = "utf-8"
            r.raise_for_status()
            return r
        except (requests.exceptions.ConnectTimeout,
                requests.exceptions.ReadTimeout,
                requests.exceptions.ConnectionError) as e:
            if attempt < MAX_RETRIES:
                print(f"\n    [timeout/error, retry {attempt}/{MAX_RETRIES} in {RETRY_WAIT}s] {e}")
                time.sleep(RETRY_WAIT)
            else:
                raise


def parse_letter_page(html: str, letter: str) -> list[dict]:
    """Parse a letter page and return list of node dicts."""
    soup = BeautifulSoup(html, "html.parser")
    rows = soup.find_all("tr", id=True)

    nodes = []
    letter_upper = letter.upper()

    # Level-1 node: the letter itself
    nodes.append({"code": letter_upper, "name": letter_upper, "level": 1, "parent": None})

    for row in rows:
        code = row["id"].upper()
        level = atc_level(code)
        if level == 0:
            continue

        tds = row.find_all("td")
        if len(tds) < 2:
            continue
        name = tds[1].get_text(separator=" ", strip=True)

        parent_len = {2: 1, 3: 3, 4: 4, 5: 5}.get(level)
        parent = code[:parent_len] if parent_len else None

        nodes.append({"code": code, "name": name, "level": level, "parent": parent})

    return nodes


def load_checkpoint() -> dict:
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"done_letters": [], "nodes": []}


def save_checkpoint(done_letters: list, nodes: list):
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump({"done_letters": done_letters, "nodes": nodes}, f, ensure_ascii=False)


def scrape_all() -> list[dict]:
    # Get letter list
    print("Fetching letter index …")
    r = fetch_with_retry(f"{BASE_URL}/drugs/atc")
    soup = BeautifulSoup(r.text, "html.parser")
    letters = [
        a["href"].split("/")[-1].upper()
        for a in soup.find_all("a", href=lambda h: h and "/drugs/atcl/" in h)
    ]
    print(f"  Found {len(letters)} letters: {' '.join(letters)}\n")

    # Load checkpoint if exists
    checkpoint = load_checkpoint()
    done_letters = checkpoint["done_letters"]
    all_nodes = checkpoint["nodes"]

    if done_letters:
        print(f"  Resuming from checkpoint — already done: {' '.join(done_letters)}\n")

    for letter in letters:
        if letter in done_letters:
            continue

        url = f"{BASE_URL}/drugs/atcl/{letter.lower()}"
        print(f"  [{letters.index(letter)+1}/{len(letters)}] {url} … ", end="", flush=True)

        r = fetch_with_retry(url)
        nodes = parse_letter_page(r.text, letter)
        print(f"{len(nodes)} nodes")

        all_nodes.extend(nodes)
        done_letters.append(letter)
        save_checkpoint(done_letters, all_nodes)

    # Remove checkpoint after successful completion
    if os.path.exists(CHECKPOINT):
        os.remove(CHECKPOINT)

    print(f"\nTotal nodes collected: {len(all_nodes)}")
    return all_nodes


def build_path_map(nodes: list[dict]) -> dict[str, str]:
    by_code = {n["code"]: n for n in nodes}
    path_cache: dict[str, str] = {}

    def get_path(code):
        if code in path_cache:
            return path_cache[code]
        node = by_code.get(code)
        if not node:
            path_cache[code] = code
            return code
        if node["parent"] is None:
            path_cache[code] = node["name"]
        else:
            parent_path = get_path(node["parent"])
            path_cache[code] = f"{parent_path} > {node['name']}"
        return path_cache[code]

    for n in nodes:
        get_path(n["code"])
    return path_cache


def write_excel(nodes: list[dict], output_path: str):
    print("Writing Excel file …")
    path_map = build_path_map(nodes)

    wb = Workbook()
    ws = wb.active
    ws.title = "АТХ Классификация"

    # Header row
    ws.append(["Код АТХ", "Наименование", "Уровень", "Полный путь"])
    for col in range(1, 5):
        cell = ws.cell(1, col)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for node in nodes:
        lvl    = node["level"]
        style  = LEVEL_STYLES[lvl]
        indent = "    " * style["indent"]

        ws.append([
            node["code"],
            indent + node["name"],
            lvl,
            path_map.get(node["code"], ""),
        ])
        ri   = ws.max_row
        fill = PatternFill("solid", fgColor=style["fill"])

        for col in range(1, 5):
            cell = ws.cell(ri, col)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font      = Font(
                bold=style["bold"],
                color=style["font_color"],
                size=style["size"],
                name="Courier New" if col == 1 else "Calibri",
            )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 90
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Saved: {output_path}  ({len(nodes)} rows)")


def main():
    print("=== vidal.ru ATC scraper ===\n")
    nodes = scrape_all()
    if not nodes:
        print("No data collected.")
        return
    write_excel(nodes, OUTPUT)
    print("\nDone!")


if __name__ == "__main__":
    main()
