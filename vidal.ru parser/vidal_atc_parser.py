import requests
import time
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
API_KEY   = "1234"
BASE_URL  = "https://www.vidal.ru/api/rest/v1"
OUTPUT    = "vidal_atc.xlsx"
DELAY     = 0.35   # seconds between requests (~3 rps, within the 3 rps limit)
PAGE_SIZE = 500
# ──────────────────────────────────────────────────────────────────────────────

HEADERS = {"X-Token": API_KEY}

# ATC level colours (light shades for readability)
LEVEL_STYLES = {
    1: {"fill": "4472C4", "font_color": "FFFFFF", "bold": True,  "indent": 0},
    2: {"fill": "D9E1F2", "font_color": "000000", "bold": True,  "indent": 1},
    3: {"fill": "EBF0FA", "font_color": "000000", "bold": False, "indent": 2},
    4: {"fill": "F5F8FE", "font_color": "000000", "bold": False, "indent": 3},
    5: {"fill": "FFFFFF", "font_color": "000000", "bold": False, "indent": 4},
}


def atc_level(code: str) -> int:
    """Return ATC level based on code length: 1→1, 3→2, 4→3, 5→4, 7→5."""
    length_to_level = {1: 1, 3: 2, 4: 3, 5: 4, 7: 5}
    return length_to_level.get(len(code), 0)


def api_get(path: str, params: dict = None) -> dict:
    """GET request with rate-limit delay and basic error handling."""
    time.sleep(DELAY)
    url = f"{BASE_URL}{path}"
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, params=params, timeout=30)
            if r.status_code == 429:
                print("  Rate limited, sleeping 5 s …")
                time.sleep(5)
                continue
            r.raise_for_status()
            data = r.json()
            if not data.get("success", True):
                print(f"  API error: {data.get('errorMsg')} — {data.get('errors')}")
            return data
        except requests.RequestException as e:
            print(f"  Request error ({attempt+1}/3): {e}")
            time.sleep(2)
    return {}


def fetch_all_pages(path: str, extra_params: dict = None) -> list:
    """Fetch every page from a paginated endpoint and return all items."""
    items = []
    page = 1
    while True:
        params = {"page": page, "limit": PAGE_SIZE}
        if extra_params:
            params.update(extra_params)
        data = api_get(path, params)
        batch = data.get("data", [])
        if not batch:
            break
        items.extend(batch)
        total = data.get("total", len(items))
        print(f"    page {page}: got {len(batch)} items (total so far {len(items)}/{total})")
        if len(items) >= total:
            break
        page += 1
    return items


def collect_tree() -> list:
    """
    Walk the ATC tree level by level and return a flat list of nodes
    sorted in tree order (parent always before its children).
    Each node: {"code": str, "name": str, "level": int, "parent": str|None}
    """
    print("Fetching top-level ATC codes (level 1) …")
    top = fetch_all_pages("/atc/list")

    # The API may return all levels at once or only the top level.
    # We index everything we receive and also fetch children explicitly.
    all_nodes: dict[str, dict] = {}

    def add(item, parent_code=None):
        code = item.get("code") or item.get("id") or ""
        name = item.get("name") or item.get("title") or ""
        if not code:
            return
        level = atc_level(code)
        if level == 0 or level > 5:
            return
        all_nodes[code] = {"code": code, "name": name, "level": level, "parent": parent_code}

    for item in top:
        add(item)

    # Determine which codes we need to expand (levels 1–4 can have children)
    def expand_level(codes_at_level: list[str]):
        for code in codes_at_level:
            level = atc_level(code)
            if level >= 5:
                continue
            print(f"  Fetching children of {code} …")
            children = fetch_all_pages("/atc/list", {"filter[parent]": code})
            if not children:
                # Try alternative param names used by some vidal API versions
                children = fetch_all_pages("/atc/list", {"filter[atcCode]": code})
            for child in children:
                add(child, parent_code=code)

    # Expand level by level so we don't miss any nodes
    for lvl in range(1, 5):
        codes_this_level = [c for c, n in all_nodes.items() if n["level"] == lvl]
        if not codes_this_level:
            break
        expand_level(codes_this_level)

    # Sort into tree order: sort by code (ATC codes sort lexicographically in tree order)
    ordered = sorted(all_nodes.values(), key=lambda n: n["code"])
    return ordered


def write_excel(nodes: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "ATC Classification"

    # Header row
    headers = ["ATC Code", "Name", "Level", "Full Path"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Build full-path lookup
    path_map: dict[str, str] = {}
    for node in nodes:
        parent_path = path_map.get(node["parent"], "") if node["parent"] else ""
        path_map[node["code"]] = f"{parent_path} > {node['name']}" if parent_path else node["name"]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for node in nodes:
        style = LEVEL_STYLES[node["level"]]
        indent = "    " * style["indent"]   # visual indent in the Name cell

        row = [
            node["code"],
            indent + node["name"],
            node["level"],
            path_map.get(node["code"], ""),
        ]
        ws.append(row)
        row_idx = ws.max_row

        fill = PatternFill("solid", fgColor=style["fill"])
        font_code = Font(bold=style["bold"], color=style["font_color"],
                         name="Courier New", size=10)
        font_name = Font(bold=style["bold"], color=style["font_color"], size=10)

        ws.cell(row_idx, 1).font = font_code
        ws.cell(row_idx, 2).font = font_name
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 80
    ws.row_dimensions[1].height = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\nSaved {len(nodes)} ATC entries to {output_path}")


def main():
    if API_KEY == "YOUR_API_KEY_HERE":
        print("ERROR: Please set your API key in the API_KEY variable at the top of the script.")
        sys.exit(1)

    print("=== vidal.ru ATC parser ===\n")
    nodes = collect_tree()
    print(f"\nTotal nodes collected: {len(nodes)}")

    if not nodes:
        print("No data collected. Check your API key and network connection.")
        sys.exit(1)

    print("Writing Excel file …")
    write_excel(nodes, OUTPUT)


if __name__ == "__main__":
    main()
